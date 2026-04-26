"""
Engine — Main orchestrator implementing Engine Protocol v2.0.

Coordinates all 6 layers, parsers, planner, and verifier to autonomously
complete Excel worksheet assignments.

Protocol v2.0:
  1. SCAN  — Read instructions, identify ALL tasks (2 min max)
  2. GROUP — Cluster tasks into logical sections (by sheet/dependency)
  3. For each section:
     a. Quick-plan: What needs doing, what order (30 sec)
     b. Verify plan is sound (no conflicts/missing prereqs)
     c. Execute aggressively
     d. Verify completion
     e. Move to next section
"""

from __future__ import annotations

import logging
import time
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from copy import deepcopy
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Callable, Optional

from excel_engine.config import EngineConfig, TaskType, Layer
from excel_engine.recalc import recalculate, RecalcResult
from excel_engine.recovery import ErrorClassifier, RecoveryStrategy, TaskError, CircuitBreaker
from excel_engine.layers.openpyxl_layer import OpenpyxlLayer
from excel_engine.layers.xlwings_layer import XlwingsLayer
from excel_engine.layers.applescript_layer import AppleScriptLayer
from excel_engine.layers.system_events import SystemEventsLayer
from excel_engine.layers.vba_layer import VBALayer
from excel_engine.layers.pyautogui_layer import PyAutoGUILayer
from excel_engine.parsers.instruction_parser import InstructionParser
from excel_engine.parsers.task_extractor import Task, TaskExtractor
from excel_engine.planner.task_planner import TaskPlanner, ExecutionPlan, Section
from excel_engine.verifier.workbook_verifier import WorkbookVerifier, SectionVerification
from excel_engine.utils.path_handler import PathHandler
from excel_engine.utils.mac_utils import MacUtils

logger = logging.getLogger(__name__)


@dataclass
class EngineResult:
    """Result of running the engine on an assignment."""
    success: bool
    workbook_path: Path
    sections_completed: int
    sections_total: int
    tasks_completed: int
    tasks_total: int
    verifications: list[SectionVerification] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)
    task_errors: list[TaskError] = field(default_factory=list)
    failed_tasks: list[str] = field(default_factory=list)
    formula_errors: Optional[RecalcResult] = None
    elapsed_seconds: float = 0.0

    def summary(self) -> str:
        status = "✓ SUCCESS" if self.success else "✗ FAILED"
        lines = [
            f"{status}: {self.workbook_path.name}",
            f"  Sections: {self.sections_completed}/{self.sections_total}",
            f"  Tasks: {self.tasks_completed}/{self.tasks_total}",
            f"  Time: {self.elapsed_seconds:.1f}s",
        ]
        if self.errors:
            lines.append(f"  Errors ({len(self.errors)}):")
            for err in self.errors[:5]:
                lines.append(f"    - {err}")
        if self.failed_tasks:
            lines.append(f"  Failed tasks ({len(self.failed_tasks)}): {', '.join(self.failed_tasks[:10])}")
        if self.task_errors:
            transient = sum(1 for e in self.task_errors if e.error_type == "transient")
            permanent = sum(1 for e in self.task_errors if e.error_type == "permanent")
            incompatible = sum(1 for e in self.task_errors if e.error_type == "layer_incompatible")
            lines.append(
                f"  Task errors: {len(self.task_errors)} "
                f"(transient={transient}, permanent={permanent}, layer_incompatible={incompatible})"
            )
        if self.formula_errors is not None:
            fe = self.formula_errors
            if fe.skipped:
                lines.append(f"  Formula recalc: skipped ({fe.warning})")
            elif fe.total_errors > 0:
                lines.append(
                    f"  Formula recalc: {fe.total_errors} errors in "
                    f"{fe.total_formulas} formulas"
                )
                for err_type, info in fe.error_summary.items():
                    lines.append(f"    {err_type}: {info['count']}")
            else:
                lines.append(f"  Formula recalc: OK ({fe.total_formulas} formulas)")
        return "\n".join(lines)


class ExcelEngine:
    """
    Main orchestrator — coordinates all layers to complete Excel assignments.

    Usage:
        engine = ExcelEngine()
        result = engine.run(
            workbook=Path("assignment.xlsx"),
            instructions=Path("instructions.docx"),
        )
        print(result.summary())
    """

    def __init__(self, config: Optional[EngineConfig] = None) -> None:
        self.config = config or EngineConfig()

        # Initialize layers
        self._openpyxl = OpenpyxlLayer()
        self._xlwings = XlwingsLayer()
        self._applescript = AppleScriptLayer(timeout=self.config.applescript_timeout)
        self._system_events = SystemEventsLayer(delay=self.config.ui_delay_between_actions)
        self._vba = VBALayer(
            execution_timeout=self.config.vba_execution_timeout,
            split_threshold=self.config.vba_split_threshold,
        )
        # Note: Layer.PYAUTOGUI is available as a last-resort fallback but no TaskTypes
        # map to it by default. It can be added to layer_order for pixel-level automation.
        self._pyautogui = PyAutoGUILayer(retina=self.config.retina_display)

        # Initialize components
        self._parser = InstructionParser()
        self._extractor = TaskExtractor()
        self._planner = TaskPlanner(config=self.config)
        self._verifier = WorkbookVerifier()
        self.path_handler = PathHandler(desktop_path=self.config.desktop_path)

        # Circuit breaker for layer cascade
        self.circuit_breaker = CircuitBreaker(
            failure_threshold=self.config.circuit_breaker_threshold,
            reset_timeout=self.config.circuit_breaker_reset_seconds,
        )

        # Layer lookup
        self._layers = {
            Layer.OPENPYXL: self._openpyxl,
            Layer.XLWINGS: self._xlwings,
            Layer.APPLESCRIPT: self._applescript,
            Layer.SYSTEM_EVENTS: self._system_events,
            Layer.VBA: self._vba,
            Layer.PYAUTOGUI: self._pyautogui,
        }

    # ── Public read-only properties ──

    @property
    def openpyxl(self) -> OpenpyxlLayer:
        return self._openpyxl

    @property
    def xlwings(self) -> XlwingsLayer:
        return self._xlwings

    @property
    def applescript(self) -> AppleScriptLayer:
        return self._applescript

    @property
    def system_events(self) -> SystemEventsLayer:
        return self._system_events

    @property
    def vba(self) -> VBALayer:
        return self._vba

    @property
    def pyautogui(self) -> PyAutoGUILayer:
        return self._pyautogui

    @property
    def parser(self) -> InstructionParser:
        return self._parser

    @property
    def extractor(self) -> TaskExtractor:
        return self._extractor

    @property
    def planner(self) -> TaskPlanner:
        return self._planner

    @property
    def verifier(self) -> WorkbookVerifier:
        return self._verifier

    # ── Main Entry Point ──

    def execute(self, plan: ExecutionPlan, workbook: Path,
                progress_callback: Optional[Callable[[dict], None]] = None) -> EngineResult:
        """Execute a pre-built plan against a workbook."""
        start_time = time.time()
        workbook = Path(workbook).resolve()

        result = EngineResult(
            success=False,
            workbook_path=workbook,
            sections_completed=0,
            sections_total=plan.section_count,
            tasks_completed=0,
            tasks_total=plan.total_tasks,
        )

        try:
            logger.info("=" * 60)
            logger.info("EXECUTE — Running %d sections", plan.section_count)
            logger.info("=" * 60)

            for section in plan.sections:
                section_ok = self._execute_section(
                    section, workbook, result,
                    progress_callback=progress_callback,
                )
                if section_ok:
                    result.sections_completed += 1

            result.success = (
                result.sections_completed == result.sections_total
                and result.tasks_completed == result.tasks_total
            )

            # ── Phase 4: RECALCULATE formulas (optional) ──
            if self.config.recalculate_formulas and result.success:
                logger.info("=" * 60)
                logger.info("PHASE 4: RECALCULATE — LibreOffice formula recalculation")
                logger.info("=" * 60)
                try:
                    recalc_result = recalculate(
                        workbook, timeout=self.config.recalc_timeout
                    )
                    result.formula_errors = recalc_result
                    if recalc_result.skipped:
                        logger.info("Recalculation skipped: %s", recalc_result.warning)
                    elif recalc_result.total_errors > 0:
                        logger.warning(
                            "Recalculation found %d formula errors",
                            recalc_result.total_errors,
                        )
                    else:
                        logger.info(
                            "Recalculation OK — %d formulas, 0 errors",
                            recalc_result.total_formulas,
                        )
                except Exception as e:
                    logger.warning("Formula recalculation failed: %s", e)
                    result.formula_errors = RecalcResult(
                        success=False, warning=str(e)
                    )

        except Exception as e:
            logger.exception("Execution failed: %s", e)
            result.errors.append(str(e))

        finally:
            result.elapsed_seconds = time.time() - start_time
            self._cleanup()

        logger.info("\n%s", result.summary())
        return result

    def run(
        self,
        workbook: Path,
        instructions: Optional[Path] = None,
        instruction_text: Optional[str] = None,
        tasks: Optional[list[Task]] = None,
        progress_callback: Optional[Callable[[dict], None]] = None,
    ) -> EngineResult:
        """
        Run the engine on a workbook with instructions.

        Provide EITHER:
          - instructions: path to instruction file (.docx, .rtfd, .pdf, .txt)
          - instruction_text: raw instruction text
          - tasks: pre-extracted Task list
        """
        workbook = Path(workbook).resolve()

        if tasks:
            tasks = deepcopy(tasks)

        try:
            # ── Phase 1: SCAN ──
            logger.info("=" * 60)
            logger.info("PHASE 1: SCAN — Parsing instructions")
            logger.info("=" * 60)

            if tasks is None:
                if instructions:
                    text = self._parser.parse(instructions)
                elif instruction_text:
                    text = instruction_text
                else:
                    raise ValueError("Provide instructions, instruction_text, or tasks")

                tasks = self._extractor.extract(text)

            logger.info("Extracted %d tasks", len(tasks))

            # ── Phase 2: GROUP ──
            logger.info("=" * 60)
            logger.info("PHASE 2: GROUP — Planning execution")
            logger.info("=" * 60)

            plan = self._planner.plan(tasks)
            logger.info("\n%s", plan.summary())

            # ── Phase 3: EXECUTE ──
            return self.execute(plan, workbook, progress_callback=progress_callback)

        except Exception as e:
            logger.exception("Engine failed during scan/plan: %s", e)
            result = EngineResult(
                success=False,
                workbook_path=workbook,
                sections_completed=0,
                sections_total=0,
                tasks_completed=0,
                tasks_total=0,
                errors=[str(e)],
            )
            return result

    # ── Section Execution ──

    def _execute_section(
        self, section: Section, workbook: Path, result: EngineResult,
        progress_callback: Optional[Callable[[dict], None]] = None,
    ) -> bool:
        """Execute all tasks in a section.

        When ``config.parallel_execution`` is enabled, tasks targeting
        *different* sheets with no cross-references run concurrently via a
        thread pool. Tasks on the *same* sheet always execute serially to
        preserve ordering semantics.
        """
        logger.info("-" * 40)
        logger.info("Section %s: %s (%d tasks)", section.id, section.name, section.task_count)
        logger.info("-" * 40)

        total_tasks = section.task_count
        section_ok = True

        if self.config.parallel_execution:
            section_ok = self._execute_section_parallel(
                section, workbook, result, total_tasks, progress_callback,
            )
        else:
            for idx, task in enumerate(section.tasks):
                if progress_callback:
                    progress_callback({
                        "phase": "executing",
                        "status": "executing",
                        "section": section.id,
                        "section_name": section.name,
                        "task": task.id,
                        "task_type": task.task_type.value,
                        "index": idx,
                        "total": total_tasks,
                    })

                task_ok, task_errors = self._execute_task(task, workbook)
                result.task_errors.extend(task_errors)

                if task_ok:
                    task.completed = True
                    result.tasks_completed += 1
                    if progress_callback:
                        progress_callback({
                            "phase": "completed",
                            "status": "completed",
                            "section": section.id,
                            "section_name": section.name,
                            "task": task.id,
                            "task_type": task.task_type.value,
                            "index": idx,
                            "total": total_tasks,
                            "success": True,
                            "passed": True,
                        })
                else:
                    section_ok = False
                    result.failed_tasks.append(task.id)
                    result.errors.append(
                        f"Task {task.id} ({task.task_type.value}) failed"
                    )
                    if progress_callback:
                        progress_callback({
                            "phase": "completed",
                            "status": "completed",
                            "section": section.id,
                            "section_name": section.name,
                            "task": task.id,
                            "task_type": task.task_type.value,
                            "index": idx,
                            "total": total_tasks,
                            "success": False,
                            "passed": False,
                        })

        # Verify section if configured
        if self.config.verify_after_each_section:
            try:
                self._save_workbook(workbook)
                self._verifier.load(workbook)
                verification = self._verifier.verify_section(section.id, section.tasks)
                result.verifications.append(verification)
                self._verifier.close()

                if not verification.all_passed:
                    logger.warning(
                        "Verification issues in section %s: %d failed",
                        section.id, verification.fail_count,
                    )
            except Exception as e:
                logger.warning("Verification error for section %s: %s", section.id, e)

        section.completed = section_ok
        return section_ok

    def _execute_section_parallel(
        self,
        section: Section,
        workbook: Path,
        result: EngineResult,
        total_tasks: int,
        progress_callback: Optional[Callable[[dict], None]] = None,
    ) -> bool:
        """Run tasks in parallel grouped by sheet.

        Tasks on the same sheet run serially (order matters). Tasks on
        different sheets with no cross-references run concurrently.
        """
        # Group tasks by target sheet
        sheet_groups: dict[str, list[Task]] = defaultdict(list)
        for task in section.tasks:
            key = task.sheet or "__default__"
            sheet_groups[key].append(task)

        section_ok = True
        completed_idx = 0

        def _run_sheet_group(sheet_key: str, tasks: list[Task]):
            nonlocal section_ok, completed_idx
            group_ok = True
            for task in tasks:
                if progress_callback:
                    progress_callback({
                        "phase": "executing",
                        "status": "executing",
                        "section": section.id,
                        "section_name": section.name,
                        "task": task.id,
                        "task_type": task.task_type.value,
                        "total": total_tasks,
                    })

                task_ok, task_errors = self._execute_task(task, workbook)
                result.task_errors.extend(task_errors)

                if task_ok:
                    task.completed = True
                    result.tasks_completed += 1
                    if progress_callback:
                        progress_callback({
                            "phase": "completed",
                            "status": "completed",
                            "section": section.id,
                            "section_name": section.name,
                            "task": task.id,
                            "task_type": task.task_type.value,
                            "total": total_tasks,
                            "success": True,
                            "passed": True,
                        })
                else:
                    group_ok = False
                    result.failed_tasks.append(task.id)
                    result.errors.append(
                        f"Task {task.id} ({task.task_type.value}) failed"
                    )
                    if progress_callback:
                        progress_callback({
                            "phase": "completed",
                            "status": "completed",
                            "section": section.id,
                            "section_name": section.name,
                            "task": task.id,
                            "task_type": task.task_type.value,
                            "total": total_tasks,
                            "success": False,
                            "passed": False,
                        })
            return group_ok

        if len(sheet_groups) == 1:
            # Single sheet — just run serially, no thread overhead
            key, tasks = next(iter(sheet_groups.items()))
            section_ok = _run_sheet_group(key, tasks)
        else:
            with ThreadPoolExecutor(max_workers=self.config.max_workers) as pool:
                futures = {
                    pool.submit(_run_sheet_group, key, tasks): key
                    for key, tasks in sheet_groups.items()
                }
                for future in as_completed(futures):
                    if not future.result():
                        section_ok = False

        return section_ok

    # ── Task Execution ──

    def _execute_task(self, task: Task, workbook: Path) -> tuple[bool, list[TaskError]]:
        """Execute a single task with retry logic and layer cascade.

        Returns (success, list_of_errors).
        """
        layers_to_try = self.config.get_layers_for_task(task.task_type)
        collected_errors: list[TaskError] = []

        if not layers_to_try:
            logger.warning("No layer can handle task type: %s", task.task_type.value)
            return False, collected_errors

        strategy = RecoveryStrategy(
            max_retries=self.config.max_retries,
            base_delay=self.config.retry_delay,
        )

        for layer_enum in layers_to_try:
            layer_name = layer_enum.name

            # A5: Circuit breaker — skip layers that are consistently failing
            if self.circuit_breaker.is_open(layer_name):
                logger.info(
                    "  Skipping Layer %s for task %s (circuit breaker open)",
                    layer_name, task.id,
                )
                continue

            attempt = 0
            while True:
                try:
                    logger.info(
                        "  Task %s: %s via Layer %s (attempt %d)",
                        task.id, task.task_type.value, layer_name, attempt + 1,
                    )
                    self._dispatch_task(task, layer_enum, workbook)
                    self.circuit_breaker.record_success(layer_name)
                    return True, collected_errors
                except Exception as e:
                    error_type = ErrorClassifier.classify(e, layer_enum)
                    task_error = TaskError(
                        task_id=task.id,
                        task_type=task.task_type,
                        layer=layer_enum,
                        error_type=error_type,
                        message=str(e),
                        timestamp=time.time(),
                    )
                    collected_errors.append(task_error)
                    self.circuit_breaker.record_failure(layer_name)

                    if strategy.should_retry(error_type, attempt):
                        delay = strategy.get_delay(attempt)
                        logger.warning(
                            "  Layer %s transient error for %s (attempt %d/%d): %s — retrying in %.1fs",
                            layer_name, task.id, attempt + 1,
                            strategy.max_retries, e, delay,
                        )
                        time.sleep(delay)
                        attempt += 1
                    else:
                        logger.warning(
                            "  Layer %s %s error for %s: %s — escalating to next layer",
                            layer_name, error_type, task.id, e,
                        )
                        break  # move to next layer

        logger.error("All layers failed for task %s", task.id)
        return False, collected_errors

    def _dispatch_task(
        self, task: Task, layer: Layer, workbook: Path
    ) -> None:
        """Dispatch a task to a specific layer."""
        dispatch_map = {
            Layer.OPENPYXL: self._exec_openpyxl,
            Layer.XLWINGS: self._exec_xlwings,
            Layer.APPLESCRIPT: self._exec_applescript,
            Layer.SYSTEM_EVENTS: self._exec_system_events,
            Layer.VBA: self._exec_vba,
            Layer.PYAUTOGUI: self._exec_pyautogui,
        }

        handler = dispatch_map.get(layer)
        if handler:
            handler(task, workbook)
        else:
            raise ValueError(f"Unknown layer: {layer}")

    # ── Layer Dispatchers ──

    def _exec_openpyxl(self, task: Task, workbook: Path) -> None:
        """Execute a task via Layer 1 (openpyxl)."""
        if not self._openpyxl.wb:
            self._openpyxl.open(workbook)

        tt = task.task_type

        if tt == TaskType.FORMULA:
            if task.cell and task.formula:
                self._openpyxl.set_formula(task.cell, task.formula, sheet=task.sheet)

        elif tt == TaskType.CELL_VALUE:
            if task.cell and task.value is not None:
                self._openpyxl.set_value(task.cell, task.value, sheet=task.sheet)

        elif tt == TaskType.TABLE_CREATE:
            name = task.params.get("name", "Table1")
            ref = task.range or task.params.get("ref", "A1:A1")
            style = task.style or task.params.get("style", "TableStyleMedium5")
            self._openpyxl.create_table(name, ref, style=style, sheet=task.sheet)

        elif tt == TaskType.NUMBER_FORMAT:
            ref = task.range or task.cell
            fmt = task.params.get("format", "$#,##0.00")
            if ref:
                self._openpyxl.set_number_format(ref, fmt, sheet=task.sheet)

        elif tt == TaskType.FONT:
            ref = task.range or task.cell
            if ref:
                self._openpyxl.set_font(
                    ref, sheet=task.sheet,
                    bold=task.params.get("bold", False),
                    italic=task.params.get("italic", False),
                    size=task.params.get("size", 11),
                    color=task.params.get("color", "000000"),
                    name=task.params.get("font_name", "Calibri"),
                )

        elif tt == TaskType.FILL:
            ref = task.range or task.cell
            color = task.params.get("color", "FFFF00")
            if ref:
                self._openpyxl.set_fill(ref, color, sheet=task.sheet)

        elif tt == TaskType.ALIGNMENT:
            ref = task.range or task.cell
            if ref:
                self._openpyxl.set_alignment(
                    ref, sheet=task.sheet,
                    horizontal=task.params.get("horizontal", "general"),
                    vertical=task.params.get("vertical", "bottom"),
                    wrap_text=task.params.get("wrap_text", False),
                )

        elif tt == TaskType.BORDER:
            ref = task.range or task.cell
            if ref:
                self._openpyxl.set_border(
                    ref, sheet=task.sheet,
                    style=task.params.get("style", "thin"),
                )

        elif tt == TaskType.COLUMN_WIDTH:
            col = task.params.get("column")
            if not col and task.cell:
                # Extract column letter from cell reference (e.g. "B1" → "B")
                import re
                m = re.match(r"([A-Za-z]+)", task.cell)
                col = m.group(1).upper() if m else "A"
            col = col or "A"
            width = task.params.get("width", 12)
            self._openpyxl.set_column_width(col, width, sheet=task.sheet)

        elif tt == TaskType.CONDITIONAL_FORMAT:
            ref = task.range or "A1:A1"
            operator = task.params.get("operator", "greaterThan")
            formula = task.params.get("formula", ["0"])
            self._openpyxl.add_conditional_format_cell_is(
                ref, operator, formula, sheet=task.sheet,
            )

        elif tt == TaskType.CHART_BAR:
            self._openpyxl.add_bar_chart(
                sheet=task.sheet,
                title=task.params.get("title", ""),
                data_range=task.params.get("data_range", task.range or "B1:B10"),
                cats_range=task.params.get("cats_range"),
                anchor=task.params.get("anchor", "E2"),
            )

        elif tt == TaskType.CHART_LINE:
            self._openpyxl.add_line_chart(
                sheet=task.sheet,
                title=task.params.get("title", ""),
                data_range=task.params.get("data_range", task.range or "B1:B10"),
                cats_range=task.params.get("cats_range"),
                anchor=task.params.get("anchor", "E2"),
            )

        elif tt == TaskType.CHART_PIE:
            self._openpyxl.add_pie_chart(
                sheet=task.sheet,
                title=task.params.get("title", ""),
                data_range=task.params.get("data_range", task.range or "B1:B10"),
                cats_range=task.params.get("cats_range"),
                anchor=task.params.get("anchor", "E2"),
            )

        elif tt == TaskType.CHART_SCATTER:
            self._openpyxl.add_scatter_chart(
                sheet=task.sheet,
                title=task.params.get("title", ""),
                x_range=task.params.get("x_range", task.range or "A1:A10"),
                y_range=task.params.get("y_range", "B1:B10"),
                anchor=task.params.get("anchor", "E2"),
                scatter_style=task.params.get("scatter_style", "line"),
            )

        elif tt == TaskType.CHART_AREA:
            self._openpyxl.add_area_chart(
                sheet=task.sheet,
                title=task.params.get("title", ""),
                data_range=task.params.get("data_range", task.range or "B1:B10"),
                cats_range=task.params.get("cats_range"),
                anchor=task.params.get("anchor", "E2"),
                grouping=task.params.get("grouping", "standard"),
            )

        elif tt == TaskType.CHART_COMBO:
            self._openpyxl.add_combo_chart(
                sheet=task.sheet,
                title=task.params.get("title", ""),
                bar_data_range=task.params.get("bar_data_range", task.range or "B1:B10"),
                line_data_range=task.params.get("line_data_range", "C1:C10"),
                cats_range=task.params.get("cats_range"),
                anchor=task.params.get("anchor", "E2"),
                secondary_axis=task.params.get("secondary_axis", True),
            )

        elif tt == TaskType.NAMED_RANGE:
            name = task.params.get("name", "NamedRange1")
            sheet = task.sheet or self._openpyxl.wb.active.title
            range_str = task.range or "$A$1"
            self._openpyxl.create_named_range(name, sheet, range_str)

        elif tt == TaskType.DATA_VALIDATION:
            ref = task.range or task.cell or "A1"
            formula1 = task.params.get("formula1")
            # Convert a "values" list to the comma-separated formula1 format
            if not formula1 and "values" in task.params:
                formula1 = '"' + ",".join(str(v) for v in task.params["values"]) + '"'
            self._openpyxl.add_data_validation(
                ref, sheet=task.sheet,
                validation_type=task.params.get("type", "list"),
                formula1=formula1,
            )

        elif tt == TaskType.FREEZE_PANES:
            cell = task.cell or "A2"
            self._openpyxl.freeze_panes(cell, sheet=task.sheet)

        elif tt == TaskType.AUTOFILTER:
            ref = task.range or task.params.get("ref", "A1:A1")
            self._openpyxl.set_autofilter(ref, sheet=task.sheet)

        elif tt == TaskType.SHEET_CREATE:
            name = task.params.get("name", task.sheet or "NewSheet")
            self._openpyxl.create_sheet(name)

        elif tt == TaskType.SHEET_RENAME:
            old = task.params.get("old_name", "Sheet1")
            new = task.params.get("new_name", task.sheet or "Renamed")
            self._openpyxl.rename_sheet(old, new)

        elif tt == TaskType.MERGE_CELLS:
            ref = task.range
            if ref:
                self._openpyxl.merge_cells(ref, sheet=task.sheet)

        elif tt == TaskType.PRINT_SETTINGS:
            if task.params.get("landscape"):
                self._openpyxl.set_page_orientation(True, sheet=task.sheet)
            if task.params.get("print_area"):
                self._openpyxl.set_print_area(task.params["print_area"], sheet=task.sheet)

        # Formula-based types — delegate to existing set_formula
        elif tt in (TaskType.TEXT_FUNCTION, TaskType.LOOKUP_FUNCTION,
                    TaskType.FILTER_FUNCTION, TaskType.SORT_FUNCTION,
                    TaskType.UNIQUE_FUNCTION, TaskType.THREE_D_REFERENCE,
                    TaskType.EXTERNAL_REFERENCE):
            if task.cell and task.formula:
                self._openpyxl.set_formula(task.cell, task.formula, sheet=task.sheet)
            elif task.cell and task.value:
                self._openpyxl.set_value(task.cell, task.value, sheet=task.sheet)

        # Table total row
        elif tt == TaskType.TABLE_TOTAL_ROW:
            # openpyxl tables support totalsRowShown
            ws = self._openpyxl._ws(task.sheet)
            for table in ws.tables.values():
                if not task.params.get("name") or table.name == task.params.get("name"):
                    table.totalsRowShown = True
                    break

        # Row height
        elif tt == TaskType.ROW_HEIGHT:
            ws = self._openpyxl._ws(task.sheet)
            row_num = int(task.cell[1:]) if task.cell else task.params.get("row", 1)
            height = task.params.get("size", task.params.get("height", 15))
            ws.row_dimensions[row_num].height = height

        # Tab color
        elif tt == TaskType.TAB_COLOR:
            ws = self._openpyxl._ws(task.sheet)
            color = task.params.get("color", "FF0000")
            ws.sheet_properties.tabColor = color

        # Page break
        elif tt == TaskType.PAGE_BREAK:
            from openpyxl.worksheet.pagebreak import Break
            ws = self._openpyxl._ws(task.sheet)
            row_num = int(task.cell[1:]) if task.cell else task.params.get("row", 1)
            ws.row_breaks.append(Break(id=row_num))

        # Hyperlink
        elif tt == TaskType.HYPERLINK:
            ws = self._openpyxl._ws(task.sheet)
            cell = ws[task.cell]
            url = task.params.get("url", task.value or "")
            display = task.params.get("display", url)
            cell.hyperlink = url
            cell.value = display
            cell.style = "Hyperlink"

        # Sheet move (reorder)
        elif tt == TaskType.SHEET_MOVE:
            if task.sheet:
                idx = task.params.get("position", 0)
                self._openpyxl.wb.move_sheet(task.sheet, offset=idx)

        # Sheet copy
        elif tt == TaskType.SHEET_COPY:
            source = task.sheet or task.params.get("source")
            new_name = task.params.get("new_name", f"{source} Copy")
            ws = self._openpyxl.wb.copy_worksheet(self._openpyxl.wb[source])
            ws.title = new_name

        # Generic FORMATTING — apply whatever params specify
        elif tt == TaskType.FORMATTING:
            if task.cell or task.range:
                target = task.cell or task.range
                ws = self._openpyxl._ws(task.sheet)
                # Delegate to font/fill/border/alignment based on params
                from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
                cells = ws[target] if ':' in str(target) else [[ws[target]]]
                for row in cells:
                    for c in (row if hasattr(row, '__iter__') else [row]):
                        if task.params.get("bold") or task.params.get("font_size") or task.params.get("font_color"):
                            c.font = Font(
                                bold=task.params.get("bold", False),
                                size=task.params.get("font_size"),
                                color=task.params.get("font_color"),
                            )

        # Table style
        elif tt == TaskType.TABLE_STYLE:
            ws = self._openpyxl._ws(task.sheet)
            style_name = task.style or task.params.get("style", "TableStyleMedium5")
            table_name = task.params.get("table_name")
            for table in ws.tables.values():
                if not table_name or table.name == table_name:
                    table.tableStyleInfo.name = style_name
                    break

        else:
            raise NotImplementedError(
                f"openpyxl handler not implemented for {tt.value}"
            )

    def _exec_xlwings(self, task: Task, workbook: Path) -> None:
        """Execute a task via Layer 2 (xlwings)."""
        safe_path = self.path_handler.safe_copy_for_xlwings(workbook)

        if not self._xlwings._wb:
            MacUtils.launch_excel()
            MacUtils.open_workbook_in_excel(safe_path)
            MacUtils.wait_for_excel_ready()
            self._xlwings.connect(safe_path)

        tt = task.task_type

        if tt in (
            TaskType.FORMULA, TaskType.CALCULATED_COLUMN,
            TaskType.TEXT_FUNCTION, TaskType.LOOKUP_FUNCTION,
            TaskType.FILTER_FUNCTION, TaskType.SORT_FUNCTION,
            TaskType.UNIQUE_FUNCTION, TaskType.THREE_D_REFERENCE,
            TaskType.EXTERNAL_REFERENCE,
        ):
            if task.cell and task.formula:
                self._xlwings.set_formula(task.cell, task.formula, sheet=task.sheet)

        elif tt == TaskType.CELL_VALUE:
            if task.cell and task.value:
                self._xlwings.write_cell(task.cell, task.value, sheet=task.sheet)

        elif tt == TaskType.SUBTOTAL:
            range_str = task.range or task.params.get("range", "A1:A1")
            self._xlwings.add_subtotal(
                range_str,
                group_by=task.params.get("group_by", 1),
                function=task.params.get("function", -4157),
                total_list=task.params.get("total_list", [2]),
                sheet=task.sheet,
            )

        elif tt == TaskType.SPLIT_PANES:
            row = task.params.get("row", 2)
            col = task.params.get("col", 1)
            self._xlwings.split_panes(row, col, sheet=task.sheet)

        elif tt == TaskType.ADVANCED_FILTER:
            self._xlwings.advanced_filter(
                list_range=task.params.get("list_range", task.range or "A1:A1"),
                criteria_range=task.params.get("criteria_range", ""),
                copy_to_range=task.params.get("copy_to_range"),
                unique=task.params.get("unique", False),
                sheet=task.sheet,
            )

        elif tt == TaskType.NUMBER_FORMAT:
            ref = task.range or task.cell
            if ref:
                self._xlwings.set_number_format(
                    ref, task.params.get("format", "$#,##0.00"), sheet=task.sheet,
                )

        elif tt == TaskType.COLUMN_WIDTH:
            col = task.params.get("column", "A")
            width = task.params.get("width", 12)
            self._xlwings.set_column_width(col, width, sheet=task.sheet)

        elif tt == TaskType.TABLE_STYLE:
            table_name = task.params.get("table_name", "Table1")
            style = task.style or task.params.get("style", "TableStyleMedium5")
            sheet = task.sheet or "Sheet1"
            self._xlwings.set_table_style(table_name, style, sheet)

        elif tt == TaskType.SAVE:
            self._xlwings.save()

        elif tt in (
            TaskType.CHART_SCATTER, TaskType.CHART_AREA,
            TaskType.CHART_COMBO, TaskType.CHART_BAR,
            TaskType.CHART_LINE,
        ):
            _TYPE_MAP = {
                TaskType.CHART_SCATTER: "scatter",
                TaskType.CHART_AREA: "area",
                TaskType.CHART_COMBO: "combo",
                TaskType.CHART_BAR: "column",
                TaskType.CHART_LINE: "line",
            }
            secondary = None
            if tt == TaskType.CHART_COMBO:
                secondary = task.params.get("secondary_axis_series", [1])
            self._xlwings.add_chart(
                chart_type=_TYPE_MAP[tt],
                data_range=task.params.get("data_range", task.range or "A1:B10"),
                sheet=task.sheet,
                title=task.params.get("title", ""),
                anchor=task.params.get("anchor", "E2"),
                secondary_axis_series=secondary,
            )

        elif tt == TaskType.SPARKLINE:
            self._xlwings.add_sparkline(
                data_range=task.params.get("data_range", task.range or "B2:M2"),
                location_range=task.params.get("location_range", task.cell or "N2"),
                sparkline_type=task.params.get("sparkline_type", "line"),
                sheet=task.sheet,
            )

        elif tt == TaskType.NAMED_RANGE:
            name = task.params.get("name", task.value or "MyRange")
            refers_to = task.params.get("refers_to", task.range or "A1")
            self._xlwings.create_named_range(name, refers_to, sheet=task.sheet)

        elif tt == TaskType.HYPERLINK:
            url = task.params.get("url", task.value or "")
            display = task.params.get("display_text", url)
            self._xlwings.add_hyperlink(
                task.cell or "A1", url, display_text=display, sheet=task.sheet,
            )

        elif tt == TaskType.SHEET_CREATE:
            name = task.params.get("name", task.value or "NewSheet")
            self._xlwings.add_sheet(name)

        elif tt == TaskType.SHEET_RENAME:
            old = task.params.get("old_name", task.sheet or "Sheet1")
            new = task.params.get("new_name", task.value or "Renamed")
            self._xlwings.rename_sheet(old, new)

        elif tt == TaskType.SHEET_MOVE:
            name = task.sheet or task.params.get("name", "Sheet1")
            before = task.params.get("before")
            after = task.params.get("after")
            self._xlwings.move_sheet(name, before=before, after=after)

        elif tt == TaskType.SHEET_COPY:
            name = task.sheet or task.params.get("name", "Sheet1")
            new_name = task.params.get("new_name")
            self._xlwings.copy_sheet(name, new_name=new_name)

        elif tt == TaskType.GOAL_SEEK:
            self._xlwings.goal_seek(
                target_cell=task.cell or task.params.get("target_cell", "A1"),
                goal_value=task.params.get("goal_value", 0),
                changing_cell=task.params.get("changing_cell", "B1"),
                sheet=task.sheet,
            )

        elif tt == TaskType.SORT:
            range_str = task.range or task.params.get("range", "A1:A1")
            keys = task.params.get("keys", [])
            self._xlwings.sort_range(range_str, keys, sheet=task.sheet)

        elif tt == TaskType.TAB_COLOR:
            color = task.params.get("color", task.value or "#FF0000")
            sheet_name = task.sheet or "Sheet1"
            self._xlwings.set_tab_color(sheet_name, color)

        elif tt == TaskType.TABLE_TOTAL_ROW:
            table_name = task.params.get("table_name", "Table1")
            show = task.params.get("show", True)
            self._xlwings.set_table_total_row(table_name, show, sheet=task.sheet)

        elif tt == TaskType.FORMATTING:
            if task.cell or task.range:
                ref = task.range or task.cell
                self._xlwings.apply_formatting(ref, task.params, sheet=task.sheet)

        else:
            raise NotImplementedError(
                f"xlwings handler not implemented for {tt.value}"
            )

        # Save and copy back if we used a desktop copy
        self._xlwings.save()
        if safe_path != workbook:
            self.path_handler.copy_back_from_desktop(safe_path, workbook)

    def _exec_applescript(self, task: Task, workbook: Path) -> None:
        """Execute a task via Layer 3 (AppleScript)."""
        tt = task.task_type

        if tt in (
            TaskType.FORMULA, TaskType.CALCULATED_COLUMN,
            TaskType.TEXT_FUNCTION, TaskType.LOOKUP_FUNCTION,
            TaskType.FILTER_FUNCTION, TaskType.SORT_FUNCTION,
            TaskType.UNIQUE_FUNCTION, TaskType.THREE_D_REFERENCE,
            TaskType.EXTERNAL_REFERENCE,
        ):
            if task.cell and task.formula:
                self._applescript.set_cell_formula(
                    task.cell, task.formula, sheet=task.sheet,
                )

        elif tt == TaskType.CELL_VALUE:
            if task.cell:
                self._applescript.set_cell_value(
                    task.cell, str(task.value or ""), sheet=task.sheet,
                )

        elif tt == TaskType.SORT:
            range_str = task.range or task.params.get("range", "A1:A1")
            keys = task.params.get("keys", [])
            self._applescript.sort_range(range_str, keys, sheet=task.sheet)

        elif tt == TaskType.AUTOFILTER:
            range_str = task.range or task.params.get("range", "A1:A1")
            field_num = task.params.get("field", 1)
            criteria = task.params.get("criteria", "*")
            self._applescript.set_autofilter(
                range_str, field_num, criteria, sheet=task.sheet,
            )

        elif tt == TaskType.FREEZE_PANES:
            cell = task.cell or "A2"
            self._applescript.freeze_panes(cell, sheet=task.sheet)

        elif tt == TaskType.SAVE:
            self._applescript.save()

        elif tt == TaskType.SAVE_AS:
            path = task.params.get("path", str(self.config.desktop_path))
            filename = task.params.get("filename", "output")
            self._applescript.save_as_xlsx(path, filename)

        else:
            raise NotImplementedError(
                f"AppleScript handler not implemented for {tt.value}"
            )

    def _exec_system_events(self, task: Task, workbook: Path) -> None:
        """Execute a task via Layer 4 (System Events)."""
        tt = task.task_type

        if tt == TaskType.SLICER:
            fields = task.params.get("fields", [])
            self._system_events.insert_slicer(fields)
            # Configure if specified
            if task.params.get("columns") or task.params.get("caption"):
                self._system_events.configure_slicer(
                    columns=task.params.get("columns"),
                    caption=task.params.get("caption"),
                )

        elif tt == TaskType.CHART_HISTOGRAM:
            self._system_events.insert_histogram(
                data_range=task.range or task.params.get("data_range"),
            )
            # Configure bins if specified
            if any(k in task.params for k in ("bin_width", "overflow", "underflow")):
                self._system_events.configure_histogram_bins(
                    bin_width=task.params.get("bin_width"),
                    overflow=task.params.get("overflow"),
                    underflow=task.params.get("underflow"),
                )

        elif tt == TaskType.TABLE_STYLE:
            style = task.style or task.params.get("style", "TableStyleMedium5")
            self._system_events.apply_table_style_via_ribbon(style)

        elif tt == TaskType.CHART_SCATTER:
            self._system_events.insert_scatter_chart(
                data_range=task.range or task.params.get("data_range"),
            )

        elif tt == TaskType.CHART_COMBO:
            self._system_events.insert_combo_chart(
                data_range=task.range or task.params.get("data_range"),
                secondary_axis=task.params.get("secondary_axis", True),
            )

        elif tt == TaskType.SPARKLINE:
            self._system_events.insert_sparkline(
                data_range=task.params.get("data_range", task.range or "B2:M2"),
                location_range=task.params.get("location_range", task.cell or "N2"),
                sparkline_type=task.params.get("sparkline_type", "Line"),
            )

        else:
            raise NotImplementedError(
                f"System Events handler not implemented for {tt.value}"
            )

    def _exec_vba(self, task: Task, workbook: Path) -> None:
        """Execute a task via Layer 5 (VBA)."""
        tt = task.task_type

        if tt == TaskType.PIVOT_TABLE:
            self._vba.create_pivot_table(
                source_sheet=task.params.get("source_sheet", task.sheet or "Sheet1"),
                source_range=task.params.get("source_range", task.range or "A1:A1"),
                dest_sheet=task.params.get("dest_sheet", "PivotSheet"),
                dest_cell=task.params.get("dest_cell", "A3"),
                table_name=task.params.get("table_name", "PivotTable1"),
                row_fields=task.params.get("row_fields"),
                column_fields=task.params.get("column_fields"),
                value_fields=task.params.get("value_fields"),
                filter_fields=task.params.get("filter_fields"),
            )

        elif tt == TaskType.PIVOT_CHART:
            self._vba.create_pivot_chart(
                pivot_table_name=task.params.get("pivot_table_name", "PivotTable1"),
                chart_type=task.params.get("chart_type", "xlColumnClustered"),
                chart_title=task.params.get("chart_title", ""),
                dest_sheet=task.params.get("dest_sheet"),
            )

        else:
            raise NotImplementedError(
                f"VBA handler not implemented for {tt.value}"
            )

    def _exec_pyautogui(self, task: Task, workbook: Path) -> None:
        """Execute a task via Layer 6 (PyAutoGUI) — last resort."""
        raise NotImplementedError(
            f"PyAutoGUI execution for {task.task_type.value} requires "
            f"screen coordinates — implement per assignment"
        )

    # ── Utility Methods ──

    def _save_workbook(self, workbook: Path) -> None:
        """Save via the best available method.

        Prefer xlwings when connected — it holds the live Excel state and is
        more likely to reflect the latest edits than the in-memory openpyxl
        workbook which may be stale after live-layer execution.
        """
        if self._xlwings._wb:
            self._xlwings.save()
        elif self._openpyxl.wb:
            self._openpyxl.save(workbook)
        else:
            try:
                self._applescript.save()
            except Exception:
                logger.warning("Could not save workbook")

    def _cleanup(self) -> None:
        """Clean up all layer connections."""
        try:
            self._openpyxl.close()
        except Exception as e:
            logger.warning("Cleanup failed for openpyxl: %s", e)
        try:
            self._xlwings.disconnect()
        except Exception as e:
            logger.warning("Cleanup failed for xlwings: %s", e)
        try:
            self._verifier.close()
        except Exception as e:
            logger.warning("Cleanup failed for verifier: %s", e)

    # ── Convenience Methods ──

    def scan(self, instructions: Path) -> list[Task]:
        """Phase 1: Scan instructions and extract tasks."""
        text = self._parser.parse(instructions)
        return self._extractor.extract(text)

    def plan(self, tasks: list[Task]) -> ExecutionPlan:
        """Phase 2: Create an execution plan from tasks."""
        return self._planner.plan(tasks)

    def verify(self, workbook: Path, tasks: list[Task]) -> SectionVerification:
        """Verify a set of tasks against a workbook."""
        self._verifier.load(workbook)
        result = self._verifier.verify_section("manual", tasks)
        self._verifier.close()
        return result

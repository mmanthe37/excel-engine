"""
Integration tests exercising the full Excel Engine pipeline:
  parse → extract → plan → execute (openpyxl) → verify

All tests run offline using openpyxl only — no Excel.app required.
"""

import pytest
from pathlib import Path

from openpyxl import load_workbook

from excel_engine.layers.openpyxl_layer import OpenpyxlLayer
from excel_engine.parsers.task_extractor import Task, TaskExtractor
from excel_engine.planner.task_planner import TaskPlanner
from excel_engine.verifier.workbook_verifier import WorkbookVerifier
from excel_engine.config import TaskType


# ═══════════════════════════════════════════════════════════════════
# Helpers
# ═══════════════════════════════════════════════════════════════════

def _open_layer(path: Path) -> OpenpyxlLayer:
    """Open a workbook with OpenpyxlLayer and return it."""
    layer = OpenpyxlLayer()
    layer.open(path)
    return layer


def _save_reload(layer: OpenpyxlLayer, path: Path) -> None:
    """Save via layer and reload so openpyxl picks up all XML changes."""
    layer.save(path)
    layer.open(path)


def _chart_title_text(chart) -> str:
    """Extract the plain-text title from a chart after save/reload.

    Before save, chart.title is a str; after reload it becomes a
    Title object with rich-text paragraphs.
    """
    title = chart.title
    if isinstance(title, str):
        return title
    if title is None:
        return ""
    # Title object → extract runs from rich text
    try:
        return "".join(
            r.t for p in title.tx.rich.paragraphs for r in p.r
        )
    except (AttributeError, TypeError):
        return str(title)


# ═══════════════════════════════════════════════════════════════════
# TestFormulaIntegration
# ═══════════════════════════════════════════════════════════════════

class TestFormulaIntegration:
    """Insert formulas via OpenpyxlLayer, verify with WorkbookVerifier."""

    def test_sum_formula(self, sample_workbook):
        layer = _open_layer(sample_workbook)
        layer.set_formula("F2", "=SUM(B2:E2)", sheet="Sales")
        layer.save(sample_workbook)

        verifier = WorkbookVerifier()
        verifier.load(sample_workbook)
        task = Task(
            id="sum1", task_type=TaskType.FORMULA,
            description="SUM row", sheet="Sales", cell="F2",
            formula="=SUM(B2:E2)",
        )
        result = verifier.verify_task(task)
        assert result.passed
        assert result.details["exact_match"] is True
        verifier.close()

    def test_average_formula(self, sample_workbook):
        layer = _open_layer(sample_workbook)
        layer.set_formula("F2", "=AVERAGE(B2:E2)", sheet="Sales")
        layer.save(sample_workbook)

        wb = load_workbook(sample_workbook)
        assert wb["Sales"]["F2"].value == "=AVERAGE(B2:E2)"
        wb.close()

    def test_if_formula(self, sample_workbook):
        layer = _open_layer(sample_workbook)
        layer.set_formula("G2", '=IF(B2>5000,"High","Low")', sheet="Sales")
        layer.save(sample_workbook)

        verifier = WorkbookVerifier()
        verifier.load(sample_workbook)
        task = Task(
            id="if1", task_type=TaskType.FORMULA,
            description="IF formula", sheet="Sales", cell="G2",
            formula='=IF(B2>5000,"High","Low")',
        )
        result = verifier.verify_task(task)
        assert result.passed
        verifier.close()

    def test_cross_sheet_formula(self, multi_sheet_workbook):
        """Insert a formula on Totals that references Jan, Feb, Mar sheets."""
        layer = _open_layer(multi_sheet_workbook)
        layer.set_formula("A1", "=Jan!B2+Feb!B2+Mar!B2", sheet="Totals")
        layer.save(multi_sheet_workbook)

        verifier = WorkbookVerifier()
        verifier.load(multi_sheet_workbook)
        task = Task(
            id="xsheet1", task_type=TaskType.FORMULA,
            description="cross-sheet sum", sheet="Totals", cell="A1",
            formula="=Jan!B2+Feb!B2+Mar!B2",
        )
        result = verifier.verify_task(task)
        assert result.passed
        assert result.details["exact_match"] is True
        verifier.close()

    def test_overwrite_existing_formula(self, sample_workbook):
        layer = _open_layer(sample_workbook)
        layer.set_formula("F2", "=SUM(B2:E2)", sheet="Sales")
        layer.save(sample_workbook)

        # Overwrite with a different formula
        layer = _open_layer(sample_workbook)
        layer.set_formula("F2", "=MAX(B2:E2)", sheet="Sales")
        layer.save(sample_workbook)

        wb = load_workbook(sample_workbook)
        assert wb["Sales"]["F2"].value == "=MAX(B2:E2)"
        wb.close()


# ═══════════════════════════════════════════════════════════════════
# TestChartIntegration
# ═══════════════════════════════════════════════════════════════════

class TestChartIntegration:
    """Add charts via OpenpyxlLayer, verify they persist after save/reload."""

    def test_bar_chart(self, sample_workbook):
        layer = _open_layer(sample_workbook)
        layer.add_bar_chart(
            sheet="Sales", title="Q1 Sales",
            data_range="B1:B11", cats_range="A2:A11",
        )
        layer.save(sample_workbook)

        wb = load_workbook(sample_workbook)
        charts = wb["Sales"]._charts
        assert len(charts) == 1
        assert _chart_title_text(charts[0]) == "Q1 Sales"
        wb.close()

    def test_scatter_chart(self, sample_workbook):
        layer = _open_layer(sample_workbook)
        layer.add_scatter_chart(
            sheet="Sales", title="Q1 vs Q2",
            x_range="B2:B11", y_range="C2:C11",
        )
        layer.save(sample_workbook)

        verifier = WorkbookVerifier()
        verifier.load(sample_workbook)
        task = Task(
            id="scatter1", task_type=TaskType.CHART_SCATTER,
            description="scatter chart", sheet="Sales",
        )
        result = verifier.verify_task(task)
        assert result.passed
        assert result.details["count"] == 1
        verifier.close()

    def test_multiple_charts_on_one_sheet(self, sample_workbook):
        layer = _open_layer(sample_workbook)
        layer.add_bar_chart(
            sheet="Sales", title="Bar Chart",
            data_range="B1:B11", anchor="E2",
        )
        layer.add_area_chart(
            sheet="Sales", title="Area Chart",
            data_range="C1:C11", anchor="E18",
        )
        layer.add_scatter_chart(
            sheet="Sales", title="Scatter Chart",
            x_range="B2:B11", y_range="D2:D11", anchor="E34",
        )
        layer.save(sample_workbook)

        wb = load_workbook(sample_workbook)
        charts = wb["Sales"]._charts
        assert len(charts) == 3
        titles = {_chart_title_text(c) for c in charts}
        assert titles == {"Bar Chart", "Area Chart", "Scatter Chart"}
        wb.close()


# ═══════════════════════════════════════════════════════════════════
# TestFormattingIntegration
# ═══════════════════════════════════════════════════════════════════

class TestFormattingIntegration:
    """Apply formatting via OpenpyxlLayer, verify with openpyxl and Verifier."""

    def test_font_bold_and_color(self, sample_workbook):
        layer = _open_layer(sample_workbook)
        layer.set_font("A1:E1", sheet="Sales", bold=True, color="FF0000", size=14)
        layer.save(sample_workbook)

        verifier = WorkbookVerifier()
        verifier.load(sample_workbook)
        task = Task(
            id="font1", task_type=TaskType.FONT,
            description="bold header", sheet="Sales", cell="A1",
        )
        result = verifier.verify_task(task)
        assert result.passed
        assert result.details["bold"] is True
        verifier.close()

    def test_fill_and_border(self, sample_workbook):
        layer = _open_layer(sample_workbook)
        layer.set_fill("A1:E1", color="4472C4", sheet="Sales")
        layer.set_border("A1:E11", sheet="Sales", style="thin", color="000000")
        layer.save(sample_workbook)

        wb = load_workbook(sample_workbook)
        cell = wb["Sales"]["A1"]
        assert cell.fill.start_color.index == "004472C4"
        assert cell.border.left.style == "thin"
        wb.close()

    def test_number_format(self, sample_workbook):
        layer = _open_layer(sample_workbook)
        layer.set_number_format("B2:E11", "$#,##0.00", sheet="Sales")
        layer.save(sample_workbook)

        verifier = WorkbookVerifier()
        verifier.load(sample_workbook)
        task = Task(
            id="numfmt1", task_type=TaskType.NUMBER_FORMAT,
            description="currency format", sheet="Sales", cell="B2",
        )
        result = verifier.verify_task(task)
        assert result.passed
        assert result.details["number_format"] == "$#,##0.00"
        verifier.close()

    def test_column_width_and_row_height(self, sample_workbook):
        layer = _open_layer(sample_workbook)
        layer.set_column_width("A", 25.0, sheet="Sales")
        layer.set_row_height(1, 30.0, sheet="Sales")
        layer.save(sample_workbook)

        wb = load_workbook(sample_workbook)
        ws = wb["Sales"]
        assert ws.column_dimensions["A"].width == 25.0
        assert ws.row_dimensions[1].height == 30.0
        wb.close()


# ═══════════════════════════════════════════════════════════════════
# TestTableIntegration
# ═══════════════════════════════════════════════════════════════════

class TestTableIntegration:
    """Create Excel tables via OpenpyxlLayer, verify via WorkbookVerifier."""

    def test_create_table_with_range(self, sample_workbook):
        layer = _open_layer(sample_workbook)
        layer.create_table(
            name="SalesTable", ref="A1:E11",
            style="TableStyleMedium5", sheet="Sales",
        )
        layer.save(sample_workbook)

        verifier = WorkbookVerifier()
        verifier.load(sample_workbook)
        task = Task(
            id="tbl1", task_type=TaskType.TABLE_CREATE,
            description="create table", sheet="Sales", range="A1:E11",
        )
        result = verifier.verify_task(task)
        assert result.passed
        assert any(t["name"] == "SalesTable" for t in result.details["tables"])
        verifier.close()

    def test_table_with_total_row(self, sample_workbook):
        layer = _open_layer(sample_workbook)
        layer.create_table(
            name="SalesTable", ref="A1:E11",
            style="TableStyleMedium5", sheet="Sales",
            show_totals=True,
        )
        layer.save(sample_workbook)

        verifier = WorkbookVerifier()
        verifier.load(sample_workbook)
        task = Task(
            id="tbl_total", task_type=TaskType.TABLE_TOTAL_ROW,
            description="table total row", sheet="Sales",
        )
        result = verifier.verify_task(task)
        assert result.passed
        verifier.close()


# ═══════════════════════════════════════════════════════════════════
# TestMultiSheetIntegration
# ═══════════════════════════════════════════════════════════════════

class TestMultiSheetIntegration:
    """Create/rename sheets and operate across them."""

    def test_create_sheets_and_cross_reference(self, sample_workbook):
        layer = _open_layer(sample_workbook)
        layer.create_sheet("Analysis")
        # Write a formula on Summary referencing Sales
        layer.set_formula("A1", "=Sales!B2", sheet="Summary")
        # Write a formula on Analysis referencing Sales
        layer.set_formula("A1", "=SUM(Sales!B2:B11)", sheet="Analysis")
        layer.save(sample_workbook)

        wb = load_workbook(sample_workbook)
        assert "Analysis" in wb.sheetnames
        assert wb["Summary"]["A1"].value == "=Sales!B2"
        assert wb["Analysis"]["A1"].value == "=SUM(Sales!B2:B11)"
        wb.close()

    def test_rename_sheet_and_operate(self, sample_workbook):
        layer = _open_layer(sample_workbook)
        layer.rename_sheet("Summary", "Overview")
        layer.set_value("A1", "Total Revenue", sheet="Overview")
        layer.set_formula("B1", "=SUM(Sales!B2:B11)", sheet="Overview")
        layer.save(sample_workbook)

        verifier = WorkbookVerifier()
        verifier.load(sample_workbook)

        sheet_task = Task(
            id="rename1", task_type=TaskType.SHEET_RENAME,
            description="rename sheet", sheet="Overview",
        )
        result = verifier.verify_task(sheet_task)
        assert result.passed

        formula_task = Task(
            id="ov_formula", task_type=TaskType.FORMULA,
            description="overview formula", sheet="Overview", cell="B1",
            formula="=SUM(Sales!B2:B11)",
        )
        result = verifier.verify_task(formula_task)
        assert result.passed
        assert result.details["exact_match"] is True
        verifier.close()


# ═══════════════════════════════════════════════════════════════════
# TestFullPipelineIntegration
# ═══════════════════════════════════════════════════════════════════

class TestFullPipelineIntegration:
    """End-to-end: instruction text → extract → plan → execute → verify."""

    def test_formula_pipeline(self, sample_workbook):
        """Parse a formula instruction, plan it, execute, and verify."""
        instruction = "In cell F2 on the Sales sheet, enter the formula =SUM(B2:E2)"

        # 1. Extract
        extractor = TaskExtractor()
        tasks = extractor.extract(instruction)
        assert len(tasks) >= 1
        formula_task = next(
            (t for t in tasks if t.task_type == TaskType.FORMULA), None
        )
        assert formula_task is not None

        # 2. Plan
        planner = TaskPlanner()
        plan = planner.plan(tasks)
        assert plan.total_tasks >= 1

        # 3. Execute manually via OpenpyxlLayer
        layer = _open_layer(sample_workbook)
        layer.set_formula("F2", "=SUM(B2:E2)", sheet="Sales")
        layer.save(sample_workbook)

        # 4. Verify
        verifier = WorkbookVerifier()
        verifier.load(sample_workbook)
        exec_task = Task(
            id="pipe_sum", task_type=TaskType.FORMULA,
            description="SUM pipeline", sheet="Sales", cell="F2",
            formula="=SUM(B2:E2)",
        )
        section_result = verifier.verify_section("sec_01", [exec_task])
        assert section_result.all_passed
        assert section_result.pass_count == 1
        verifier.close()

    def test_chart_pipeline(self, sample_workbook):
        """Parse a chart instruction, plan it, execute, and verify."""
        instruction = "Add a bar chart of the Q1 data on the Sales sheet"

        # 1. Extract
        extractor = TaskExtractor()
        tasks = extractor.extract(instruction)
        assert len(tasks) >= 1
        chart_task = next(
            (t for t in tasks if t.task_type == TaskType.CHART_BAR), None
        )
        assert chart_task is not None

        # 2. Plan
        planner = TaskPlanner()
        plan = planner.plan(tasks)
        assert plan.total_tasks >= 1

        # 3. Execute
        layer = _open_layer(sample_workbook)
        layer.add_bar_chart(
            sheet="Sales", title="Q1 Sales",
            data_range="B1:B11", cats_range="A2:A11",
        )
        layer.save(sample_workbook)

        # 4. Verify
        verifier = WorkbookVerifier()
        verifier.load(sample_workbook)
        verify_task = Task(
            id="pipe_chart", task_type=TaskType.CHART_BAR,
            description="bar chart pipeline", sheet="Sales",
        )
        section_result = verifier.verify_section("sec_02", [verify_task])
        assert section_result.all_passed
        verifier.close()

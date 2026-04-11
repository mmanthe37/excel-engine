"""Formula recalculation and error scanning via LibreOffice."""

from __future__ import annotations

import logging
import os
import platform
import shutil
import subprocess
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

EXCEL_ERRORS = ["#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A"]

MACRO_XML = """\
<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" script:name="Module1" script:language="StarBasic">
    Sub RecalculateAndSave()
      ThisComponent.calculateAll()
      ThisComponent.store()
      ThisComponent.close(True)
    End Sub
</script:module>
"""

MACRO_SCRIPT_URI = (
    "vnd.sun.star.script:Standard.Module1.RecalculateAndSave"
    "?language=Basic&location=application"
)


@dataclass
class RecalcResult:
    success: bool
    skipped: bool = False
    total_errors: int = 0
    total_formulas: int = 0
    error_summary: dict = field(default_factory=dict)
    warning: str = ""


# ---------------------------------------------------------------------------
# Macro installation
# ---------------------------------------------------------------------------

def _macro_dir() -> Path:
    """Return the LibreOffice user macro directory for the current platform."""
    system = platform.system()
    if system == "Darwin":
        return Path.home() / "Library/Application Support/LibreOffice/4/user/basic/Standard"
    # Linux (and other POSIX)
    return Path.home() / ".config/libreoffice/4/user/basic/Standard"


def _ensure_macro_installed() -> None:
    """Install the RecalculateAndSave macro if it is not already present."""
    macro_path = _macro_dir() / "Module1.xba"
    if macro_path.exists():
        logger.debug("LibreOffice macro already installed at %s", macro_path)
        return

    logger.info("Installing LibreOffice RecalculateAndSave macro to %s", macro_path)
    macro_path.parent.mkdir(parents=True, exist_ok=True)
    macro_path.write_text(MACRO_XML, encoding="utf-8")


# ---------------------------------------------------------------------------
# LibreOffice detection
# ---------------------------------------------------------------------------

def _find_soffice() -> str | None:
    """Return the soffice binary path, or *None* if LibreOffice is not installed."""
    return shutil.which("soffice") or shutil.which("libreoffice")


def _find_timeout_cmd() -> str | None:
    """Return a ``timeout`` command name available on this platform."""
    system = platform.system()
    if system == "Linux":
        return shutil.which("timeout")
    if system == "Darwin":
        # coreutils provides gtimeout on macOS
        return shutil.which("gtimeout")
    return None


# ---------------------------------------------------------------------------
# Recalculation
# ---------------------------------------------------------------------------

def recalculate(filepath: Path, timeout: int = 30) -> RecalcResult:
    """Recalculate *filepath* via LibreOffice headless, then scan for errors.

    If LibreOffice is not available the file is still scanned for cached
    formula errors but ``RecalcResult.skipped`` will be ``True``.
    """
    filepath = Path(filepath).resolve()

    soffice = _find_soffice()
    if soffice is None:
        warning = "LibreOffice not installed, formulas not recalculated"
        logger.warning(warning)
        errors = scan_formula_errors(filepath)
        return RecalcResult(
            success=True,
            skipped=True,
            total_errors=errors["total_errors"],
            total_formulas=errors["total_formulas"],
            error_summary=errors["error_summary"],
            warning=warning,
        )

    _ensure_macro_installed()

    env = os.environ.copy()
    env["SAL_USE_VCLPLUGIN"] = "svp"

    cmd: list[str] = []
    timeout_bin = _find_timeout_cmd()
    if timeout_bin:
        cmd.extend([timeout_bin, str(timeout)])
    cmd.extend([soffice, "--headless", "--norestore", MACRO_SCRIPT_URI, str(filepath)])

    logger.info("Running LibreOffice recalculation: %s", " ".join(cmd))

    try:
        result = subprocess.run(
            cmd,
            env=env,
            capture_output=True,
            timeout=timeout + 10 if not timeout_bin else None,
        )

        if result.returncode != 0:
            stderr = result.stderr.decode(errors="replace").strip()
            logger.error(
                "LibreOffice exited with code %d: %s", result.returncode, stderr
            )
            errors = scan_formula_errors(filepath)
            return RecalcResult(
                success=False,
                total_errors=errors["total_errors"],
                total_formulas=errors["total_formulas"],
                error_summary=errors["error_summary"],
                warning=f"LibreOffice exited with code {result.returncode}",
            )

    except subprocess.TimeoutExpired:
        logger.error("LibreOffice timed out after %d seconds", timeout)
        errors = scan_formula_errors(filepath)
        return RecalcResult(
            success=False,
            total_errors=errors["total_errors"],
            total_formulas=errors["total_formulas"],
            error_summary=errors["error_summary"],
            warning=f"LibreOffice timed out after {timeout}s",
        )

    except OSError as exc:
        # Catch AF_UNIX / sandbox socket errors and similar OS-level failures
        if "AF_UNIX" in str(exc) or "socket" in str(exc).lower():
            warning = (
                "Sandboxed environment detected (AF_UNIX sockets blocked); "
                "skipping LibreOffice recalculation"
            )
            logger.warning(warning)
            errors = scan_formula_errors(filepath)
            return RecalcResult(
                success=True,
                skipped=True,
                total_errors=errors["total_errors"],
                total_formulas=errors["total_formulas"],
                error_summary=errors["error_summary"],
                warning=warning,
            )
        raise

    logger.info("LibreOffice recalculation completed successfully")
    errors = scan_formula_errors(filepath)
    return RecalcResult(
        success=True,
        total_errors=errors["total_errors"],
        total_formulas=errors["total_formulas"],
        error_summary=errors["error_summary"],
    )


# ---------------------------------------------------------------------------
# Error scanning
# ---------------------------------------------------------------------------

def scan_formula_errors(filepath: Path) -> dict:
    """Scan *filepath* for Excel formula errors and return a summary dict.

    Works independently of LibreOffice — only requires ``openpyxl``.

    Returns::

        {
            "total_errors": int,
            "total_formulas": int,
            "error_summary": {
                "#VALUE!": {"count": int, "locations": ["Sheet1!A2", ...]},
                ...
            },
        }
    """
    filepath = Path(filepath)

    # --- Pass 1: count formulas (data_only=False) ---
    total_formulas = 0
    wb_formulas = load_workbook(filepath, data_only=False, read_only=True)
    try:
        for ws in wb_formulas.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        total_formulas += 1
    finally:
        wb_formulas.close()

    # --- Pass 2: scan cached values for error strings (data_only=True) ---
    error_summary: dict[str, dict] = {}
    total_errors = 0

    wb_values = load_workbook(filepath, data_only=True, read_only=True)
    try:
        for ws in wb_values.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    val = cell.value
                    if not isinstance(val, str):
                        continue
                    if val not in EXCEL_ERRORS:
                        continue

                    total_errors += 1
                    loc = f"{ws.title}!{cell.coordinate}"

                    if val not in error_summary:
                        error_summary[val] = {"count": 0, "locations": []}

                    entry = error_summary[val]
                    entry["count"] += 1
                    if len(entry["locations"]) < 20:
                        entry["locations"].append(loc)
    finally:
        wb_values.close()

    return {
        "total_errors": total_errors,
        "total_formulas": total_formulas,
        "error_summary": error_summary,
    }

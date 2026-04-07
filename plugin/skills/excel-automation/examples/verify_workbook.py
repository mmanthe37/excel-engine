#!/usr/bin/env python3
"""
Example: Verify an Excel workbook against expected values.

Reads a workbook and checks that specific cells contain expected
formulas, values, or formatting. Useful for validating assignment
completion without re-running the engine.

Usage:
    python verify_workbook.py /path/to/workbook.xlsx
"""

import sys
from pathlib import Path

from openpyxl import load_workbook


def verify_cell(ws, cell_ref, expected_value=None, expected_formula=None,
                expected_format=None):
    """Verify a single cell meets expectations."""
    cell = ws[cell_ref]
    results = []

    if expected_formula is not None:
        actual = cell.value
        if isinstance(actual, str) and actual.startswith("="):
            if actual.upper() == expected_formula.upper():
                results.append(f"  ✅ {cell_ref} formula: {actual}")
            else:
                results.append(f"  ❌ {cell_ref} formula: got {actual}, expected {expected_formula}")
        else:
            results.append(f"  ❌ {cell_ref} no formula found (value: {actual})")

    if expected_value is not None:
        actual = cell.value
        if actual == expected_value:
            results.append(f"  ✅ {cell_ref} value: {actual}")
        else:
            results.append(f"  ❌ {cell_ref} value: got {actual}, expected {expected_value}")

    if expected_format is not None:
        actual = cell.number_format
        if actual == expected_format:
            results.append(f"  ✅ {cell_ref} format: {actual}")
        else:
            results.append(f"  ❌ {cell_ref} format: got {actual}, expected {expected_format}")

    return results


def main():
    if len(sys.argv) < 2:
        print("Usage: verify_workbook.py <workbook.xlsx>")
        sys.exit(1)

    path = Path(sys.argv[1]).resolve()
    if not path.exists():
        print(f"Error: File not found: {path}")
        sys.exit(1)

    print(f"=== Workbook Verification ===")
    print(f"File: {path.name}")
    print()

    wb = load_workbook(str(path))

    # Print workbook structure
    print("Sheets:", ", ".join(wb.sheetnames))
    print()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"--- {sheet_name} ---")
        print(f"  Dimensions: {ws.dimensions}")
        print(f"  Tables: {list(ws.tables.keys()) if ws.tables else 'none'}")

        # Check for formulas
        formula_count = 0
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_count += 1
        print(f"  Formulas: {formula_count}")

        # Check for conditional formatting
        cf_count = len(ws.conditional_formatting)
        if cf_count:
            print(f"  Conditional formats: {cf_count}")

        print()

    # Example verification checks — customize per assignment
    # Uncomment and modify for specific assignments:
    #
    # ws = wb['Summary']
    # results = []
    # results += verify_cell(ws, 'B5', expected_formula='=SUM(B2:B4)')
    # results += verify_cell(ws, 'B5', expected_format='$#,##0.00')
    # for r in results:
    #     print(r)

    wb.close()
    print("Verification complete.")


if __name__ == "__main__":
    main()

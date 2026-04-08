"""
Shared pytest fixtures for integration tests.

Generates sample .xlsx workbooks programmatically using openpyxl
so that tests run offline without Excel.app.
"""

import pytest
from pathlib import Path
from openpyxl import Workbook


@pytest.fixture
def sample_workbook(tmp_path):
    """Create a sample workbook with Sales data and an empty Summary sheet."""
    wb = Workbook()

    # Sheet 1: Sales data with headers in row 1, data in rows 2-11
    ws = wb.active
    ws.title = "Sales"
    headers = ["Product", "Q1", "Q2", "Q3", "Q4"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    products = [
        "Widget A", "Widget B", "Widget C", "Gadget X", "Gadget Y",
        "Gizmo 1", "Gizmo 2", "Part Alpha", "Part Beta", "Part Gamma",
    ]
    import random
    random.seed(42)
    for row_idx, product in enumerate(products, 2):
        ws.cell(row=row_idx, column=1, value=product)
        for col_idx in range(2, 6):
            ws.cell(row=row_idx, column=col_idx, value=random.randint(100, 9999))

    # Sheet 2: Empty "Summary" sheet for formula targets
    wb.create_sheet("Summary")

    path = tmp_path / "test_workbook.xlsx"
    wb.save(path)
    return path


@pytest.fixture
def multi_sheet_workbook(tmp_path):
    """Create a workbook with three data sheets for cross-sheet testing."""
    wb = Workbook()

    for name in ["Jan", "Feb", "Mar"]:
        if name == "Jan":
            ws = wb.active
            ws.title = name
        else:
            ws = wb.create_sheet(name)
        ws["A1"] = "Region"
        ws["B1"] = "Revenue"
        for i in range(2, 7):
            ws.cell(row=i, column=1, value=f"Region {i - 1}")
            ws.cell(row=i, column=2, value=(i - 1) * 1000)

    # Totals sheet
    wb.create_sheet("Totals")

    path = tmp_path / "multi_sheet.xlsx"
    wb.save(path)
    return path

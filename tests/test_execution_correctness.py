"""
Execution Correctness Tests — Verify that every major TaskType
actually produces the correct result in the .xlsx file.

These tests exercise the full pipeline: Task → plan → execute → save → verify.
"""

import pytest
import openpyxl
from pathlib import Path

from excel_engine import ExcelEngine, EngineConfig, TaskType, Layer
from excel_engine.parsers.task_extractor import Task


@pytest.fixture
def workbook_path(tmp_path):
    """Create a pre-populated test workbook."""
    fpath = tmp_path / "test.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for i in range(1, 11):
        ws[f"A{i}"] = i * 10
        ws[f"B{i}"] = f"Item {i}"
    wb.save(str(fpath))
    return fpath


@pytest.fixture
def engine():
    return ExcelEngine(EngineConfig())


def run_tasks(engine, workbook_path, tasks):
    """Run tasks and return (result, reloaded_workbook)."""
    result = engine.run(workbook_path, tasks=tasks)
    wb = openpyxl.load_workbook(str(workbook_path))
    return result, wb


# ═══════════════════════════════════════════════════════════════════════════
# A) TASKS THAT EXECUTE CORRECTLY
# ═══════════════════════════════════════════════════════════════════════════


class TestCorrectExecution:
    """Tests that verify working task types."""

    def test_cell_value_string(self, workbook_path, engine):
        result, wb = run_tasks(engine, workbook_path, [
            Task(id="t1", task_type=TaskType.CELL_VALUE, description="d",
                 cell="C1", value="Hello", sheet="Sheet1"),
        ])
        assert wb.active["C1"].value == "Hello"

    def test_formula_with_equals(self, workbook_path, engine):
        result, wb = run_tasks(engine, workbook_path, [
            Task(id="t1", task_type=TaskType.FORMULA, description="d",
                 cell="C1", formula="=SUM(A1:A10)", sheet="Sheet1"),
        ])
        assert wb.active["C1"].value == "=SUM(A1:A10)"

    def test_formula_without_equals(self, workbook_path, engine):
        """OpenpyxlLayer.set_formula auto-prepends = if missing."""
        result, wb = run_tasks(engine, workbook_path, [
            Task(id="t1", task_type=TaskType.FORMULA, description="d",
                 cell="C1", formula="SUM(A1:A10)", sheet="Sheet1"),
        ])
        assert wb.active["C1"].value == "=SUM(A1:A10)"

    def test_formatting_bold(self, workbook_path, engine):
        result, wb = run_tasks(engine, workbook_path, [
            Task(id="t1", task_type=TaskType.FORMATTING, description="d",
                 cell="A1", sheet="Sheet1", params={"bold": True}),
        ])
        assert wb.active["A1"].font.bold is True

    def test_column_width_with_params(self, workbook_path, engine):
        result, wb = run_tasks(engine, workbook_path, [
            Task(id="t1", task_type=TaskType.COLUMN_WIDTH, description="d",
                 sheet="Sheet1", params={"column": "A", "width": 25}),
        ])
        assert wb.active.column_dimensions["A"].width == 25

    def test_row_height(self, workbook_path, engine):
        result, wb = run_tasks(engine, workbook_path, [
            Task(id="t1", task_type=TaskType.ROW_HEIGHT, description="d",
                 cell="A5", sheet="Sheet1", params={"height": 30}),
        ])
        assert wb.active.row_dimensions[5].height == 30

    def test_merge_cells(self, workbook_path, engine):
        result, wb = run_tasks(engine, workbook_path, [
            Task(id="t1", task_type=TaskType.MERGE_CELLS, description="d",
                 range="A1:D1", sheet="Sheet1"),
        ])
        assert "A1:D1" in [str(m) for m in wb.active.merged_cells.ranges]

    def test_named_range(self, workbook_path, engine):
        result, wb = run_tasks(engine, workbook_path, [
            Task(id="t1", task_type=TaskType.NAMED_RANGE, description="d",
                 sheet="Sheet1", params={"name": "TestRange", "range": "$A$1:$A$10"}),
        ])
        assert "TestRange" in list(wb.defined_names)

    def test_tab_color(self, workbook_path, engine):
        result, wb = run_tasks(engine, workbook_path, [
            Task(id="t1", task_type=TaskType.TAB_COLOR, description="d",
                 sheet="Sheet1", params={"color": "FF0000"}),
        ])
        assert wb.active.sheet_properties.tabColor is not None

    def test_hyperlink(self, workbook_path, engine):
        result, wb = run_tasks(engine, workbook_path, [
            Task(id="t1", task_type=TaskType.HYPERLINK, description="d",
                 cell="A1", sheet="Sheet1",
                 params={"url": "https://example.com", "display": "Click"}),
        ])
        assert wb.active["A1"].hyperlink is not None

    def test_freeze_panes(self, workbook_path, engine):
        result, wb = run_tasks(engine, workbook_path, [
            Task(id="t1", task_type=TaskType.FREEZE_PANES, description="d",
                 cell="A2", sheet="Sheet1"),
        ])
        assert wb.active.freeze_panes == "A2"

    def test_autofilter(self, workbook_path, engine):
        result, wb = run_tasks(engine, workbook_path, [
            Task(id="t1", task_type=TaskType.AUTOFILTER, description="d",
                 range="A1:B10", sheet="Sheet1"),
        ])
        assert wb.active.auto_filter.ref is not None

    def test_table_create(self, workbook_path, engine):
        result, wb = run_tasks(engine, workbook_path, [
            Task(id="t1", task_type=TaskType.TABLE_CREATE, description="d",
                 range="A1:B10", sheet="Sheet1", params={"name": "TestTable"}),
        ])
        assert "TestTable" in wb.active.tables

    def test_conditional_format(self, workbook_path, engine):
        result, wb = run_tasks(engine, workbook_path, [
            Task(id="t1", task_type=TaskType.CONDITIONAL_FORMAT, description="d",
                 range="A1:A10", sheet="Sheet1",
                 params={"operator": "greaterThan", "formula": ["50"]}),
        ])
        assert len(wb.active.conditional_formatting) > 0

    def test_data_validation_formula1(self, workbook_path, engine):
        result, wb = run_tasks(engine, workbook_path, [
            Task(id="t1", task_type=TaskType.DATA_VALIDATION, description="d",
                 cell="C1", sheet="Sheet1",
                 params={"type": "list", "formula1": '"Yes,No,Maybe"'}),
        ])
        assert len(wb.active.data_validations.dataValidation) > 0

    def test_sheet_create(self, workbook_path, engine):
        result, wb = run_tasks(engine, workbook_path, [
            Task(id="t1", task_type=TaskType.SHEET_CREATE, description="d",
                 sheet="Sheet1", params={"name": "NewSheet"}),
        ])
        assert "NewSheet" in wb.sheetnames

    def test_sheet_rename(self, workbook_path, engine):
        result, wb = run_tasks(engine, workbook_path, [
            Task(id="t1", task_type=TaskType.SHEET_RENAME, description="d",
                 sheet="Sheet1",
                 params={"old_name": "Sheet1", "new_name": "Data"}),
        ])
        assert "Data" in wb.sheetnames

    def test_chart_bar(self, workbook_path, engine):
        result, wb = run_tasks(engine, workbook_path, [
            Task(id="t1", task_type=TaskType.CHART_BAR, description="d",
                 sheet="Sheet1",
                 params={"title": "Sales", "data_range": "A1:A10"}),
        ])
        assert len(wb.active._charts) > 0

    def test_chart_line(self, workbook_path, engine):
        result, wb = run_tasks(engine, workbook_path, [
            Task(id="t1", task_type=TaskType.CHART_LINE, description="d",
                 sheet="Sheet1", params={"data_range": "A1:A10"}),
        ])
        assert len(wb.active._charts) > 0

    def test_chart_pie(self, workbook_path, engine):
        result, wb = run_tasks(engine, workbook_path, [
            Task(id="t1", task_type=TaskType.CHART_PIE, description="d",
                 sheet="Sheet1", params={"data_range": "A1:A10"}),
        ])
        assert len(wb.active._charts) > 0

    def test_chart_scatter(self, workbook_path, engine):
        result, wb = run_tasks(engine, workbook_path, [
            Task(id="t1", task_type=TaskType.CHART_SCATTER, description="d",
                 sheet="Sheet1",
                 params={"x_range": "A1:A10", "y_range": "B1:B10"}),
        ])
        assert len(wb.active._charts) > 0

    def test_chart_area(self, workbook_path, engine):
        result, wb = run_tasks(engine, workbook_path, [
            Task(id="t1", task_type=TaskType.CHART_AREA, description="d",
                 sheet="Sheet1", params={"data_range": "A1:A10"}),
        ])
        assert len(wb.active._charts) > 0

    def test_chart_combo(self, workbook_path, engine):
        result, wb = run_tasks(engine, workbook_path, [
            Task(id="t1", task_type=TaskType.CHART_COMBO, description="d",
                 sheet="Sheet1",
                 params={"bar_data_range": "A1:A10",
                         "line_data_range": "B1:B10"}),
        ])
        assert len(wb.active._charts) > 0

    def test_page_break(self, workbook_path, engine):
        result, wb = run_tasks(engine, workbook_path, [
            Task(id="t1", task_type=TaskType.PAGE_BREAK, description="d",
                 cell="A5", sheet="Sheet1"),
        ])
        assert len(wb.active.row_breaks.brk) > 0

    def test_print_settings(self, workbook_path, engine):
        result, wb = run_tasks(engine, workbook_path, [
            Task(id="t1", task_type=TaskType.PRINT_SETTINGS, description="d",
                 sheet="Sheet1",
                 params={"landscape": True, "print_area": "A1:B10"}),
        ])
        assert wb.active.page_setup.orientation == "landscape"

    def test_text_function(self, workbook_path, engine):
        result, wb = run_tasks(engine, workbook_path, [
            Task(id="t1", task_type=TaskType.TEXT_FUNCTION, description="d",
                 cell="C1", formula="=UPPER(B1)", sheet="Sheet1"),
        ])
        assert wb.active["C1"].value == "=UPPER(B1)"

    def test_sheet_move(self, workbook_path, engine):
        # Add sheets first
        wb0 = openpyxl.load_workbook(str(workbook_path))
        wb0.create_sheet("Sheet2")
        wb0.create_sheet("Sheet3")
        wb0.save(str(workbook_path))

        result, wb = run_tasks(engine, workbook_path, [
            Task(id="t1", task_type=TaskType.SHEET_MOVE, description="d",
                 sheet="Sheet1", params={"position": 2}),
        ])
        assert result.success

    def test_sheet_copy(self, workbook_path, engine):
        result, wb = run_tasks(engine, workbook_path, [
            Task(id="t1", task_type=TaskType.SHEET_COPY, description="d",
                 sheet="Sheet1", params={"new_name": "Backup"}),
        ])
        assert "Backup" in wb.sheetnames

    def test_lookup_function(self, workbook_path, engine):
        result, wb = run_tasks(engine, workbook_path, [
            Task(id="t1", task_type=TaskType.LOOKUP_FUNCTION, description="d",
                 cell="C1", formula="=VLOOKUP(A1,A1:B10,2,FALSE)",
                 sheet="Sheet1"),
        ])
        assert "VLOOKUP" in str(wb.active["C1"].value)

    def test_table_total_row(self, workbook_path, engine):
        # Pre-create table
        from openpyxl.worksheet.table import Table, TableStyleInfo
        wb0 = openpyxl.load_workbook(str(workbook_path))
        ws0 = wb0.active
        t = Table(displayName="MyTable", ref="A1:B10")
        t.tableStyleInfo = TableStyleInfo(name="TableStyleMedium5")
        ws0.add_table(t)
        wb0.save(str(workbook_path))

        result, wb = run_tasks(engine, workbook_path, [
            Task(id="t1", task_type=TaskType.TABLE_TOTAL_ROW, description="d",
                 sheet="Sheet1", params={"name": "MyTable"}),
        ])
        assert any(t.totalsRowShown for t in wb.active.tables.values())

    def test_empty_task_list(self, workbook_path, engine):
        result, wb = run_tasks(engine, workbook_path, [])
        assert result.success

    def test_multi_task_pipeline(self, workbook_path, engine):
        """Multiple tasks in sequence all apply correctly."""
        result, wb = run_tasks(engine, workbook_path, [
            Task(id="m1", task_type=TaskType.CELL_VALUE, description="d",
                 cell="C1", value="Total", sheet="Sheet1"),
            Task(id="m2", task_type=TaskType.FORMULA, description="d",
                 cell="C2", formula="=SUM(A1:A10)", sheet="Sheet1"),
            Task(id="m3", task_type=TaskType.FORMATTING, description="d",
                 cell="C1", sheet="Sheet1", params={"bold": True}),
            Task(id="m4", task_type=TaskType.NUMBER_FORMAT, description="d",
                 cell="C2", sheet="Sheet1", params={"format": "$#,##0.00"}),
        ])
        ws = wb.active
        assert ws["C1"].value == "Total"
        assert ws["C2"].value == "=SUM(A1:A10)"
        assert ws["C1"].font.bold is True
        assert ws["C2"].number_format == "$#,##0.00"


# ═══════════════════════════════════════════════════════════════════════════
# B) TASKS THAT FAIL OR PRODUCE WRONG RESULTS
# ═══════════════════════════════════════════════════════════════════════════


class TestBugCellValueFalsy:
    """
    [HIGH] CELL_VALUE — Falsy values were silently dropped (now FIXED).

    Root cause was: engine.py line 415 used `if task.cell and task.value:`
    which evaluated False for value=0, value="", value=False.
    Fix: changed to `if task.cell and task.value is not None:`
    """

    def test_cell_value_zero(self, workbook_path, engine):
        result, wb = run_tasks(engine, workbook_path, [
            Task(id="t1", task_type=TaskType.CELL_VALUE, description="d",
                 cell="C1", value=0, sheet="Sheet1"),
        ])
        assert wb.active["C1"].value == 0

    def test_cell_value_empty_string(self, workbook_path, engine):
        result, wb = run_tasks(engine, workbook_path, [
            Task(id="t1", task_type=TaskType.CELL_VALUE, description="d",
                 cell="C1", value="", sheet="Sheet1"),
        ])
        # openpyxl converts "" → None on save/reload; engine correctly
        # passes the value through (no longer dropped by falsy guard)
        assert wb.active["C1"].value in ("", None)

    def test_cell_value_false(self, workbook_path, engine):
        result, wb = run_tasks(engine, workbook_path, [
            Task(id="t1", task_type=TaskType.CELL_VALUE, description="d",
                 cell="C1", value=False, sheet="Sheet1"),
        ])
        assert wb.active["C1"].value is False


class TestBugSingleCellIteration:
    """
    [HIGH] NUMBER_FORMAT, FONT, FILL, ALIGNMENT, BORDER — were crashing on
    single cells (now FIXED).

    Root cause was: openpyxl_layer.py methods did `for row in ws[cells]:`
    When cells="A1" (single cell), ws["A1"] returns a Cell object, not
    a tuple-of-tuples. Iterating a Cell raised:
        TypeError: 'Cell' object is not iterable

    Fix: Detect single cell (no ':' in cells) and apply directly.
    """

    def test_number_format_single_cell(self, workbook_path, engine):
        result, wb = run_tasks(engine, workbook_path, [
            Task(id="t1", task_type=TaskType.NUMBER_FORMAT, description="d",
                 cell="A1", sheet="Sheet1", params={"format": "$#,##0.00"}),
        ])
        assert wb.active["A1"].number_format == "$#,##0.00"

    def test_font_single_cell(self, workbook_path, engine):
        result, wb = run_tasks(engine, workbook_path, [
            Task(id="t1", task_type=TaskType.FONT, description="d",
                 cell="A1", sheet="Sheet1",
                 params={"bold": True, "size": 14}),
        ])
        assert wb.active["A1"].font.bold is True
        assert wb.active["A1"].font.size == 14

    def test_fill_single_cell(self, workbook_path, engine):
        result, wb = run_tasks(engine, workbook_path, [
            Task(id="t1", task_type=TaskType.FILL, description="d",
                 cell="A1", sheet="Sheet1", params={"color": "FFFF00"}),
        ])
        assert "FFFF00" in str(wb.active["A1"].fill.start_color.rgb)

    def test_alignment_single_cell(self, workbook_path, engine):
        result, wb = run_tasks(engine, workbook_path, [
            Task(id="t1", task_type=TaskType.ALIGNMENT, description="d",
                 cell="A1", sheet="Sheet1",
                 params={"horizontal": "center", "wrap_text": True}),
        ])
        assert wb.active["A1"].alignment.horizontal == "center"

    def test_border_single_cell(self, workbook_path, engine):
        result, wb = run_tasks(engine, workbook_path, [
            Task(id="t1", task_type=TaskType.BORDER, description="d",
                 cell="A1", sheet="Sheet1", params={"style": "thin"}),
        ])
        assert wb.active["A1"].border.left.style == "thin"


class TestBugColumnWidthNoColumnParam:
    """
    [MEDIUM] COLUMN_WIDTH — was ignoring task.cell when params["column"] missing
    (now FIXED).

    Root cause was: engine.py did `col = task.params.get("column", "A")`
    Fix: Extract column letter from task.cell as fallback.
    """

    def test_column_width_from_cell(self, workbook_path, engine):
        result, wb = run_tasks(engine, workbook_path, [
            Task(id="t1", task_type=TaskType.COLUMN_WIDTH, description="d",
                 cell="B1", sheet="Sheet1", params={"width": 30}),
        ])
        assert wb.active.column_dimensions["B"].width == 30


class TestBugDVValuesNotConverted:
    """
    [MEDIUM] DATA_VALIDATION — params["values"] list was not converted to
    formula1 (now FIXED).

    Root cause was: engine.py passed params.get("formula1") directly.
    Fix: Convert ["A","B","C"] to '"A,B,C"' when no formula1 provided.
    """

    def test_dv_values_param(self, workbook_path, engine):
        result, wb = run_tasks(engine, workbook_path, [
            Task(id="t1", task_type=TaskType.DATA_VALIDATION, description="d",
                 cell="C1", sheet="Sheet1",
                 params={"type": "list", "values": ["A", "B", "C"]}),
        ])
        dv = wb.active.data_validations.dataValidation[0]
        assert dv.formula1 == '"A,B,C"'


# ═══════════════════════════════════════════════════════════════════════════
# C) TASKS THAT SILENTLY DO NOTHING
# ═══════════════════════════════════════════════════════════════════════════


class TestSilentNoOps:
    """Tasks that produce no error but have no effect."""

    def test_formatting_no_cell_no_range(self, workbook_path, engine):
        """FORMATTING with no cell/range silently skips (engine.py line 643)."""
        result, wb = run_tasks(engine, workbook_path, [
            Task(id="t1", task_type=TaskType.FORMATTING, description="d",
                 sheet="Sheet1", params={"bold": True}),
        ])
        # No error, but nothing happened — this is arguably acceptable behavior
        assert result.success

    def test_cell_value_no_cell(self, workbook_path, engine):
        """CELL_VALUE with no cell silently does nothing."""
        result, wb = run_tasks(engine, workbook_path, [
            Task(id="t1", task_type=TaskType.CELL_VALUE, description="d",
                 value="Hello", sheet="Sheet1"),
        ])
        # No error, no effect — the task "succeeds" with no write
        assert result.tasks_completed == 1  # Marked complete despite no-op


# ═══════════════════════════════════════════════════════════════════════════
# D) TASK ORDERING ISSUES
# ═══════════════════════════════════════════════════════════════════════════


class TestTaskOrdering:
    """Test that tasks in a section execute in correct order."""

    def test_data_before_formatting(self, workbook_path, engine):
        """Formatting should apply AFTER data entry."""
        result, wb = run_tasks(engine, workbook_path, [
            Task(id="t1", task_type=TaskType.CELL_VALUE, description="d",
                 cell="C1", value="Label", sheet="Sheet1"),
            Task(id="t2", task_type=TaskType.FORMATTING, description="d",
                 cell="C1", sheet="Sheet1", params={"bold": True}),
        ])
        ws = wb.active
        assert ws["C1"].value == "Label"
        assert ws["C1"].font.bold is True

    def test_data_before_table(self, workbook_path, engine):
        """Table creation should work on pre-populated data."""
        result, wb = run_tasks(engine, workbook_path, [
            Task(id="t1", task_type=TaskType.TABLE_CREATE, description="d",
                 range="A1:B10", sheet="Sheet1",
                 params={"name": "OrderTable"}),
        ])
        assert "OrderTable" in wb.active.tables


# ═══════════════════════════════════════════════════════════════════════════
# E) VERIFICATION GAPS
# ═══════════════════════════════════════════════════════════════════════════


class TestVerificationGaps:
    """Tests that expose gaps in the verifier."""

    def test_verifier_cell_value_no_op(self, workbook_path, engine):
        """Verifier marks CELL_VALUE as passed even with no cell target."""
        result, wb = run_tasks(engine, workbook_path, [
            Task(id="t1", task_type=TaskType.CELL_VALUE, description="d",
                 value="Hello", sheet="Sheet1"),
        ])
        # The task "completes" with no error but nothing was written.
        # The verifier doesn't flag this.
        assert result.tasks_completed == 1

    def test_verifier_number_format_now_passes(self, workbook_path, engine):
        """NUMBER_FORMAT on single cells now works correctly (FIXED)."""
        result, wb = run_tasks(engine, workbook_path, [
            Task(id="t1", task_type=TaskType.NUMBER_FORMAT, description="d",
                 cell="A1", sheet="Sheet1", params={"format": "$#,##0.00"}),
        ])
        assert wb.active["A1"].number_format == "$#,##0.00"
        assert result.tasks_completed == 1


# ═══════════════════════════════════════════════════════════════════════════
# F) MISSING CAPABILITIES / LAYER MAPPING ISSUES
# ═══════════════════════════════════════════════════════════════════════════


class TestMissingCapabilities:
    """TaskTypes that exist but can't actually execute via openpyxl."""

    def test_sheet_copy_in_openpyxl_map(self, workbook_path, engine):
        """
        SHEET_COPY now includes OPENPYXL in TASK_LAYER_MAP (FIXED).
        The handler at engine.py:635 uses wb.copy_worksheet().
        """
        from excel_engine.config import TASK_LAYER_MAP
        layers = TASK_LAYER_MAP[TaskType.SHEET_COPY]
        assert Layer.OPENPYXL in layers

    def test_lookup_function_in_openpyxl_map(self, workbook_path, engine):
        """
        LOOKUP_FUNCTION now includes OPENPYXL as a fallback (FIXED).
        openpyxl can write the formula text even though it can't evaluate it.
        """
        from excel_engine.config import TASK_LAYER_MAP
        layers = TASK_LAYER_MAP[TaskType.LOOKUP_FUNCTION]
        assert Layer.OPENPYXL in layers

    def test_get_layers_with_restricted_layer_order(self):
        """
        get_layers_for_task no longer crashes with ValueError when
        layer_order doesn't contain all candidate layers (FIXED).

        Fix: Filters candidates to only layers present in layer_order
        before sorting.
        """
        config = EngineConfig()
        config.layer_order = [Layer.OPENPYXL]
        layers = config.get_layers_for_task(TaskType.FORMULA)
        assert layers == [Layer.OPENPYXL]


# ═══════════════════════════════════════════════════════════════════════════
# G) EDGE CASES
# ═══════════════════════════════════════════════════════════════════════════


class TestEdgeCases:

    def test_nonexistent_sheet_handled(self, workbook_path, engine):
        """Task targeting a missing sheet should fail gracefully."""
        result, wb = run_tasks(engine, workbook_path, [
            Task(id="t1", task_type=TaskType.CELL_VALUE, description="d",
                 cell="A1", value="X", sheet="NoSuchSheet"),
        ])
        # Original data preserved
        assert wb.active["A1"].value == 10
        # Task correctly marked as failed
        assert "t1" in result.failed_tasks

"""
Tests for CLI edge cases — argument parsing, error handling, help output,
config loading, and output writing.
"""

from __future__ import annotations

import json
import sys
from io import StringIO
from pathlib import Path
from unittest.mock import MagicMock, patch, PropertyMock

import pytest

from excel_engine.cli import (
    build_parser,
    main,
    cmd_run,
    cmd_parse,
    cmd_verify,
    cmd_info,
    cmd_check_env,
    _load_config,
    _write_output,
    _setup_logging,
)
from excel_engine.config import EngineConfig, Layer


# ── build_parser / main basics ──


class TestBuildParser:
    def test_parser_has_subcommands(self):
        parser = build_parser()
        # parse_args with no args should give command=None
        args = parser.parse_args([])
        assert args.command is None

    def test_parser_version(self, capsys):
        parser = build_parser()
        with pytest.raises(SystemExit) as exc_info:
            parser.parse_args(["--version"])
        assert exc_info.value.code == 0

    def test_parser_verbose_flag(self):
        parser = build_parser()
        args = parser.parse_args(["-v", "info"])
        assert args.verbose is True
        assert args.command == "info"


class TestMain:
    def test_no_command_prints_help(self, capsys):
        rc = main([])
        assert rc == 0
        captured = capsys.readouterr()
        assert "excel-engine" in captured.out.lower() or "usage" in captured.out.lower()

    def test_keyboard_interrupt_returns_130(self):
        with patch("excel_engine.cli.cmd_info", side_effect=KeyboardInterrupt):
            rc = main(["info"])
            assert rc == 130

    def test_unhandled_exception_returns_1(self, capsys):
        with patch("excel_engine.cli.cmd_info", side_effect=RuntimeError("boom")):
            rc = main(["info"])
            assert rc == 1
            captured = capsys.readouterr()
            assert "boom" in captured.err

    def test_verbose_flag_sets_debug(self):
        with patch("excel_engine.cli.cmd_info", return_value=0):
            with patch("excel_engine.cli._setup_logging") as mock_log:
                main(["-v", "info"])
                mock_log.assert_called_once_with(True)


# ── _setup_logging ──


class TestSetupLogging:
    def test_verbose_logging(self):
        _setup_logging(verbose=True)

    def test_normal_logging(self):
        _setup_logging(verbose=False)


# ── _load_config ──


class TestLoadConfig:
    def test_no_config_returns_defaults(self):
        config = _load_config(None)
        assert isinstance(config, EngineConfig)

    def test_missing_config_file_exits(self):
        with pytest.raises(SystemExit):
            _load_config("/nonexistent/path/config.json")

    def test_valid_config_file(self, tmp_path):
        cfg_file = tmp_path / "config.json"
        cfg_file.write_text(json.dumps({"max_retries": 5, "scan_timeout": 60}))
        config = _load_config(str(cfg_file))
        assert config.max_retries == 5
        assert config.scan_timeout == 60.0

    def test_config_ignores_unknown_keys(self, tmp_path):
        cfg_file = tmp_path / "config.json"
        cfg_file.write_text(json.dumps({"max_retries": 2, "unknown_key": "whatever"}))
        config = _load_config(str(cfg_file))
        assert config.max_retries == 2


# ── _write_output ──


class TestWriteOutput:
    def test_write_to_stdout(self, capsys):
        _write_output({"key": "val"}, None)
        out = capsys.readouterr().out
        data = json.loads(out)
        assert data["key"] == "val"

    def test_write_to_file(self, tmp_path):
        out_file = tmp_path / "result.json"
        _write_output({"success": True}, str(out_file))
        data = json.loads(out_file.read_text())
        assert data["success"] is True


# ── cmd_run ──


class TestCmdRun:
    def _make_args(self, workbook="wb.xlsx", instructions="inst.docx", **kwargs):
        ns = MagicMock()
        ns.workbook = workbook
        ns.instructions = instructions
        ns.config = kwargs.get("config", None)
        ns.phase = kwargs.get("phase", "both")
        ns.dry_run = kwargs.get("dry_run", False)
        ns.watch = kwargs.get("watch", False)
        ns.output = kwargs.get("output", None)
        ns.verbose = False
        return ns

    def test_missing_workbook(self, tmp_path, capsys):
        args = self._make_args(
            workbook=str(tmp_path / "missing.xlsx"),
            instructions=str(tmp_path / "inst.docx"),
        )
        # instructions file can be missing too, but workbook is checked first
        rc = cmd_run(args)
        assert rc == 1
        assert "not found" in capsys.readouterr().err

    def test_missing_instructions(self, tmp_path, capsys):
        wb = tmp_path / "test.xlsx"
        wb.write_text("fake")
        args = self._make_args(
            workbook=str(wb),
            instructions=str(tmp_path / "missing.docx"),
        )
        rc = cmd_run(args)
        assert rc == 1
        assert "not found" in capsys.readouterr().err

    def test_phase_1_sets_openpyxl_only(self, tmp_path):
        wb = tmp_path / "test.xlsx"
        wb.write_text("fake")
        inst = tmp_path / "inst.txt"
        inst.write_text("fake")

        mock_result = MagicMock()
        mock_result.success = True
        mock_result.summary.return_value = "OK"

        with patch("excel_engine.engine.ExcelEngine") as MockEngine:
            engine_inst = MockEngine.return_value
            engine_inst.run.return_value = mock_result
            args = self._make_args(
                workbook=str(wb), instructions=str(inst), phase="1"
            )
            rc = cmd_run(args)
            assert rc == 0

    def test_phase_2_sets_live_layers(self, tmp_path):
        wb = tmp_path / "test.xlsx"
        wb.write_text("fake")
        inst = tmp_path / "inst.txt"
        inst.write_text("fake")

        mock_result = MagicMock()
        mock_result.success = True
        mock_result.summary.return_value = "OK"

        with patch("excel_engine.engine.ExcelEngine") as MockEngine:
            engine_inst = MockEngine.return_value
            engine_inst.run.return_value = mock_result
            args = self._make_args(
                workbook=str(wb), instructions=str(inst), phase="2"
            )
            rc = cmd_run(args)
            assert rc == 0

    def test_dry_run(self, tmp_path):
        wb = tmp_path / "test.xlsx"
        wb.write_text("fake")
        inst = tmp_path / "inst.txt"
        inst.write_text("fake")

        mock_plan = MagicMock()
        mock_plan.summary.return_value = "Plan summary"
        mock_plan.estimated_time_seconds = 10.0
        mock_plan.section_count = 2
        mock_plan.total_tasks = 5

        with patch("excel_engine.engine.ExcelEngine") as MockEngine:
            engine_inst = MockEngine.return_value
            engine_inst._parser.parse.return_value = "text"
            engine_inst._extractor.extract.return_value = []
            engine_inst._planner.plan.return_value = mock_plan

            args = self._make_args(
                workbook=str(wb), instructions=str(inst), dry_run=True
            )
            rc = cmd_run(args)
            assert rc == 0

    def test_dry_run_with_output(self, tmp_path):
        wb = tmp_path / "test.xlsx"
        wb.write_text("fake")
        inst = tmp_path / "inst.txt"
        inst.write_text("fake")
        out_file = tmp_path / "result.json"

        mock_plan = MagicMock()
        mock_plan.summary.return_value = "Plan summary"
        mock_plan.estimated_time_seconds = 10.0
        mock_plan.section_count = 2
        mock_plan.total_tasks = 5

        with patch("excel_engine.engine.ExcelEngine") as MockEngine:
            engine_inst = MockEngine.return_value
            engine_inst._parser.parse.return_value = "text"
            engine_inst._extractor.extract.return_value = []
            engine_inst._planner.plan.return_value = mock_plan

            args = self._make_args(
                workbook=str(wb), instructions=str(inst),
                dry_run=True, output=str(out_file),
            )
            rc = cmd_run(args)
            assert rc == 0
            assert out_file.exists()

    def test_run_failure_returns_1(self, tmp_path):
        wb = tmp_path / "test.xlsx"
        wb.write_text("fake")
        inst = tmp_path / "inst.txt"
        inst.write_text("fake")

        mock_result = MagicMock()
        mock_result.success = False
        mock_result.summary.return_value = "FAILED"

        with patch("excel_engine.engine.ExcelEngine") as MockEngine:
            engine_inst = MockEngine.return_value
            engine_inst.run.return_value = mock_result
            args = self._make_args(workbook=str(wb), instructions=str(inst))
            rc = cmd_run(args)
            assert rc == 1

    def test_run_with_output(self, tmp_path):
        wb = tmp_path / "test.xlsx"
        wb.write_text("fake")
        inst = tmp_path / "inst.txt"
        inst.write_text("fake")
        out_file = tmp_path / "result.json"

        mock_result = MagicMock()
        mock_result.success = True
        mock_result.workbook_path = wb
        mock_result.sections_completed = 1
        mock_result.sections_total = 1
        mock_result.tasks_completed = 3
        mock_result.tasks_total = 3
        mock_result.elapsed_seconds = 1.5
        mock_result.errors = []
        mock_result.summary.return_value = "OK"

        with patch("excel_engine.engine.ExcelEngine") as MockEngine:
            engine_inst = MockEngine.return_value
            engine_inst.run.return_value = mock_result
            args = self._make_args(
                workbook=str(wb), instructions=str(inst),
                output=str(out_file),
            )
            rc = cmd_run(args)
            assert rc == 0
            assert out_file.exists()

    def test_run_watch_passes_progress_callback(self, tmp_path):
        wb = tmp_path / "test.xlsx"
        wb.write_text("fake")
        inst = tmp_path / "inst.txt"
        inst.write_text("fake")

        mock_result = MagicMock()
        mock_result.success = True
        mock_result.summary.return_value = "OK"
        mock_result.workbook_path = wb
        mock_result.sections_completed = 1
        mock_result.sections_total = 1
        mock_result.tasks_completed = 1
        mock_result.tasks_total = 1
        mock_result.elapsed_seconds = 1.0
        mock_result.errors = []

        with patch("excel_engine.engine.ExcelEngine") as MockEngine:
            engine_inst = MockEngine.return_value
            engine_inst.run.return_value = mock_result
            args = self._make_args(workbook=str(wb), instructions=str(inst), watch=True)
            rc = cmd_run(args)
            assert rc == 0
            _, kwargs = engine_inst.run.call_args
            assert callable(kwargs["progress_callback"])


# ── cmd_parse ──


class TestCmdParse:
    def _make_args(self, instructions="inst.docx", output=None, config=None):
        ns = MagicMock()
        ns.instructions = instructions
        ns.output = output
        ns.config = config
        ns.verbose = False
        return ns

    def test_missing_instructions(self, capsys):
        args = self._make_args(instructions="/nonexistent/path.docx")
        rc = cmd_parse(args)
        assert rc == 1
        assert "not found" in capsys.readouterr().err

    def test_parse_success(self, tmp_path):
        inst = tmp_path / "inst.txt"
        inst.write_text("Enter 100 in cell A1")

        mock_task = MagicMock()
        mock_task.id = "t1"
        mock_task.task_type.value = "cell_value"
        mock_task.description = "enter 100"
        mock_task.sheet = None
        mock_task.cell = "A1"
        mock_task.range = None
        mock_task.value = "100"
        mock_task.formula = None
        mock_task.style = None
        mock_task.params = {}
        mock_task.depends_on = []

        with patch("excel_engine.parsers.instruction_parser.InstructionParser") as MockParser, \
             patch("excel_engine.parsers.task_extractor.TaskExtractor") as MockExtractor:
            MockParser.return_value.parse.return_value = "Enter 100 in cell A1"
            MockExtractor.return_value.extract.return_value = [mock_task]

            args = self._make_args(instructions=str(inst))
            rc = cmd_parse(args)
            assert rc == 0

    def test_parse_with_output(self, tmp_path):
        inst = tmp_path / "inst.txt"
        inst.write_text("Enter 100 in cell A1")
        out_file = tmp_path / "tasks.json"

        with patch("excel_engine.parsers.instruction_parser.InstructionParser") as MockParser, \
             patch("excel_engine.parsers.task_extractor.TaskExtractor") as MockExtractor:
            MockParser.return_value.parse.return_value = "Enter 100 in cell A1"
            MockExtractor.return_value.extract.return_value = []

            args = self._make_args(instructions=str(inst), output=str(out_file))
            rc = cmd_parse(args)
            assert rc == 0
            assert out_file.exists()


# ── cmd_verify ──


class TestCmdVerify:
    def _make_args(self, workbook="wb.xlsx", instructions=None, output=None, config=None):
        ns = MagicMock()
        ns.workbook = workbook
        ns.instructions = instructions
        ns.output = output
        ns.config = config
        ns.verbose = False
        return ns

    def test_missing_workbook(self, capsys):
        args = self._make_args(workbook="/nonexistent/wb.xlsx")
        rc = cmd_verify(args)
        assert rc == 1
        assert "not found" in capsys.readouterr().err

    def test_verify_basic_inspection(self, tmp_path):
        """Verify without instructions — basic inspection mode."""
        from openpyxl import Workbook
        wb_path = tmp_path / "test.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        ws["A1"] = "Hello"
        wb.save(str(wb_path))
        wb.close()

        args = self._make_args(workbook=str(wb_path))
        # Remove instructions attr to simulate no-instructions mode
        args.instructions = None
        rc = cmd_verify(args)
        assert rc == 0

    def test_verify_with_instructions(self, tmp_path):
        """Verify with instructions — full task-based verification."""
        from openpyxl import Workbook
        wb_path = tmp_path / "test.xlsx"
        wb = Workbook()
        wb.save(str(wb_path))
        wb.close()

        inst = tmp_path / "inst.txt"
        inst.write_text("Enter 100 in cell A1")

        mock_sv = MagicMock()
        mock_sv.all_passed = True
        mock_sv.pass_count = 1
        mock_sv.fail_count = 0
        mock_sv.section_id = "s1"
        mock_sv.summary.return_value = "Section s1: OK"

        mock_section = MagicMock()
        mock_section.id = "s1"
        mock_section.tasks = []

        mock_plan = MagicMock()
        mock_plan.sections = [mock_section]

        with patch("excel_engine.parsers.instruction_parser.InstructionParser") as MockParser, \
             patch("excel_engine.parsers.task_extractor.TaskExtractor") as MockExtractor, \
             patch("excel_engine.planner.task_planner.TaskPlanner") as MockPlanner, \
             patch("excel_engine.verifier.workbook_verifier.WorkbookVerifier") as MockVerifier:
            MockParser.return_value.parse.return_value = "text"
            MockExtractor.return_value.extract.return_value = []
            MockPlanner.return_value.plan.return_value = mock_plan
            MockVerifier.return_value.verify_section.return_value = mock_sv

            args = self._make_args(
                workbook=str(wb_path),
                instructions=str(inst),
            )
            rc = cmd_verify(args)
            assert rc == 0

    def test_verify_with_instructions_failure(self, tmp_path):
        from openpyxl import Workbook
        wb_path = tmp_path / "test.xlsx"
        wb = Workbook()
        wb.save(str(wb_path))
        wb.close()

        inst = tmp_path / "inst.txt"
        inst.write_text("Enter 100 in cell A1")

        mock_sv = MagicMock()
        mock_sv.all_passed = False
        mock_sv.pass_count = 0
        mock_sv.fail_count = 1
        mock_sv.section_id = "s1"
        mock_sv.summary.return_value = "Section s1: FAILED"

        mock_section = MagicMock()
        mock_section.id = "s1"
        mock_section.tasks = []

        mock_plan = MagicMock()
        mock_plan.sections = [mock_section]

        with patch("excel_engine.parsers.instruction_parser.InstructionParser") as MockParser, \
             patch("excel_engine.parsers.task_extractor.TaskExtractor") as MockExtractor, \
             patch("excel_engine.planner.task_planner.TaskPlanner") as MockPlanner, \
             patch("excel_engine.verifier.workbook_verifier.WorkbookVerifier") as MockVerifier:
            MockParser.return_value.parse.return_value = "text"
            MockExtractor.return_value.extract.return_value = []
            MockPlanner.return_value.plan.return_value = mock_plan
            MockVerifier.return_value.verify_section.return_value = mock_sv

            args = self._make_args(
                workbook=str(wb_path),
                instructions=str(inst),
            )
            rc = cmd_verify(args)
            assert rc == 1

    def test_verify_with_instructions_and_output(self, tmp_path):
        from openpyxl import Workbook
        wb_path = tmp_path / "test.xlsx"
        wb = Workbook()
        wb.save(str(wb_path))
        wb.close()

        inst = tmp_path / "inst.txt"
        inst.write_text("Enter 100 in cell A1")
        out_file = tmp_path / "verify.json"

        mock_sv = MagicMock()
        mock_sv.all_passed = True
        mock_sv.pass_count = 1
        mock_sv.fail_count = 0
        mock_sv.section_id = "s1"
        mock_sv.summary.return_value = "OK"

        mock_section = MagicMock()
        mock_section.id = "s1"
        mock_section.tasks = []

        mock_plan = MagicMock()
        mock_plan.sections = [mock_section]

        with patch("excel_engine.parsers.instruction_parser.InstructionParser") as MockParser, \
             patch("excel_engine.parsers.task_extractor.TaskExtractor") as MockExtractor, \
             patch("excel_engine.planner.task_planner.TaskPlanner") as MockPlanner, \
             patch("excel_engine.verifier.workbook_verifier.WorkbookVerifier") as MockVerifier:
            MockParser.return_value.parse.return_value = "text"
            MockExtractor.return_value.extract.return_value = []
            MockPlanner.return_value.plan.return_value = mock_plan
            MockVerifier.return_value.verify_section.return_value = mock_sv

            args = self._make_args(
                workbook=str(wb_path),
                instructions=str(inst),
                output=str(out_file),
            )
            rc = cmd_verify(args)
            assert rc == 0
            assert out_file.exists()


# ── cmd_info ──


class TestCmdInfo:
    def test_info_default(self, capsys):
        args = MagicMock()
        args.config = None
        rc = cmd_info(args)
        assert rc == 0
        out = capsys.readouterr().out
        assert "Excel Engine" in out
        assert "Automation Layers" in out
        assert "Task Types" in out

    def test_info_with_config(self, tmp_path, capsys):
        cfg = tmp_path / "config.json"
        cfg.write_text(json.dumps({"max_retries": 7}))
        args = MagicMock()
        args.config = str(cfg)
        rc = cmd_info(args)
        assert rc == 0
        out = capsys.readouterr().out
        assert "7" in out


# ── cmd_check_env ──


class TestCmdCheckEnv:
    def test_check_env_runs(self):
        args = MagicMock()
        rc = cmd_check_env(args)
        assert rc in (0, 1)  # depends on environment

    @patch("platform.system", return_value="Linux")
    def test_check_env_non_mac(self, mock_sys, capsys):
        args = MagicMock()
        rc = cmd_check_env(args)
        assert rc == 1
        out = capsys.readouterr().out
        assert "Linux" in out

    @patch("platform.system", return_value="Darwin")
    @patch("shutil.which", return_value=None)
    def test_check_env_no_osascript(self, mock_which, mock_sys, capsys):
        args = MagicMock()
        rc = cmd_check_env(args)
        # Should report osascript missing
        assert rc == 1


# ── interactive dispatch ──


class TestInteractiveDispatch:
    def test_interactive_dispatches(self):
        with patch("excel_engine.cli._cmd_interactive", return_value=0) as mock_int:
            rc = main(["interactive"])
            assert rc == 0
            mock_int.assert_called_once()

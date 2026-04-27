"""workbook_report — MCP resource that returns a full workbook audit as JSON.

The resource URI includes the workbook path so Copilot Studio can reference
the report as a tool output and pass it back to the agent.

URI pattern: excel-engine://workbook-report?path=<url-encoded-path>
"""

from __future__ import annotations

import asyncio
import json
import sys
from pathlib import Path
from urllib.parse import urlencode, urlparse, parse_qs

_ENGINE_ROOT = Path(__file__).resolve().parent.parent.parent
if str(_ENGINE_ROOT) not in sys.path:
    sys.path.insert(0, str(_ENGINE_ROOT))

URI = "excel-engine://workbook-report"

DEFINITION: dict = {
    "uri": URI,
    "name": "workbook_report",
    "description": (
        "Full structural audit of an Excel workbook returned as a JSON resource. "
        "Includes sheet names, table names, chart counts, named ranges, formula counts, "
        "and cell dimensions. Access by reading the resource URI returned by verify_workbook."
    ),
    "mimeType": "application/json",
}


def _resolve(p: str) -> Path:
    resolved = Path(p).expanduser().resolve()
    if not resolved.is_relative_to(Path.home()):
        raise ValueError(f"Path must be under home directory: {resolved}")
    return resolved


def _build_report_sync(workbook_path: str) -> str:
    from openpyxl import load_workbook as opx_load

    wb_path = _resolve(workbook_path)
    if not wb_path.exists():
        return json.dumps({"error": f"Workbook not found: {wb_path}"})

    oxwb = opx_load(str(wb_path), data_only=False)

    sheets: list[dict] = []
    for ws_name in oxwb.sheetnames:
        ws = oxwb[ws_name]
        formula_cells: list[str] = []
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_cells.append(cell.coordinate)

        sheets.append({
            "name": ws_name,
            "dimensions": ws.dimensions,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "tables": [t.name for t in ws.tables.values()],
            "chart_count": len(ws._charts),
            "formula_count": len(formula_cells),
            "formula_cells_sample": formula_cells[:20],
        })

    report = {
        "workbook": str(wb_path),
        "sheet_count": len(oxwb.sheetnames),
        "sheets": sheets,
        "named_ranges": [str(nr) for nr in oxwb.defined_names.definedName],
    }
    oxwb.close()
    return json.dumps(report, indent=2)


async def read(params: dict) -> dict:
    """Return a resource content dict compatible with MCP resources/read."""
    # Extract path from URI query string or from params directly
    uri = params.get("uri", "")
    parsed = urlparse(uri)
    qs = parse_qs(parsed.query)
    workbook_path = (qs.get("path") or [""])[0] or params.get("workbook_path", "")

    if not workbook_path:
        return {
            "uri": uri,
            "mimeType": "application/json",
            "text": json.dumps({"error": "workbook_path query parameter is required"}),
        }

    loop = asyncio.get_event_loop()
    report_json = await loop.run_in_executor(None, _build_report_sync, workbook_path)

    # Build a full resource URI with the path embedded
    resource_uri = f"{URI}?{urlencode({'path': workbook_path})}"

    return {
        "uri": resource_uri,
        "mimeType": "application/json",
        "text": report_json,
    }

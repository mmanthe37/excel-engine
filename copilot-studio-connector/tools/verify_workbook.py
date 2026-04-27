"""verify_workbook — verify assignment completion against expected tasks."""

from __future__ import annotations

import asyncio
import json
import sys
from pathlib import Path

_ENGINE_ROOT = Path(__file__).resolve().parent.parent.parent
if str(_ENGINE_ROOT) not in sys.path:
    sys.path.insert(0, str(_ENGINE_ROOT))

DEFINITION: dict = {
    "name": "verify_workbook",
    "description": (
        "Verify that an Excel workbook meets expected task requirements. "
        "When expected_tasks is provided, checks each task specifically and returns a "
        "pass/fail score per task. When omitted, performs a general structural audit: "
        "sheet names, table names, chart count, named ranges, and formula count. "
        "Returns a JSON verification report."
    ),
    "inputSchema": {
        "type": "object",
        "required": ["workbook_path"],
        "properties": {
            "workbook_path": {
                "type": "string",
                "description": "Absolute path to the .xlsx workbook file to verify",
            },
            "expected_tasks": {
                "type": "array",
                "description": (
                    "Optional list of task descriptions to verify specifically. "
                    "Leave empty for a general structural audit. "
                    "Example: ['Apply SUM formula in C10', 'Format header row bold']"
                ),
                "items": {"type": "string"},
            },
        },
    },
}


def _resolve(p: str) -> Path:
    resolved = Path(p).expanduser().resolve()
    if not resolved.is_relative_to(Path.home()):
        raise ValueError(f"Path must be under home directory: {resolved}")
    return resolved


def _run_sync(workbook_path: str, expected_tasks: list[str]) -> str:
    from excel_engine.verifier.workbook_verifier import WorkbookVerifier

    wb = _resolve(workbook_path)
    if not wb.exists():
        return json.dumps({"error": f"Workbook not found: {wb}"})

    verifier = WorkbookVerifier()
    verifier.load(wb)

    if expected_tasks:
        from excel_engine.parsers.task_extractor import TaskExtractor

        extractor = TaskExtractor()
        task_text = "\n".join(f"- {t}" for t in expected_tasks)
        task_objs = extractor.extract(task_text)
        verification = verifier.verify_section("copilot-studio-check", task_objs)
        verifier.close()

        total = len(verification.results)
        return json.dumps(
            {
                "workbook": str(wb),
                "section_id": verification.section_id,
                "total_tasks": total,
                "passed": verification.pass_count,
                "failed": verification.fail_count,
                "all_passed": verification.all_passed,
                "score": f"{verification.pass_count}/{total}",
                "results": [
                    {
                        "task_id": r.task_id,
                        "task_type": r.task_type.value,
                        "passed": r.passed,
                        "message": r.message,
                    }
                    for r in verification.results
                ],
            },
            indent=2,
        )

    # General structural audit
    from openpyxl import load_workbook as opx_load

    oxwb = opx_load(str(wb), data_only=False)
    report: dict = {
        "workbook": str(wb),
        "sheets": oxwb.sheetnames,
        "sheet_count": len(oxwb.sheetnames),
        "tables": {},
        "charts": {},
        "named_ranges": [str(nr) for nr in oxwb.defined_names.definedName],
        "formulas_found": 0,
    }
    for ws_name in oxwb.sheetnames:
        ws = oxwb[ws_name]
        report["tables"][ws_name] = [t.name for t in ws.tables.values()]
        report["charts"][ws_name] = len(ws._charts)
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    report["formulas_found"] += 1

    oxwb.close()
    verifier.close()
    return json.dumps(report, indent=2)


async def run(arguments: dict) -> str:
    workbook_path = arguments.get("workbook_path", "").strip()
    expected_tasks = arguments.get("expected_tasks") or []

    if not workbook_path:
        raise ValueError("workbook_path is required")

    loop = asyncio.get_event_loop()
    return await loop.run_in_executor(None, _run_sync, workbook_path, list(expected_tasks))

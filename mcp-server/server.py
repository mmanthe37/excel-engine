"""
Excel Engine MCP Server — Model Context Protocol interface for autonomous
Excel automation on macOS.

Exposes the 6-layer Excel Engine via MCP tools so that AI assistants
(GitHub Copilot CLI, Claude Desktop, etc.) can drive spreadsheet
assignments through natural-language tool calls.

Run:
    python server.py
    # or
    python -m mcp_server
"""

from __future__ import annotations

import json
import logging
import sys
from pathlib import Path
from typing import Any, Optional

from mcp.server.fastmcp import FastMCP

# ── Bootstrap the excel_engine package from the parent directory ──
_ENGINE_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_ENGINE_ROOT))

from excel_engine import ExcelEngine, EngineConfig, __version__ as engine_version
from excel_engine.config import Layer, TaskType, TASK_LAYER_MAP
from excel_engine.parsers.instruction_parser import InstructionParser
from excel_engine.parsers.task_extractor import Task, TaskExtractor
from excel_engine.verifier.workbook_verifier import WorkbookVerifier

# ── Logging ──
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
)
logger = logging.getLogger("excel-engine-mcp")

# ── MCP Server ──
mcp = FastMCP(
    "excel-engine",
    instructions=(
        "Autonomous Excel automation engine for macOS. "
        "Completes spreadsheet assignments using a 6-layer architecture: "
        "openpyxl → xlwings → AppleScript → System Events → VBA → PyAutoGUI."
    ),
)


# ─────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────

def _resolve_path(p: str) -> Path:
    """Expand ~ and resolve to absolute path, with boundary validation."""
    resolved = Path(p).expanduser().resolve()
    _ALLOWED_DIRS = [Path.home()]
    if not any(resolved.is_relative_to(d) for d in _ALLOWED_DIRS):
        raise ValueError(f"Path must be under home directory: {resolved}")
    return resolved


def _engine_result_to_dict(result) -> dict[str, Any]:
    """Convert an EngineResult dataclass to a JSON-safe dict."""
    verifications = []
    for v in result.verifications:
        verifications.append({
            "section_id": v.section_id,
            "all_passed": v.all_passed,
            "pass_count": v.pass_count,
            "fail_count": v.fail_count,
            "results": [
                {
                    "task_id": r.task_id,
                    "task_type": r.task_type.value,
                    "passed": r.passed,
                    "message": r.message,
                }
                for r in v.results
            ],
        })
    return {
        "success": result.success,
        "workbook": str(result.workbook_path),
        "sections_completed": result.sections_completed,
        "sections_total": result.sections_total,
        "tasks_completed": result.tasks_completed,
        "tasks_total": result.tasks_total,
        "elapsed_seconds": round(result.elapsed_seconds, 2),
        "errors": result.errors,
        "verifications": verifications,
        "summary": result.summary(),
    }


def _make_progress_callback():
    """Create a progress callback that logs task progress via the MCP logger."""
    def callback(info: dict) -> None:
        phase = info.get("phase", "unknown")
        task_id = info.get("task", "?")
        total = info.get("total", "?")
        if phase == "executing":
            logger.info("[progress] Executing task %s (of %s)", task_id, total)
        elif phase == "completed":
            success = info.get("success", False)
            status = "✓" if success else "✗"
            logger.info("[progress] %s Task %s (of %s)", status, task_id, total)
        else:
            logger.info("[progress] %s — task %s", phase, task_id)
    return callback


def _task_to_dict(t: Task) -> dict[str, Any]:
    """Convert a Task dataclass to a JSON-safe dict."""
    return {
        "id": t.id,
        "task_type": t.task_type.value,
        "description": t.description,
        "sheet": t.sheet,
        "cell": t.cell,
        "range": t.range,
        "value": t.value,
        "formula": t.formula,
        "style": t.style,
        "params": t.params,
        "depends_on": t.depends_on,
    }


# ─────────────────────────────────────────────────────────────────────
# Tool 1: complete_assignment
# ─────────────────────────────────────────────────────────────────────

@mcp.tool()
def complete_assignment(
    workbook_path: str,
    instruction_path: str,
    options: Optional[dict] = None,
) -> str:
    """Complete an Excel assignment autonomously.

    Parses the instruction file, plans execution across 6 automation layers,
    executes all tasks, and verifies the results.

    Args:
        workbook_path: Path to the .xlsx workbook file.
        instruction_path: Path to the instruction file (.docx, .rtfd, .pdf, .txt).
        options: Optional engine config overrides (max_retries, verify_after_each_section, etc.).

    Returns:
        JSON report with success status, task counts, verification results, and timing.
    """
    try:
        wb = _resolve_path(workbook_path)
        inst = _resolve_path(instruction_path)

        if not wb.exists():
            return json.dumps({"error": f"Workbook not found: {wb}"})
        if not inst.exists() and not inst.is_dir():
            return json.dumps({"error": f"Instruction file not found: {inst}"})

        config = EngineConfig()
        _SAFE_OPTIONS = {"max_retries", "verify_after_each_section", "retina_display", "scan_timeout"}
        if options:
            for key, val in options.items():
                if key in _SAFE_OPTIONS and hasattr(config, key):
                    setattr(config, key, val)

        logger.info("Starting assignment: %s with %s", wb.name, inst.name)
        engine = ExcelEngine(config=config)
        progress_cb = _make_progress_callback()
        result = engine.run(workbook=wb, instructions=inst, progress_callback=progress_cb)

        return json.dumps(_engine_result_to_dict(result), indent=2)

    except Exception as exc:
        logger.exception("complete_assignment failed")
        return json.dumps({"error": f"Internal error: {type(exc).__name__}"})


# ─────────────────────────────────────────────────────────────────────
# Tool 2: parse_instructions
# ─────────────────────────────────────────────────────────────────────

@mcp.tool()
def parse_instructions(instruction_path: str) -> str:
    """Parse an instruction file into structured tasks.

    Reads a .docx, .rtfd, .pdf, or .txt instruction file, extracts the raw
    text, then identifies individual Excel operations (formulas, tables,
    formatting, charts, etc.) as structured Task objects.

    Args:
        instruction_path: Path to the instruction file.

    Returns:
        JSON with raw_text (truncated), task_count, and a list of task objects.
    """
    try:
        inst = _resolve_path(instruction_path)
        if not inst.exists() and not inst.is_dir():
            return json.dumps({"error": f"File not found: {inst}"})

        parser = InstructionParser()
        text = parser.parse(inst)

        extractor = TaskExtractor()
        tasks = extractor.extract(text)

        return json.dumps({
            "instruction_file": str(inst),
            "raw_text_preview": text[:500] + ("..." if len(text) > 500 else ""),
            "raw_text_length": len(text),
            "task_count": len(tasks),
            "tasks": [_task_to_dict(t) for t in tasks],
        }, indent=2)

    except Exception as exc:
        logger.exception("parse_instructions failed")
        return json.dumps({"error": f"Internal error: {type(exc).__name__}"})


# ─────────────────────────────────────────────────────────────────────
# Tool 3: execute_openpyxl
# ─────────────────────────────────────────────────────────────────────

@mcp.tool()
def execute_openpyxl(workbook_path: str, tasks: list[str]) -> str:
    """Run Phase 1 offline operations via the openpyxl layer.

    Executes a list of task descriptions against the workbook using only
    the openpyxl layer (Layer 1). This works entirely offline — Excel does
    not need to be open.

    Args:
        workbook_path: Path to the .xlsx workbook file.
        tasks: List of task description strings (natural language or structured).

    Returns:
        JSON report of executed tasks and any errors.
    """
    try:
        wb = _resolve_path(workbook_path)
        if not wb.exists():
            return json.dumps({"error": f"Workbook not found: {wb}"})

        if len(tasks) > 500:
            return json.dumps({"error": "Too many tasks (max 500)"})
        if any(len(t) > 10000 for t in tasks):
            return json.dumps({"error": "Task too long (max 10000 chars)"})

        extractor = TaskExtractor()
        task_text = "\n".join(f"- {t}" for t in tasks)
        task_objs = extractor.extract(task_text)

        config = EngineConfig()
        config.layer_order = [Layer.OPENPYXL]
        engine = ExcelEngine(config=config)

        logger.info("Executing %d openpyxl tasks on %s", len(task_objs), wb.name)
        result = engine.run(workbook=wb, tasks=task_objs)

        return json.dumps(_engine_result_to_dict(result), indent=2)

    except Exception as exc:
        logger.exception("execute_openpyxl failed")
        return json.dumps({"error": f"Internal error: {type(exc).__name__}"})


# ─────────────────────────────────────────────────────────────────────
# Tool 4: execute_live
# ─────────────────────────────────────────────────────────────────────

@mcp.tool()
def execute_live(workbook_path: str, tasks: list[str]) -> str:
    """Run Phase 2 live Excel operations.

    Executes tasks that require Excel to be open: xlwings, AppleScript,
    System Events, VBA, or PyAutoGUI (Layers 2-6). Use this for operations
    like sorting, slicers, pivot tables, or anything that needs the live app.

    Args:
        workbook_path: Path to the .xlsx workbook file.
        tasks: List of task description strings (natural language or structured).

    Returns:
        JSON report of executed tasks and any errors.
    """
    try:
        wb = _resolve_path(workbook_path)
        if not wb.exists():
            return json.dumps({"error": f"Workbook not found: {wb}"})

        if len(tasks) > 500:
            return json.dumps({"error": "Too many tasks (max 500)"})
        if any(len(t) > 10000 for t in tasks):
            return json.dumps({"error": "Task too long (max 10000 chars)"})

        extractor = TaskExtractor()
        task_text = "\n".join(f"- {t}" for t in tasks)
        task_objs = extractor.extract(task_text)

        config = EngineConfig()
        config.layer_order = [
            Layer.XLWINGS,
            Layer.APPLESCRIPT,
            Layer.SYSTEM_EVENTS,
            Layer.VBA,
            Layer.PYAUTOGUI,
        ]
        engine = ExcelEngine(config=config)

        logger.info("Executing %d live tasks on %s", len(task_objs), wb.name)
        result = engine.run(workbook=wb, tasks=task_objs)

        return json.dumps(_engine_result_to_dict(result), indent=2)

    except Exception as exc:
        logger.exception("execute_live failed")
        return json.dumps({"error": f"Internal error: {type(exc).__name__}"})


# ─────────────────────────────────────────────────────────────────────
# Tool 5: verify_workbook
# ─────────────────────────────────────────────────────────────────────

@mcp.tool()
def verify_workbook(
    workbook_path: str,
    expected_tasks: Optional[list[str]] = None,
) -> str:
    """Verify assignment completion against expected tasks.

    Opens the workbook and checks whether each expected task was completed.
    If no expected_tasks are given, performs a general structural check
    (sheets, tables, charts, named ranges, formulas).

    Args:
        workbook_path: Path to the .xlsx workbook file.
        expected_tasks: Optional list of task descriptions to verify.

    Returns:
        JSON verification report with pass/fail per task and overall score.
    """
    try:
        wb = _resolve_path(workbook_path)
        if not wb.exists():
            return json.dumps({"error": f"Workbook not found: {wb}"})

        verifier = WorkbookVerifier()
        verifier.load(wb)

        if expected_tasks:
            extractor = TaskExtractor()
            task_text = "\n".join(f"- {t}" for t in expected_tasks)
            task_objs = extractor.extract(task_text)

            verification = verifier.verify_section("manual-check", task_objs)
            verifier.close()

            total = len(verification.results)
            return json.dumps({
                "workbook": str(wb),
                "section_id": verification.section_id,
                "total_tasks": total,
                "passed": verification.pass_count,
                "failed": verification.fail_count,
                "all_passed": verification.all_passed,
                "score": f"{verification.pass_count}/{total}",
                "results": [
                    {
                        "task_id": r.task_id,
                        "task_type": r.task_type.value,
                        "passed": r.passed,
                        "message": r.message,
                    }
                    for r in verification.results
                ],
            }, indent=2)
        else:
            # General structural audit
            from openpyxl import load_workbook as opx_load

            oxwb = opx_load(str(wb), data_only=False)
            report = {
                "workbook": str(wb),
                "sheets": oxwb.sheetnames,
                "sheet_count": len(oxwb.sheetnames),
                "tables": {},
                "charts": {},
                "named_ranges": [str(nr) for nr in oxwb.defined_names.definedName],
                "formulas_found": 0,
            }

            for ws_name in oxwb.sheetnames:
                ws = oxwb[ws_name]
                report["tables"][ws_name] = [t.name for t in ws.tables.values()]
                report["charts"][ws_name] = len(ws._charts)
                for row in ws.iter_rows():
                    for cell in row:
                        if isinstance(cell.value, str) and cell.value.startswith("="):
                            report["formulas_found"] += 1

            oxwb.close()
            verifier.close()
            return json.dumps(report, indent=2)

    except Exception as exc:
        logger.exception("verify_workbook failed")
        return json.dumps({"error": f"Internal error: {type(exc).__name__}"})


# ─────────────────────────────────────────────────────────────────────
# Tool 6: get_engine_status
# ─────────────────────────────────────────────────────────────────────

@mcp.tool()
def get_engine_status() -> str:
    """Get current engine configuration, version, and available layers.

    Returns:
        JSON with engine version, available layers, supported task types,
        and current configuration.
    """
    config = EngineConfig()
    return json.dumps({
        "engine_version": engine_version,
        "mcp_server_version": "1.1.0",
        "available_layers": [
            {"number": layer.value, "name": layer.name}
            for layer in Layer
        ],
        "supported_task_types": [tt.value for tt in TaskType],
        "task_layer_mapping": {
            tt.value: [l.name for l in layers]
            for tt, layers in TASK_LAYER_MAP.items()
        },
        "config": {
            "scan_timeout": config.scan_timeout,
            "section_timeout": config.section_timeout,
            "applescript_timeout": config.applescript_timeout,
            "max_retries": config.max_retries,
            "retry_delay": config.retry_delay,
            "layer_order": [l.name for l in config.layer_order],
            "verify_after_each_section": config.verify_after_each_section,
            "sam_fingerprint_protected": config.sam_fingerprint_protected,
            "retina_display": config.retina_display,
            "vba_split_threshold": config.vba_split_threshold,
            "parallel_execution": getattr(config, "parallel_execution", False),
            "max_workers": getattr(config, "max_workers", 4),
            "circuit_breaker_threshold": getattr(config, "circuit_breaker_threshold", 5),
            "circuit_breaker_reset_seconds": getattr(config, "circuit_breaker_reset_seconds", 300),
            "recalculate_formulas": getattr(config, "recalculate_formulas", False),
            "recalc_timeout": getattr(config, "recalc_timeout", 60),
        },
    }, indent=2)


# ─────────────────────────────────────────────────────────────────────
# Entrypoint
# ─────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    mcp.run(transport="stdio")

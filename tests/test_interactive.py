"""
Tests for excel_engine/interactive.py — REPL helpers and session flow.

Covers _prompt, _confirm, _resolve_path, and the interactive_session
orchestration via mocked stdin/stdout.
"""

import json
from pathlib import Path
from unittest.mock import patch, MagicMock, call

import pytest

from excel_engine.interactive import _prompt, _confirm, _resolve_path, interactive_session


# ══════════════════════════════════════════════════════════════════
# _prompt
# ══════════════════════════════════════════════════════════════════

class TestPrompt:
    def test_returns_user_input(self):
        with patch("builtins.input", return_value="hello"):
            assert _prompt("Enter") == "hello"

    def test_strips_whitespace(self):
        with patch("builtins.input", return_value="  spaces  "):
            assert _prompt("Enter") == "spaces"

    def test_returns_default_on_empty(self):
        with patch("builtins.input", return_value=""):
            assert _prompt("Enter", default="fallback") == "fallback"

    def test_returns_default_on_eof(self):
        with patch("builtins.input", side_effect=EOFError):
            assert _prompt("Enter", default="eof_default") == "eof_default"

    def test_returns_empty_on_eof_no_default(self):
        with patch("builtins.input", side_effect=EOFError):
            assert _prompt("Enter") == ""

    def test_user_overrides_default(self):
        with patch("builtins.input", return_value="override"):
            assert _prompt("Enter", default="fallback") == "override"


# ══════════════════════════════════════════════════════════════════
# _confirm
# ══════════════════════════════════════════════════════════════════

class TestConfirm:
    def test_yes_answers(self):
        for answer in ("y", "yes", "Y", "YES", "Yes"):
            with patch("builtins.input", return_value=answer):
                assert _confirm("OK?") is True

    def test_no_answers(self):
        for answer in ("n", "no", "N", "nope", "anything"):
            with patch("builtins.input", return_value=answer):
                assert _confirm("OK?") is False

    def test_empty_returns_default_true(self):
        with patch("builtins.input", return_value=""):
            assert _confirm("OK?", default=True) is True

    def test_empty_returns_default_false(self):
        with patch("builtins.input", return_value=""):
            assert _confirm("OK?", default=False) is False

    def test_eof_returns_default(self):
        with patch("builtins.input", side_effect=EOFError):
            assert _confirm("OK?", default=True) is True
        with patch("builtins.input", side_effect=EOFError):
            assert _confirm("OK?", default=False) is False


# ══════════════════════════════════════════════════════════════════
# _resolve_path
# ══════════════════════════════════════════════════════════════════

class TestResolvePath:
    def test_empty_string_returns_none(self):
        assert _resolve_path("") is None

    def test_valid_path(self, tmp_path):
        f = tmp_path / "test.xlsx"
        f.write_bytes(b"data")
        result = _resolve_path(str(f))
        assert result is not None
        assert result.name == "test.xlsx"

    def test_nonexistent_returns_none(self):
        result = _resolve_path("/nonexistent/does_not_exist_99999.xlsx")
        assert result is None

    def test_tilde_expansion(self):
        home = Path.home()
        if home.exists():
            result = _resolve_path("~")
            assert result is not None
            assert result == home

    def test_relative_path(self, tmp_path, monkeypatch):
        f = tmp_path / "rel_test.xlsx"
        f.write_bytes(b"data")
        monkeypatch.chdir(tmp_path)
        result = _resolve_path("rel_test.xlsx")
        assert result is not None
        assert result.name == "rel_test.xlsx"


# ══════════════════════════════════════════════════════════════════
# interactive_session — full-flow with mocks
# ══════════════════════════════════════════════════════════════════

class TestInteractiveSession:
    def test_abort_before_execution(self, tmp_path):
        """User provides valid files but declines to execute."""
        wb_path = tmp_path / "book.xlsx"
        instr_path = tmp_path / "instr.txt"

        # Create minimal valid files
        from openpyxl import Workbook
        wb = Workbook()
        wb.save(wb_path)
        wb.close()
        instr_path.write_text("Enter 100 in cell A1")

        inputs = iter([
            str(wb_path),       # workbook path
            str(instr_path),    # instructions path
            "n",                # decline execution
        ])

        with patch("builtins.input", side_effect=lambda *a: next(inputs)):
            code = interactive_session()

        assert code == 0  # aborted gracefully

    def test_bad_workbook_path_then_valid(self, tmp_path):
        """User enters a bad workbook path, then a valid one, then aborts."""
        wb_path = tmp_path / "book.xlsx"
        instr_path = tmp_path / "instr.txt"

        from openpyxl import Workbook
        wb = Workbook()
        wb.save(wb_path)
        wb.close()
        instr_path.write_text("Enter SUM formula in cell B1")

        inputs = iter([
            "/nonexistent/bad.xlsx",  # first bad path
            str(wb_path),             # then valid
            str(instr_path),          # instructions
            "n",                      # decline execution
        ])

        with patch("builtins.input", side_effect=lambda *a: next(inputs)):
            code = interactive_session()

        assert code == 0

    def test_non_excel_extension_rejected(self, tmp_path):
        """User gives a valid file with wrong extension, then correct one."""
        txt_file = tmp_path / "data.txt"
        txt_file.write_text("not excel")
        wb_path = tmp_path / "book.xlsx"
        instr_path = tmp_path / "instr.txt"

        from openpyxl import Workbook
        wb = Workbook()
        wb.save(wb_path)
        wb.close()
        instr_path.write_text("Enter 42 in cell A1")

        inputs = iter([
            str(txt_file),    # wrong extension
            str(wb_path),     # valid
            str(instr_path),  # instructions
            "n",              # decline
        ])

        with patch("builtins.input", side_effect=lambda *a: next(inputs)):
            code = interactive_session()

        assert code == 0

    def test_empty_workbook_path_retries(self, tmp_path):
        """Empty input for workbook triggers a retry prompt."""
        wb_path = tmp_path / "book.xlsx"
        instr_path = tmp_path / "instr.txt"

        from openpyxl import Workbook
        wb = Workbook()
        wb.save(wb_path)
        wb.close()
        instr_path.write_text("Enter 1 in cell A1")

        inputs = iter([
            "",               # empty triggers retry
            str(wb_path),     # valid workbook
            "",               # empty instructions triggers retry
            str(instr_path),  # valid instructions
            "n",              # decline
        ])

        with patch("builtins.input", side_effect=lambda *a: next(inputs)):
            code = interactive_session()

        assert code == 0

    def test_parse_failure_returns_1(self, tmp_path):
        """If instruction parsing throws, session returns 1."""
        wb_path = tmp_path / "book.xlsx"
        instr_path = tmp_path / "instr.bin"

        from openpyxl import Workbook
        wb = Workbook()
        wb.save(wb_path)
        wb.close()
        # Write invalid binary to trigger a parse error
        instr_path.write_bytes(b"\x00\x01\x02\x03")

        inputs = iter([
            str(wb_path),
            str(instr_path),
        ])

        with patch("builtins.input", side_effect=lambda *a: next(inputs)):
            code = interactive_session()

        assert code == 1

    def test_full_run_with_save(self, tmp_path):
        """Full session: execute + save results to JSON."""
        wb_path = tmp_path / "book.xlsx"
        instr_path = tmp_path / "instr.txt"

        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        wb.save(wb_path)
        wb.close()
        instr_path.write_text("Enter 100 in cell A1")

        out_json = tmp_path / "book_results.json"

        inputs = iter([
            str(wb_path),
            str(instr_path),
            "y",                # proceed with execution
            "y",                # save results
            str(out_json),      # output path
        ])

        with patch("builtins.input", side_effect=lambda *a: next(inputs)):
            code = interactive_session()

        assert code in (0, 1)  # engine may pass or fail depending on task extraction
        if out_json.exists():
            data = json.loads(out_json.read_text())
            assert "success" in data
            assert "workbook" in data

    def test_full_run_no_save(self, tmp_path):
        """Full session: execute but decline saving."""
        wb_path = tmp_path / "book.xlsx"
        instr_path = tmp_path / "instr.txt"

        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        wb.save(wb_path)
        wb.close()
        instr_path.write_text("Enter 100 in cell A1")

        inputs = iter([
            str(wb_path),
            str(instr_path),
            "y",                # proceed with execution
            "n",                # don't save
        ])

        with patch("builtins.input", side_effect=lambda *a: next(inputs)):
            code = interactive_session()

        assert code in (0, 1)

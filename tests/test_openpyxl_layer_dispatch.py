"""
Tests for OpenpyxlLayer — task dispatch, edge cases, and method coverage.
"""

from __future__ import annotations

from pathlib import Path
from unittest.mock import MagicMock, patch

import pytest
from openpyxl import Workbook

from excel_engine.layers.openpyxl_layer import OpenpyxlLayer


@pytest.fixture
def layer(tmp_path):
    """Create an OpenpyxlLayer with a fresh workbook."""
    lyr = OpenpyxlLayer()
    wb_path = tmp_path / "test.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    # Populate some data for chart/table tests
    ws["A1"] = "Name"
    ws["B1"] = "Value"
    for i in range(2, 12):
        ws[f"A{i}"] = f"Item{i}"
        ws[f"B{i}"] = i * 10
    wb.save(str(wb_path))
    wb.close()

    lyr.open(wb_path)
    return lyr


@pytest.fixture
def empty_layer():
    """An OpenpyxlLayer with no workbook open."""
    return OpenpyxlLayer()


# ── Lifecycle ──


class TestLifecycle:
    def test_open_and_close(self, layer):
        assert layer.wb is not None
        layer.close()
        assert layer.wb is None

    def test_create_new_workbook(self, tmp_path):
        lyr = OpenpyxlLayer()
        path = tmp_path / "new.xlsx"
        lyr.create(path)
        assert lyr.wb is not None
        lyr.save()
        assert path.exists()
        lyr.close()

    def test_save_to_alternate_path(self, layer, tmp_path):
        alt = tmp_path / "alt.xlsx"
        layer.save(alt)
        assert alt.exists()

    def test_save_no_workbook_raises(self, empty_layer):
        with pytest.raises(ValueError, match="No workbook"):
            empty_layer.save()

    def test_close_idempotent(self, empty_layer):
        empty_layer.close()
        empty_layer.close()  # should not raise

    def test_open_xlsm(self, tmp_path):
        lyr = OpenpyxlLayer()
        path = tmp_path / "macro.xlsm"
        wb = Workbook()
        wb.save(str(path))
        wb.close()
        lyr.open(path)
        assert lyr.wb is not None
        lyr.close()


# ── _ws helper ──


class TestWsHelper:
    def test_no_workbook_raises(self, empty_layer):
        with pytest.raises(RuntimeError, match="No workbook"):
            empty_layer._ws()

    def test_get_active_sheet(self, layer):
        ws = layer._ws()
        assert ws is not None

    def test_get_named_sheet(self, layer):
        ws = layer._ws("Sheet1")
        assert ws.title == "Sheet1"

    def test_get_nonexistent_sheet_raises(self, layer):
        with pytest.raises(KeyError):
            layer._ws("NonExistent")


# ── _qualify_range ──


class TestQualifyRange:
    def test_already_qualified(self, layer):
        ws = layer._ws()
        result = OpenpyxlLayer._qualify_range(ws, "Sheet1!A1:B10")
        assert result == "Sheet1!A1:B10"

    def test_unqualified(self, layer):
        ws = layer._ws()
        result = OpenpyxlLayer._qualify_range(ws, "A1:B10")
        assert "!" in result
        assert "A1:B10" in result

    def test_sheet_with_space(self, layer):
        layer.create_sheet("My Sheet")
        ws = layer._ws("My Sheet")
        result = OpenpyxlLayer._qualify_range(ws, "A1:B10")
        assert "'My Sheet'!A1:B10" == result


# ── Sheet Management ──


class TestSheetManagement:
    def test_create_sheet(self, layer):
        layer.create_sheet("NewSheet")
        assert "NewSheet" in layer.wb.sheetnames

    def test_create_sheet_at_index(self, layer):
        layer.create_sheet("First", index=0)
        assert layer.wb.sheetnames[0] == "First"

    def test_create_sheet_no_workbook(self, empty_layer):
        with pytest.raises(RuntimeError, match="No workbook"):
            empty_layer.create_sheet("Test")

    def test_rename_sheet(self, layer):
        layer.rename_sheet("Sheet1", "Renamed")
        assert "Renamed" in layer.wb.sheetnames
        assert "Sheet1" not in layer.wb.sheetnames

    def test_delete_sheet(self, layer):
        layer.create_sheet("ToDelete")
        assert "ToDelete" in layer.wb.sheetnames
        layer.delete_sheet("ToDelete")
        assert "ToDelete" not in layer.wb.sheetnames

    def test_delete_sheet_no_workbook(self, empty_layer):
        with pytest.raises(RuntimeError, match="No workbook"):
            empty_layer.delete_sheet("Missing")


# ── Cell Operations ──


class TestCellOperations:
    def test_set_and_get_value(self, layer):
        layer.set_value("C1", 42, sheet="Sheet1")
        assert layer.get_value("C1", sheet="Sheet1") == 42

    def test_set_formula(self, layer):
        layer.set_formula("C1", "SUM(B2:B11)", sheet="Sheet1")
        val = layer.get_value("C1", sheet="Sheet1")
        assert val == "=SUM(B2:B11)"

    def test_set_formula_with_equals(self, layer):
        layer.set_formula("C2", "=SUM(B2:B3)", sheet="Sheet1")
        val = layer.get_value("C2", sheet="Sheet1")
        assert val == "=SUM(B2:B3)"

    def test_set_values_range(self, layer):
        layer.set_values_range("D1", [["a", "b"], ["c", "d"]], sheet="Sheet1")
        assert layer.get_value("D1", sheet="Sheet1") == "a"
        assert layer.get_value("E1", sheet="Sheet1") == "b"
        assert layer.get_value("D2", sheet="Sheet1") == "c"
        assert layer.get_value("E2", sheet="Sheet1") == "d"


# ── Tables ──


class TestTables:
    def test_create_table(self, layer):
        layer.create_table("TestTable", "A1:B11", sheet="Sheet1")
        ws = layer._ws("Sheet1")
        assert len(ws.tables) == 1

    def test_create_table_with_totals(self, layer):
        layer.create_table("TotalTable", "A1:B11", sheet="Sheet1", show_totals=True)
        ws = layer._ws("Sheet1")
        table = list(ws.tables.values())[0]
        assert table.totalsRowShown is True

    def test_create_table_custom_style(self, layer):
        layer.create_table("Styled", "A1:B11", style="TableStyleLight1", sheet="Sheet1")
        ws = layer._ws("Sheet1")
        table = list(ws.tables.values())[0]
        assert table.tableStyleInfo.name == "TableStyleLight1"


# ── Formatting ──


class TestFormatting:
    def test_set_font_single_cell(self, layer):
        layer.set_font("A1", sheet="Sheet1", bold=True, size=14)
        ws = layer._ws("Sheet1")
        assert ws["A1"].font.bold is True
        assert ws["A1"].font.size == 14

    def test_set_font_range(self, layer):
        layer.set_font("A1:B2", sheet="Sheet1", italic=True)
        ws = layer._ws("Sheet1")
        assert ws["A1"].font.italic is True
        assert ws["B2"].font.italic is True

    def test_set_fill_single_cell(self, layer):
        layer.set_fill("A1", "FF0000", sheet="Sheet1")
        ws = layer._ws("Sheet1")
        assert ws["A1"].fill.start_color.rgb == "00FF0000"

    def test_set_fill_range(self, layer):
        layer.set_fill("A1:A3", "00FF00", sheet="Sheet1")
        ws = layer._ws("Sheet1")
        assert ws["A1"].fill.start_color is not None

    def test_set_alignment_single_cell(self, layer):
        layer.set_alignment("A1", sheet="Sheet1", horizontal="center", wrap_text=True)
        ws = layer._ws("Sheet1")
        assert ws["A1"].alignment.horizontal == "center"
        assert ws["A1"].alignment.wrap_text is True

    def test_set_alignment_range(self, layer):
        layer.set_alignment("A1:B2", sheet="Sheet1", vertical="top")

    def test_set_number_format_single_cell(self, layer):
        layer.set_number_format("B2", "$#,##0.00", sheet="Sheet1")
        ws = layer._ws("Sheet1")
        assert ws["B2"].number_format == "$#,##0.00"

    def test_set_number_format_range(self, layer):
        layer.set_number_format("B2:B5", "0.00%", sheet="Sheet1")

    def test_set_border_single_cell(self, layer):
        layer.set_border("A1", sheet="Sheet1", style="thick")
        ws = layer._ws("Sheet1")
        assert ws["A1"].border.left.style == "thick"

    def test_set_border_range(self, layer):
        layer.set_border("A1:B2", sheet="Sheet1")

    def test_set_column_width(self, layer):
        layer.set_column_width("A", 25, sheet="Sheet1")
        ws = layer._ws("Sheet1")
        assert ws.column_dimensions["A"].width == 25

    def test_set_row_height(self, layer):
        layer.set_row_height(1, 30, sheet="Sheet1")
        ws = layer._ws("Sheet1")
        assert ws.row_dimensions[1].height == 30

    def test_auto_column_widths(self, layer):
        layer.auto_column_widths(sheet="Sheet1")
        ws = layer._ws("Sheet1")
        # Column A has "Name" (4 chars) → width should be > 0
        assert ws.column_dimensions["A"].width > 0


# ── Conditional Formatting ──


class TestConditionalFormatting:
    def test_cell_is_rule(self, layer):
        layer.add_conditional_format_cell_is(
            "B2:B11", "greaterThan", ["50"], sheet="Sheet1"
        )

    def test_formula_rule(self, layer):
        layer.add_conditional_format_formula(
            "A2:A11", '$B2>50', sheet="Sheet1"
        )

    def test_color_scale_3_color(self, layer):
        layer.add_color_scale("B2:B11", sheet="Sheet1")

    def test_color_scale_2_color(self, layer):
        layer.add_color_scale(
            "B2:B11", mid_color=None, sheet="Sheet1"
        )


# ── Charts ──


class TestCharts:
    def test_bar_chart(self, layer):
        layer.add_bar_chart(sheet="Sheet1", title="Bar", data_range="B1:B11")

    def test_bar_chart_with_cats(self, layer):
        layer.add_bar_chart(
            sheet="Sheet1", title="Bar+Cats",
            data_range="B1:B11", cats_range="A1:A11",
        )

    def test_line_chart(self, layer):
        layer.add_line_chart(sheet="Sheet1", title="Line", data_range="B1:B11")

    def test_line_chart_with_cats(self, layer):
        layer.add_line_chart(
            sheet="Sheet1", title="Line+Cats",
            data_range="B1:B11", cats_range="A1:A11",
        )

    def test_pie_chart(self, layer):
        layer.add_pie_chart(sheet="Sheet1", title="Pie", data_range="B1:B11")

    def test_pie_chart_with_cats(self, layer):
        layer.add_pie_chart(
            sheet="Sheet1", title="Pie+Cats",
            data_range="B1:B11", cats_range="A1:A11",
        )

    def test_scatter_chart(self, layer):
        layer.add_scatter_chart(
            sheet="Sheet1", title="Scatter",
            x_range="A2:A11", y_range="B2:B11",
        )

    def test_scatter_chart_marker_style(self, layer):
        layer.add_scatter_chart(
            sheet="Sheet1", title="Scatter Marker",
            x_range="A2:A11", y_range="B2:B11",
            scatter_style="marker",
        )

    def test_area_chart(self, layer):
        layer.add_area_chart(sheet="Sheet1", title="Area", data_range="B1:B11")

    def test_area_chart_with_cats(self, layer):
        layer.add_area_chart(
            sheet="Sheet1", title="Area+Cats",
            data_range="B1:B11", cats_range="A1:A11",
        )

    def test_combo_chart(self, layer):
        # Need a C column for line data
        ws = layer._ws("Sheet1")
        for i in range(1, 12):
            ws[f"C{i}"] = i * 5
        layer.add_combo_chart(
            sheet="Sheet1", title="Combo",
            bar_data_range="B1:B11",
            line_data_range="C1:C11",
        )

    def test_combo_chart_no_secondary(self, layer):
        ws = layer._ws("Sheet1")
        for i in range(1, 12):
            ws[f"C{i}"] = i * 5
        layer.add_combo_chart(
            sheet="Sheet1", title="Combo",
            bar_data_range="B1:B11",
            line_data_range="C1:C11",
            secondary_axis=False,
        )

    def test_combo_chart_with_cats(self, layer):
        ws = layer._ws("Sheet1")
        for i in range(1, 12):
            ws[f"C{i}"] = i * 5
        layer.add_combo_chart(
            sheet="Sheet1", title="Combo+Cats",
            bar_data_range="B1:B11",
            line_data_range="C1:C11",
            cats_range="A1:A11",
        )


# ── Named Ranges ──


class TestNamedRanges:
    def test_create_named_range(self, layer):
        layer.create_named_range("TestRange", "Sheet1", "$A$1:$B$11")
        names = [name for name in layer.wb.defined_names]
        assert "TestRange" in names

    def test_create_named_range_no_workbook(self, empty_layer):
        with pytest.raises(RuntimeError, match="No workbook"):
            empty_layer.create_named_range("X", "Sheet1", "$A$1")


# ── Data Validation ──


class TestDataValidation:
    def test_list_validation(self, layer):
        layer.add_data_validation(
            "C2:C10", validation_type="list",
            formula1='"Yes,No,Maybe"', sheet="Sheet1",
        )

    def test_whole_validation(self, layer):
        layer.add_data_validation(
            "D2:D10", validation_type="whole",
            formula1="1", formula2="100", sheet="Sheet1",
        )


# ── View / Layout ──


class TestViewLayout:
    def test_freeze_panes(self, layer):
        layer.freeze_panes("A2", sheet="Sheet1")
        ws = layer._ws("Sheet1")
        assert ws.freeze_panes == "A2"

    def test_autofilter(self, layer):
        layer.set_autofilter("A1:B11", sheet="Sheet1")
        ws = layer._ws("Sheet1")
        assert ws.auto_filter.ref == "A1:B11"

    def test_merge_cells(self, layer):
        layer.merge_cells("D1:E1", sheet="Sheet1")

    def test_unmerge_cells(self, layer):
        layer.merge_cells("D1:E1", sheet="Sheet1")
        layer.unmerge_cells("D1:E1", sheet="Sheet1")

    def test_print_area(self, layer):
        layer.set_print_area("A1:B11", sheet="Sheet1")

    def test_print_title_rows(self, layer):
        layer.set_print_title_rows("1:1", sheet="Sheet1")

    def test_page_orientation_landscape(self, layer):
        layer.set_page_orientation(True, sheet="Sheet1")
        ws = layer._ws("Sheet1")
        assert ws.page_setup.orientation == "landscape"

    def test_page_orientation_portrait(self, layer):
        layer.set_page_orientation(False, sheet="Sheet1")
        ws = layer._ws("Sheet1")
        assert ws.page_setup.orientation == "portrait"

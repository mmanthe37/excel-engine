"""
Tests for Track A verification improvements (A1–A5).
"""

from __future__ import annotations

import math
import time
from dataclasses import field
from pathlib import Path
from unittest.mock import patch, MagicMock

import pytest
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference, LineChart
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule

from excel_engine.config import EngineConfig, TaskType, Layer
from excel_engine.parsers.task_extractor import Task
from excel_engine.verifier.workbook_verifier import (
    WorkbookVerifier,
    VerificationResult,
    _extract_sheet_references,
    _expand_range,
    _ranges_overlap,
)
from excel_engine.recovery import CircuitBreaker


# ── Helpers ──

def _make_task(task_type=TaskType.CELL_VALUE, **kwargs):
    defaults = dict(
        id="t1", task_type=task_type, description="test",
        sheet=None, cell=None, range=None, value=None,
        formula=None, style=None, params={},
    )
    defaults.update(kwargs)
    return Task(**defaults)


def _verifier_with_wb(wb, tmp_path):
    p = tmp_path / "test.xlsx"
    wb.save(p)
    v = WorkbookVerifier()
    v.load(p)
    return v


# ══════════════════════════════════════════════════════════════════
# A1 — Floating-Point Leniency
# ══════════════════════════════════════════════════════════════════

class TestA1FloatingPointLeniency:
    """_verify_cell_value should pass when numbers are nearly equal."""

    def test_classic_0_1_plus_0_2(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"].value = 0.1 + 0.2  # 0.30000000000000004
        v = _verifier_with_wb(wb, tmp_path)

        task = _make_task(cell="A1", value="0.3")
        result = v.verify_task(task)
        assert result.passed, result.message

    def test_large_numbers(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws["A1"].value = 1e15 + 0.1
        v = _verifier_with_wb(wb, tmp_path)

        task = _make_task(cell="A1", value=str(1e15 + 0.1))
        result = v.verify_task(task)
        assert result.passed

    def test_negative_numbers(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws["A1"].value = -0.1 - 0.2
        v = _verifier_with_wb(wb, tmp_path)

        task = _make_task(cell="A1", value="-0.3")
        result = v.verify_task(task)
        assert result.passed

    def test_exact_string_still_works(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws["A1"].value = "Hello"
        v = _verifier_with_wb(wb, tmp_path)

        task = _make_task(cell="A1", value="Hello")
        result = v.verify_task(task)
        assert result.passed

    def test_genuinely_different_numbers_fail(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws["A1"].value = 1.0
        v = _verifier_with_wb(wb, tmp_path)

        task = _make_task(cell="A1", value="2.0")
        result = v.verify_task(task)
        assert not result.passed

    def test_integer_vs_float_match(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws["A1"].value = 42
        v = _verifier_with_wb(wb, tmp_path)

        task = _make_task(cell="A1", value="42.0")
        result = v.verify_task(task)
        assert result.passed

    def test_non_numeric_strings_use_string_compare(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws["A1"].value = "abc"
        v = _verifier_with_wb(wb, tmp_path)

        task = _make_task(cell="A1", value="def")
        result = v.verify_task(task)
        assert not result.passed

    def test_zero_tolerance(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws["A1"].value = 1e-13
        v = _verifier_with_wb(wb, tmp_path)

        task = _make_task(cell="A1", value="0")
        result = v.verify_task(task)
        assert result.passed  # within abs_tol=1e-12


# ══════════════════════════════════════════════════════════════════
# A2 — Enhanced Chart Verification
# ══════════════════════════════════════════════════════════════════

class TestA2ChartVerification:
    """_verify_chart should check title, legend, series, and axis labels."""

    def _wb_with_chart(self, title=None, legend=True, series_count=1,
                       x_axis_title=None, y_axis_title=None):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        # Put some data
        for i in range(1, 6):
            ws.cell(row=i, column=1, value=i)
            ws.cell(row=i, column=2, value=i * 10)

        chart = BarChart()
        if title:
            chart.title = title
        if not legend:
            chart.legend = None
        for s in range(series_count):
            data = Reference(ws, min_col=2, min_row=1, max_row=5)
            chart.add_data(data)
        if x_axis_title:
            chart.x_axis.title = x_axis_title
        if y_axis_title:
            chart.y_axis.title = y_axis_title
        ws.add_chart(chart)
        return wb

    def test_chart_with_title_legend_series(self, tmp_path):
        wb = self._wb_with_chart(title="Sales", series_count=2)
        v = _verifier_with_wb(wb, tmp_path)

        task = _make_task(
            task_type=TaskType.CHART_BAR,
            params={"title": "Sales", "legend": True, "series_count": 2},
        )
        result = v.verify_task(task)
        assert result.passed, result.message

    def test_chart_wrong_title_fails(self, tmp_path):
        wb = self._wb_with_chart(title="Revenue")
        v = _verifier_with_wb(wb, tmp_path)

        task = _make_task(
            task_type=TaskType.CHART_BAR,
            params={"title": "Sales"},
        )
        result = v.verify_task(task)
        assert not result.passed

    def test_chart_no_legend_when_expected(self, tmp_path):
        wb = self._wb_with_chart(legend=False)
        v = _verifier_with_wb(wb, tmp_path)

        task = _make_task(
            task_type=TaskType.CHART_BAR,
            params={"legend": True},
        )
        result = v.verify_task(task)
        assert not result.passed

    def test_chart_no_series_fails(self, tmp_path):
        wb = self._wb_with_chart(series_count=0)
        v = _verifier_with_wb(wb, tmp_path)

        task = _make_task(task_type=TaskType.CHART_BAR)
        result = v.verify_task(task)
        assert not result.passed

    def test_chart_partial_specs_pass(self, tmp_path):
        """No specific params = existence check only (with series)."""
        wb = self._wb_with_chart()
        v = _verifier_with_wb(wb, tmp_path)

        task = _make_task(task_type=TaskType.CHART_BAR)
        result = v.verify_task(task)
        assert result.passed

    def test_no_chart_fails(self, tmp_path):
        wb = Workbook()
        v = _verifier_with_wb(wb, tmp_path)

        task = _make_task(task_type=TaskType.CHART_BAR)
        result = v.verify_task(task)
        assert not result.passed

    def test_axis_labels(self, tmp_path):
        wb = self._wb_with_chart(x_axis_title="Month", y_axis_title="Revenue")
        v = _verifier_with_wb(wb, tmp_path)

        task = _make_task(
            task_type=TaskType.CHART_BAR,
            params={"x_axis_title": "Month", "y_axis_title": "Revenue"},
        )
        result = v.verify_task(task)
        assert result.passed


# ══════════════════════════════════════════════════════════════════
# A3 — Conditional Format Verification
# ══════════════════════════════════════════════════════════════════

class TestA3ConditionalFormatVerification:
    """_verify_conditional_format should check rule type and range overlap."""

    def test_matching_rule_type(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        rule = CellIsRule(operator="greaterThan", formula=["5"], fill=None)
        ws.conditional_formatting.add("A1:A10", rule)
        v = _verifier_with_wb(wb, tmp_path)

        task = _make_task(
            task_type=TaskType.CONDITIONAL_FORMAT,
            params={"rule_type": "cellIs"},
        )
        result = v.verify_task(task)
        assert result.passed, result.message

    def test_non_matching_rule_type(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        rule = CellIsRule(operator="greaterThan", formula=["5"], fill=None)
        ws.conditional_formatting.add("A1:A10", rule)
        v = _verifier_with_wb(wb, tmp_path)

        task = _make_task(
            task_type=TaskType.CONDITIONAL_FORMAT,
            params={"rule_type": "colorScale"},
        )
        result = v.verify_task(task)
        assert not result.passed

    def test_matching_range(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        rule = CellIsRule(operator="greaterThan", formula=["5"], fill=None)
        ws.conditional_formatting.add("A1:A10", rule)
        v = _verifier_with_wb(wb, tmp_path)

        task = _make_task(
            task_type=TaskType.CONDITIONAL_FORMAT,
            range="A1:A10",
        )
        result = v.verify_task(task)
        assert result.passed

    def test_non_matching_range(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        rule = CellIsRule(operator="greaterThan", formula=["5"], fill=None)
        ws.conditional_formatting.add("A1:A10", rule)
        v = _verifier_with_wb(wb, tmp_path)

        task = _make_task(
            task_type=TaskType.CONDITIONAL_FORMAT,
            range="Z1:Z50",
        )
        result = v.verify_task(task)
        assert not result.passed

    def test_existence_fallback(self, tmp_path):
        """No type/range specified = existence check."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        rule = CellIsRule(operator="greaterThan", formula=["5"], fill=None)
        ws.conditional_formatting.add("B1:B5", rule)
        v = _verifier_with_wb(wb, tmp_path)

        task = _make_task(task_type=TaskType.CONDITIONAL_FORMAT)
        result = v.verify_task(task)
        assert result.passed

    def test_no_rules_fails(self, tmp_path):
        wb = Workbook()
        v = _verifier_with_wb(wb, tmp_path)

        task = _make_task(task_type=TaskType.CONDITIONAL_FORMAT)
        result = v.verify_task(task)
        assert not result.passed

    def test_partial_range_overlap(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        rule = CellIsRule(operator="greaterThan", formula=["5"], fill=None)
        ws.conditional_formatting.add("A1:C10", rule)
        v = _verifier_with_wb(wb, tmp_path)

        task = _make_task(
            task_type=TaskType.CONDITIONAL_FORMAT,
            range="B5:D15",  # overlaps at B5:C10
        )
        result = v.verify_task(task)
        assert result.passed


# ══════════════════════════════════════════════════════════════════
# A4 — Cross-Sheet Formula Link Validation
# ══════════════════════════════════════════════════════════════════

class TestA4CrossSheetFormula:
    """_verify_formula should warn about broken cross-sheet refs."""

    def test_extract_sheet_references(self):
        refs = _extract_sheet_references("=Sheet2!A1+Sheet3!B2")
        assert ("Sheet2", "A1") in refs
        assert ("Sheet3", "B2") in refs

    def test_extract_quoted_sheet(self):
        refs = _extract_sheet_references("='My Sheet'!C5")
        assert ("My Sheet", "C5") in refs

    def test_valid_cross_sheet_ref(self, tmp_path):
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Summary"
        ws2 = wb.create_sheet("Data")
        ws2["A1"].value = 100
        ws1["A1"].value = "=Data!A1"

        v = _verifier_with_wb(wb, tmp_path)
        task = _make_task(
            task_type=TaskType.FORMULA,
            cell="A1", formula="=Data!A1",
        )
        result = v.verify_task(task)
        assert result.passed
        assert not result.details.get("cross_sheet_warnings")

    def test_broken_sheet_ref_warns(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"
        ws["A1"].value = "=NonExistent!A1"

        v = _verifier_with_wb(wb, tmp_path)
        task = _make_task(
            task_type=TaskType.FORMULA,
            cell="A1", formula="=NonExistent!A1",
        )
        result = v.verify_task(task)
        # Should still pass (warnings, not failures)
        assert result.passed
        assert len(result.details.get("cross_sheet_warnings", [])) > 0

    def test_empty_referenced_cell_warns(self, tmp_path):
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Summary"
        ws2 = wb.create_sheet("Data")
        # Data!A1 left empty
        ws1["A1"].value = "=Data!A1"

        v = _verifier_with_wb(wb, tmp_path)
        task = _make_task(
            task_type=TaskType.FORMULA,
            cell="A1", formula="=Data!A1",
        )
        result = v.verify_task(task)
        assert result.passed
        warnings = result.details.get("cross_sheet_warnings", [])
        assert any("empty" in w.lower() for w in warnings)

    def test_external_ref_no_crash(self, tmp_path):
        """External refs like [Book2]Sheet1!A1 shouldn't crash."""
        wb = Workbook()
        ws = wb.active
        ws["A1"].value = "=[Book2]Sheet1!A1"
        v = _verifier_with_wb(wb, tmp_path)

        task = _make_task(
            task_type=TaskType.FORMULA,
            cell="A1",
        )
        result = v.verify_task(task)
        # Should not crash — the regex won't match [Book2] style refs
        assert result.passed

    def test_formula_without_cross_ref(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws["A1"].value = "=SUM(B1:B10)"
        v = _verifier_with_wb(wb, tmp_path)

        task = _make_task(
            task_type=TaskType.FORMULA,
            cell="A1", formula="=SUM(B1:B10)",
        )
        result = v.verify_task(task)
        assert result.passed
        assert not result.details.get("cross_sheet_warnings")


# ══════════════════════════════════════════════════════════════════
# A5 — Circuit Breaker
# ══════════════════════════════════════════════════════════════════

class TestA5CircuitBreaker:
    """Unit tests for CircuitBreaker."""

    def test_starts_closed(self):
        cb = CircuitBreaker(failure_threshold=3, reset_timeout=60)
        assert not cb.is_open("OPENPYXL")

    def test_opens_after_threshold(self):
        cb = CircuitBreaker(failure_threshold=3, reset_timeout=60)
        for _ in range(3):
            cb.record_failure("OPENPYXL")
        assert cb.is_open("OPENPYXL")

    def test_stays_closed_below_threshold(self):
        cb = CircuitBreaker(failure_threshold=3, reset_timeout=60)
        cb.record_failure("OPENPYXL")
        cb.record_failure("OPENPYXL")
        assert not cb.is_open("OPENPYXL")

    def test_success_resets_count(self):
        cb = CircuitBreaker(failure_threshold=3, reset_timeout=60)
        cb.record_failure("OPENPYXL")
        cb.record_failure("OPENPYXL")
        cb.record_success("OPENPYXL")
        cb.record_failure("OPENPYXL")
        cb.record_failure("OPENPYXL")
        assert not cb.is_open("OPENPYXL")

    def test_half_open_after_timeout(self):
        cb = CircuitBreaker(failure_threshold=2, reset_timeout=0.1)
        cb.record_failure("XLWINGS")
        cb.record_failure("XLWINGS")
        assert cb.is_open("XLWINGS")

        time.sleep(0.15)
        # After timeout, first call should return False (half-open probe)
        assert not cb.is_open("XLWINGS")
        # Second call (while still half-open) should return True
        assert cb.is_open("XLWINGS")

    def test_half_open_success_resets(self):
        cb = CircuitBreaker(failure_threshold=2, reset_timeout=0.1)
        cb.record_failure("XLWINGS")
        cb.record_failure("XLWINGS")
        time.sleep(0.15)
        cb.is_open("XLWINGS")  # transition to half-open
        cb.record_success("XLWINGS")  # probe succeeded
        assert not cb.is_open("XLWINGS")

    def test_half_open_failure_reopens(self):
        cb = CircuitBreaker(failure_threshold=2, reset_timeout=0.1)
        cb.record_failure("XLWINGS")
        cb.record_failure("XLWINGS")
        time.sleep(0.15)
        cb.is_open("XLWINGS")  # transition to half-open
        cb.record_failure("XLWINGS")
        cb.record_failure("XLWINGS")
        assert cb.is_open("XLWINGS")

    def test_manual_reset(self):
        cb = CircuitBreaker(failure_threshold=2, reset_timeout=300)
        cb.record_failure("OPENPYXL")
        cb.record_failure("OPENPYXL")
        assert cb.is_open("OPENPYXL")
        cb.reset("OPENPYXL")
        assert not cb.is_open("OPENPYXL")

    def test_independent_layers(self):
        cb = CircuitBreaker(failure_threshold=2, reset_timeout=300)
        cb.record_failure("OPENPYXL")
        cb.record_failure("OPENPYXL")
        assert cb.is_open("OPENPYXL")
        assert not cb.is_open("XLWINGS")


class TestA5CircuitBreakerConfig:
    """Config fields for circuit breaker."""

    def test_default_config(self):
        cfg = EngineConfig()
        assert cfg.circuit_breaker_threshold == 5
        assert cfg.circuit_breaker_reset_seconds == 300

    def test_custom_config(self):
        cfg = EngineConfig(circuit_breaker_threshold=10, circuit_breaker_reset_seconds=60)
        assert cfg.circuit_breaker_threshold == 10
        assert cfg.circuit_breaker_reset_seconds == 60


class TestA5CircuitBreakerEngineIntegration:
    """Circuit breaker integration with engine._execute_task."""

    def test_engine_has_circuit_breaker(self):
        from excel_engine.engine import ExcelEngine
        engine = ExcelEngine()
        assert hasattr(engine, "circuit_breaker")
        assert isinstance(engine.circuit_breaker, CircuitBreaker)

    def test_engine_cb_uses_config(self):
        from excel_engine.engine import ExcelEngine
        cfg = EngineConfig(circuit_breaker_threshold=7, circuit_breaker_reset_seconds=120)
        engine = ExcelEngine(config=cfg)
        assert engine.circuit_breaker.failure_threshold == 7
        assert engine.circuit_breaker.reset_timeout == 120


# ══════════════════════════════════════════════════════════════════
# Helper function unit tests
# ══════════════════════════════════════════════════════════════════

class TestHelperFunctions:
    def test_expand_range_single_cell(self):
        cells = _expand_range("A1")
        assert "A1" in cells

    def test_expand_range_block(self):
        cells = _expand_range("A1:B2")
        assert cells == {"A1", "A2", "B1", "B2"}

    def test_ranges_overlap_true(self):
        expected = _expand_range("A1:C3")
        assert _ranges_overlap("B2:D4", expected)

    def test_ranges_overlap_false(self):
        expected = _expand_range("A1:A2")
        assert not _ranges_overlap("C3:D4", expected)

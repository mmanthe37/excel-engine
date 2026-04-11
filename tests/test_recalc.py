"""Tests for formula recalculation, error scanning, verifier, and financial presets."""

import subprocess
from pathlib import Path
from unittest.mock import MagicMock, patch

import pytest
from openpyxl import Workbook

from excel_engine.config import TaskType
from excel_engine.engine import EngineResult
from excel_engine.presets.financial import (
    FinancialPreset,
    _detect_cell_category,
    apply_financial_preset,
    apply_number_formats,
)
from excel_engine.recalc import (
    EXCEL_ERRORS,
    RecalcResult,
    _find_soffice,
    _find_timeout_cmd,
    recalculate,
    scan_formula_errors,
)
from excel_engine.verifier.workbook_verifier import (
    VerificationResult,
    WorkbookVerifier,
)


# ── Helpers ──────────────────────────────────────────────────────────────────


def _save_wb(wb: Workbook, tmp_path: Path, name: str = "test.xlsx") -> Path:
    """Save a workbook to tmp_path and return the path."""
    p = tmp_path / name
    wb.save(p)
    return p


# ═══════════════════════════════════════════════════════════════════════════════
# A. scan_formula_errors
# ═══════════════════════════════════════════════════════════════════════════════


class TestScanFormulaErrors:
    """Tests for recalc.scan_formula_errors."""

    def test_counts_formulas(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = 10
        ws["A2"] = 20
        ws["A3"] = "=SUM(A1:A2)"
        ws["B1"] = "=A1*2"
        ws["B2"] = "hello"
        path = _save_wb(wb, tmp_path)

        result = scan_formula_errors(path)
        assert result["total_formulas"] == 2

    def test_no_errors_returns_zero(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = 10
        ws["A2"] = "=A1+5"
        path = _save_wb(wb, tmp_path)

        result = scan_formula_errors(path)
        assert result["total_errors"] == 0
        assert result["error_summary"] == {}

    def test_detects_cached_error_strings(self, tmp_path):
        """Workbooks saved by Excel cache error values as plain strings in data_only mode."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "#REF!"
        ws["A2"] = "#DIV/0!"
        ws["A3"] = "normal text"
        path = _save_wb(wb, tmp_path)

        result = scan_formula_errors(path)
        assert result["total_errors"] == 2
        assert "#REF!" in result["error_summary"]
        assert "#DIV/0!" in result["error_summary"]

    def test_error_location_format(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["B3"] = "#NAME?"
        path = _save_wb(wb, tmp_path)

        result = scan_formula_errors(path)
        locs = result["error_summary"]["#NAME?"]["locations"]
        assert locs == ["Sheet1!B3"]

    def test_multiple_sheets(self, tmp_path):
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Data"
        ws1["A1"] = "#VALUE!"

        ws2 = wb.create_sheet("Results")
        ws2["C5"] = "#N/A"

        path = _save_wb(wb, tmp_path)
        result = scan_formula_errors(path)
        assert result["total_errors"] == 2

        all_locs = []
        for info in result["error_summary"].values():
            all_locs.extend(info["locations"])
        assert "Data!A1" in all_locs
        assert "Results!C5" in all_locs

    def test_location_count_capped_at_20(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        for i in range(1, 30):
            ws.cell(row=i, column=1, value="#REF!")
        path = _save_wb(wb, tmp_path)

        result = scan_formula_errors(path)
        assert result["total_errors"] == 29
        assert result["error_summary"]["#REF!"]["count"] == 29
        assert len(result["error_summary"]["#REF!"]["locations"]) == 20

    def test_empty_workbook(self, tmp_path):
        wb = Workbook()
        path = _save_wb(wb, tmp_path)

        result = scan_formula_errors(path)
        assert result["total_formulas"] == 0
        assert result["total_errors"] == 0


# ═══════════════════════════════════════════════════════════════════════════════
# B. recalculate (mocked LibreOffice)
# ═══════════════════════════════════════════════════════════════════════════════


class TestRecalculate:
    """Tests for recalc.recalculate with mocked subprocess."""

    def _make_file(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "=1+1"
        return _save_wb(wb, tmp_path)

    @patch("excel_engine.recalc._find_soffice", return_value=None)
    def test_no_libreoffice_skips(self, mock_soffice, tmp_path):
        path = self._make_file(tmp_path)
        result = recalculate(path)

        assert result.skipped is True
        assert result.success is True
        assert "not installed" in result.warning.lower()

    @patch("excel_engine.recalc._ensure_macro_installed")
    @patch("excel_engine.recalc._find_timeout_cmd", return_value=None)
    @patch("excel_engine.recalc._find_soffice", return_value="/usr/bin/soffice")
    @patch("subprocess.run")
    def test_successful_recalculation(
        self, mock_run, mock_soffice, mock_timeout, mock_macro, tmp_path
    ):
        mock_run.return_value = MagicMock(returncode=0)
        path = self._make_file(tmp_path)

        result = recalculate(path)
        assert result.success is True
        assert result.skipped is False
        mock_run.assert_called_once()

    @patch("excel_engine.recalc._ensure_macro_installed")
    @patch("excel_engine.recalc._find_timeout_cmd", return_value=None)
    @patch("excel_engine.recalc._find_soffice", return_value="/usr/bin/soffice")
    @patch("subprocess.run")
    def test_nonzero_exit_code(
        self, mock_run, mock_soffice, mock_timeout, mock_macro, tmp_path
    ):
        mock_run.return_value = MagicMock(
            returncode=1, stderr=b"some error"
        )
        path = self._make_file(tmp_path)

        result = recalculate(path)
        assert result.success is False
        assert "code 1" in result.warning

    @patch("excel_engine.recalc._ensure_macro_installed")
    @patch("excel_engine.recalc._find_timeout_cmd", return_value=None)
    @patch("excel_engine.recalc._find_soffice", return_value="/usr/bin/soffice")
    @patch("subprocess.run", side_effect=subprocess.TimeoutExpired(cmd="soffice", timeout=30))
    def test_timeout_expired(
        self, mock_run, mock_soffice, mock_timeout, mock_macro, tmp_path
    ):
        path = self._make_file(tmp_path)

        result = recalculate(path)
        assert result.success is False
        assert "timed out" in result.warning.lower()

    @patch("excel_engine.recalc._ensure_macro_installed")
    @patch("excel_engine.recalc._find_timeout_cmd", return_value=None)
    @patch("excel_engine.recalc._find_soffice", return_value="/usr/bin/soffice")
    @patch("subprocess.run", side_effect=OSError("AF_UNIX path too long"))
    def test_af_unix_oserror_skips(
        self, mock_run, mock_soffice, mock_timeout, mock_macro, tmp_path
    ):
        path = self._make_file(tmp_path)

        result = recalculate(path)
        assert result.skipped is True
        assert result.success is True
        assert "AF_UNIX" in result.warning or "sandbox" in result.warning.lower()

    @patch("excel_engine.recalc._ensure_macro_installed")
    @patch("excel_engine.recalc._find_timeout_cmd", return_value=None)
    @patch("excel_engine.recalc._find_soffice", return_value="/usr/bin/soffice")
    @patch("subprocess.run", side_effect=OSError("permission denied"))
    def test_non_socket_oserror_propagates(
        self, mock_run, mock_soffice, mock_timeout, mock_macro, tmp_path
    ):
        path = self._make_file(tmp_path)
        with pytest.raises(OSError, match="permission denied"):
            recalculate(path)


# ═══════════════════════════════════════════════════════════════════════════════
# B-extra. _find_soffice / _find_timeout_cmd helpers
# ═══════════════════════════════════════════════════════════════════════════════


class TestRecalcHelpers:
    @patch("shutil.which", return_value=None)
    def test_find_soffice_not_installed(self, mock_which):
        assert _find_soffice() is None

    @patch("shutil.which", side_effect=lambda x: "/usr/bin/soffice" if x == "soffice" else None)
    def test_find_soffice_found(self, mock_which):
        assert _find_soffice() == "/usr/bin/soffice"

    @patch("excel_engine.recalc.platform.system", return_value="Linux")
    @patch("shutil.which", return_value="/usr/bin/timeout")
    def test_find_timeout_linux(self, mock_which, mock_system):
        assert _find_timeout_cmd() == "/usr/bin/timeout"

    @patch("excel_engine.recalc.platform.system", return_value="Windows")
    def test_find_timeout_windows(self, mock_system):
        assert _find_timeout_cmd() is None


# ═══════════════════════════════════════════════════════════════════════════════
# C. WorkbookVerifier: verify_formula_errors / count_formulas
# ═══════════════════════════════════════════════════════════════════════════════


class TestVerifierFormulaErrors:
    """Tests for WorkbookVerifier formula error methods."""

    def test_no_errors_returns_passed(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = 42
        ws["A2"] = "=A1+1"
        path = _save_wb(wb, tmp_path)

        v = WorkbookVerifier()
        v.load(path)
        v.load_with_values(path)
        results = v.verify_formula_errors()
        v.close()

        assert len(results) == 1
        assert results[0].passed is True
        assert results[0].task_id == "formula_error_scan"

    def test_detects_error_strings(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "#REF!"
        ws["B1"] = "#DIV/0!"
        path = _save_wb(wb, tmp_path)

        v = WorkbookVerifier()
        v.load(path)
        v.load_with_values(path)
        results = v.verify_formula_errors()
        v.close()

        failed = [r for r in results if not r.passed]
        assert len(failed) == 2
        error_types = {r.details["error_type"] for r in failed}
        assert "#REF!" in error_types
        assert "#DIV/0!" in error_types

    def test_no_workbook_path_skips_gracefully(self):
        v = WorkbookVerifier()
        results = v.verify_formula_errors()

        assert len(results) == 1
        assert results[0].passed is True
        assert "skipping" in results[0].message.lower()

    def test_loads_data_only_wb_on_demand(self, tmp_path):
        """When _wb_values is None but _path is set, it opens data_only=True on the fly."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "#NUM!"
        path = _save_wb(wb, tmp_path)

        v = WorkbookVerifier()
        v.load(path)
        # Intentionally do NOT call load_with_values
        results = v.verify_formula_errors()
        v.close()

        failed = [r for r in results if not r.passed]
        assert len(failed) == 1
        assert failed[0].details["error_type"] == "#NUM!"

    def test_count_formulas(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "=SUM(B1:B5)"
        ws["A2"] = "=A1*2"
        ws["A3"] = "plain text"
        ws["A4"] = 42
        path = _save_wb(wb, tmp_path)

        v = WorkbookVerifier()
        v.load(path)
        assert v.count_formulas() == 2
        v.close()

    def test_count_formulas_no_workbook_raises(self):
        v = WorkbookVerifier()
        with pytest.raises(RuntimeError, match="No workbook loaded"):
            v.count_formulas()


# ═══════════════════════════════════════════════════════════════════════════════
# D. FinancialPreset and apply_financial_preset
# ═══════════════════════════════════════════════════════════════════════════════


class TestDetectCellCategory:
    """Tests for _detect_cell_category."""

    def test_formula(self):
        cell = MagicMock(value="=SUM(A1:A5)")
        assert _detect_cell_category(cell) == "formula"

    def test_crosssheet(self):
        cell = MagicMock(value="=Sheet2!A1+10")
        assert _detect_cell_category(cell) == "crosssheet"

    def test_external(self):
        cell = MagicMock(value="=[Budget.xlsx]Sheet1!A1")
        assert _detect_cell_category(cell) == "external"

    def test_input_int(self):
        cell = MagicMock(value=42)
        assert _detect_cell_category(cell) == "input"

    def test_input_float(self):
        cell = MagicMock(value=3.14)
        assert _detect_cell_category(cell) == "input"

    def test_text(self):
        cell = MagicMock(value="Revenue")
        assert _detect_cell_category(cell) == "text"

    def test_none(self):
        cell = MagicMock(value=None)
        assert _detect_cell_category(cell) == "text"


class TestApplyFinancialPreset:
    """Tests for apply_financial_preset."""

    def _build_workbook(self):
        """Create a workbook with various cell categories."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Model"

        # Row 1 — headers
        ws["A1"] = "Item"
        ws["B1"] = "Value"

        # Row 2 — hardcoded input number
        ws["A2"] = "Revenue"
        ws["B2"] = 100000

        # Row 3 — formula
        ws["A3"] = "COGS"
        ws["B3"] = "=B2*0.6"

        # Row 4 — cross-sheet formula
        ws["A4"] = "Tax Rate"
        ws["B4"] = "=Assumptions!B1"

        # Row 5 — external link
        ws["A5"] = "Benchmark"
        ws["B5"] = "=[Market.xlsx]Data!A1"

        return wb

    def test_color_coding_inputs_blue(self):
        wb = self._build_workbook()
        preset = FinancialPreset()
        apply_financial_preset(wb, preset)

        ws = wb["Model"]
        # B2 is a numeric input → blue
        assert ws["B2"].font.color.rgb == "00" + preset.input_font_color or \
               ws["B2"].font.color.rgb == preset.input_font_color

    def test_color_coding_formula_black(self):
        wb = self._build_workbook()
        preset = FinancialPreset()
        apply_financial_preset(wb, preset)

        ws = wb["Model"]
        # B3 is a formula → black
        color = ws["B3"].font.color.rgb
        assert preset.formula_font_color in color

    def test_color_coding_crosssheet_green(self):
        wb = self._build_workbook()
        preset = FinancialPreset()
        apply_financial_preset(wb, preset)

        ws = wb["Model"]
        color = ws["B4"].font.color.rgb
        assert preset.crosssheet_font_color in color

    def test_color_coding_external_red(self):
        wb = self._build_workbook()
        preset = FinancialPreset()
        apply_financial_preset(wb, preset)

        ws = wb["Model"]
        color = ws["B5"].font.color.rgb
        assert preset.external_font_color in color

    def test_header_row_bold_and_sized(self):
        wb = self._build_workbook()
        preset = FinancialPreset()
        apply_financial_preset(wb, preset)

        ws = wb["Model"]
        assert ws["A1"].font.bold is True
        assert ws["A1"].font.size == preset.header_font_size

    def test_summary_counts(self):
        wb = self._build_workbook()
        summary = apply_financial_preset(wb)

        assert summary["inputs_colored"] >= 1  # B2
        assert summary["formulas_colored"] >= 1  # B3
        assert summary["crosssheet_colored"] >= 1  # B4
        assert summary["external_colored"] >= 1  # B5
        assert summary["total_cells"] > 0

    def test_empty_sheet(self):
        wb = Workbook()
        # Active sheet is empty
        summary = apply_financial_preset(wb)
        assert summary["total_cells"] == 0

    def test_specific_sheets_filter(self):
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "IncludeMe"
        ws1["A1"] = 42

        ws2 = wb.create_sheet("SkipMe")
        ws2["A1"] = 99

        summary = apply_financial_preset(wb, sheets=["IncludeMe"])
        assert summary["total_cells"] == 1  # only A1 from IncludeMe


class TestApplyNumberFormats:
    """Tests for apply_number_formats."""

    def test_currency_format_applied(self):
        wb = Workbook()
        ws = wb.active
        ws["B1"] = "Revenue"
        ws["B2"] = 50000
        ws["B3"] = 75000

        preset = FinancialPreset()
        apply_number_formats(wb, preset, currency_columns=["B"])

        assert ws["B2"].number_format == preset.currency_format
        assert ws["B3"].number_format == preset.currency_format

    def test_percentage_format_applied(self):
        wb = Workbook()
        ws = wb.active
        ws["C2"] = 0.15

        preset = FinancialPreset()
        apply_number_formats(wb, preset, percentage_columns=["C"])

        assert ws["C2"].number_format == preset.percentage_format

    def test_no_columns_is_noop(self):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = 100
        original_fmt = ws["A1"].number_format

        apply_number_formats(wb)
        assert ws["A1"].number_format == original_fmt

    def test_skips_none_cells(self):
        wb = Workbook()
        ws = wb.active
        ws["B1"] = 100
        # B2 is intentionally empty

        preset = FinancialPreset()
        apply_number_formats(wb, preset, currency_columns=["B"])

        # Should not raise; B2 stays untouched
        assert ws["B2"].value is None


# ═══════════════════════════════════════════════════════════════════════════════
# E. EngineResult integration (formula_errors in summary)
# ═══════════════════════════════════════════════════════════════════════════════


class TestEngineResultFormulaSummary:
    """Tests that EngineResult.summary() integrates RecalcResult correctly."""

    def _base_result(self, **overrides) -> EngineResult:
        defaults = dict(
            success=True,
            workbook_path=Path("test.xlsx"),
            sections_completed=1,
            sections_total=1,
            tasks_completed=5,
            tasks_total=5,
        )
        defaults.update(overrides)
        return EngineResult(**defaults)

    def test_summary_ok(self):
        er = self._base_result(
            formula_errors=RecalcResult(
                success=True, total_formulas=10, total_errors=0
            ),
        )
        s = er.summary()
        assert "Formula recalc: OK" in s
        assert "10 formulas" in s

    def test_summary_skipped(self):
        er = self._base_result(
            formula_errors=RecalcResult(
                success=True, skipped=True, warning="LibreOffice not installed"
            ),
        )
        s = er.summary()
        assert "skipped" in s.lower()

    def test_summary_with_errors(self):
        er = self._base_result(
            formula_errors=RecalcResult(
                success=True,
                total_formulas=20,
                total_errors=3,
                error_summary={
                    "#REF!": {"count": 2, "locations": ["Sheet1!A1", "Sheet1!B2"]},
                    "#DIV/0!": {"count": 1, "locations": ["Sheet2!C3"]},
                },
            ),
        )
        s = er.summary()
        assert "3 errors" in s
        assert "20 formulas" in s
        assert "#REF!" in s

    def test_summary_no_formula_errors_field(self):
        er = self._base_result(formula_errors=None)
        s = er.summary()
        # Should not crash and should not mention formula recalc
        assert "Formula recalc" not in s

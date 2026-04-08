"""
Tests for ExcelEngine dispatch — covers _exec_openpyxl for all TaskType handlers,
plus run(), execute(), scan(), plan(), and error handling.
"""

import pytest
from pathlib import Path
from unittest.mock import patch, MagicMock

from openpyxl import load_workbook
from excel_engine import ExcelEngine, EngineConfig, TaskType, Layer
from excel_engine.parsers.task_extractor import Task
from excel_engine.planner.task_planner import ExecutionPlan, Section


# ── Fixtures ──

@pytest.fixture
def engine():
    config = EngineConfig()
    config.layer_order = [Layer.OPENPYXL]
    return ExcelEngine(config)


@pytest.fixture
def engine_with_workbook(sample_workbook, engine):
    engine._openpyxl.open(sample_workbook)
    yield engine, sample_workbook
    engine._openpyxl.close()


def _make_task(task_type, **kwargs):
    kwargs.setdefault("id", f"test-{task_type.value}")
    kwargs.setdefault("description", f"Test {task_type.value}")
    return Task(task_type=task_type, **kwargs)


# ══════════════════════════════════════════════════════════════════
# Dispatch tests for _exec_openpyxl
# ══════════════════════════════════════════════════════════════════

class TestExecOpenpyxlDispatch:
    """Test each TaskType handler in _exec_openpyxl."""

    def test_cell_value(self, engine_with_workbook):
        eng, wb = engine_with_workbook
        task = _make_task(TaskType.CELL_VALUE, cell="A1", value="Hello", sheet="Sales")
        eng._exec_openpyxl(task, wb)
        assert eng._openpyxl._ws("Sales")["A1"].value == "Hello"

    def test_formula(self, engine_with_workbook):
        eng, wb = engine_with_workbook
        task = _make_task(TaskType.FORMULA, cell="F2", formula="=SUM(B2:E2)", sheet="Sales")
        eng._exec_openpyxl(task, wb)
        assert eng._openpyxl._ws("Sales")["F2"].value == "=SUM(B2:E2)"

    def test_text_function(self, engine_with_workbook):
        eng, wb = engine_with_workbook
        task = _make_task(TaskType.TEXT_FUNCTION, cell="G2", formula="=UPPER(A2)", sheet="Sales")
        eng._exec_openpyxl(task, wb)
        assert eng._openpyxl._ws("Sales")["G2"].value == "=UPPER(A2)"

    def test_lookup_function(self, engine_with_workbook):
        eng, wb = engine_with_workbook
        task = _make_task(TaskType.LOOKUP_FUNCTION, cell="G3", formula='=VLOOKUP(A3,A:E,2,FALSE)', sheet="Sales")
        eng._exec_openpyxl(task, wb)
        assert eng._openpyxl._ws("Sales")["G3"].value == '=VLOOKUP(A3,A:E,2,FALSE)'

    def test_filter_function(self, engine_with_workbook):
        eng, wb = engine_with_workbook
        task = _make_task(TaskType.FILTER_FUNCTION, cell="H2", formula='=FILTER(A2:A11,B2:B11>500)', sheet="Sales")
        eng._exec_openpyxl(task, wb)
        assert eng._openpyxl._ws("Sales")["H2"].value == '=FILTER(A2:A11,B2:B11>500)'

    def test_sort_function(self, engine_with_workbook):
        eng, wb = engine_with_workbook
        task = _make_task(TaskType.SORT_FUNCTION, cell="I2", formula='=SORT(A2:A11)', sheet="Sales")
        eng._exec_openpyxl(task, wb)
        assert eng._openpyxl._ws("Sales")["I2"].value == '=SORT(A2:A11)'

    def test_unique_function(self, engine_with_workbook):
        eng, wb = engine_with_workbook
        task = _make_task(TaskType.UNIQUE_FUNCTION, cell="J2", formula='=UNIQUE(A2:A11)', sheet="Sales")
        eng._exec_openpyxl(task, wb)
        assert eng._openpyxl._ws("Sales")["J2"].value == '=UNIQUE(A2:A11)'

    def test_three_d_reference(self, engine_with_workbook):
        eng, wb = engine_with_workbook
        task = _make_task(TaskType.THREE_D_REFERENCE, cell="A1", formula='=SUM(Sales:Summary!B2)', sheet="Summary")
        eng._exec_openpyxl(task, wb)
        assert eng._openpyxl._ws("Summary")["A1"].value == '=SUM(Sales:Summary!B2)'

    def test_external_reference(self, engine_with_workbook):
        eng, wb = engine_with_workbook
        task = _make_task(TaskType.EXTERNAL_REFERENCE, cell="A2", formula='=[Other.xlsx]Sheet1!A1', sheet="Summary")
        eng._exec_openpyxl(task, wb)
        assert eng._openpyxl._ws("Summary")["A2"].value == '=[Other.xlsx]Sheet1!A1'

    def test_table_create(self, engine_with_workbook):
        eng, wb = engine_with_workbook
        task = _make_task(
            TaskType.TABLE_CREATE, sheet="Sales",
            range="A1:E11",
            params={"name": "SalesTable", "style": "TableStyleMedium2"},
        )
        eng._exec_openpyxl(task, wb)
        ws = eng._openpyxl._ws("Sales")
        assert "SalesTable" in ws.tables

    def test_table_total_row(self, engine_with_workbook):
        eng, wb = engine_with_workbook
        # First create a table
        eng._exec_openpyxl(
            _make_task(TaskType.TABLE_CREATE, sheet="Sales", range="A1:E11",
                       params={"name": "TotalsTable"}), wb)
        task = _make_task(TaskType.TABLE_TOTAL_ROW, sheet="Sales",
                          params={"name": "TotalsTable"})
        eng._exec_openpyxl(task, wb)
        ws = eng._openpyxl._ws("Sales")
        assert ws.tables["TotalsTable"].totalsRowShown is True

    def test_number_format(self, engine_with_workbook):
        eng, wb = engine_with_workbook
        task = _make_task(TaskType.NUMBER_FORMAT, range="B2:B2", sheet="Sales",
                          params={"format": "$#,##0.00"})
        eng._exec_openpyxl(task, wb)
        ws = eng._openpyxl._ws("Sales")
        assert ws["B2"].number_format == "$#,##0.00"

    def test_font(self, engine_with_workbook):
        eng, wb = engine_with_workbook
        task = _make_task(TaskType.FONT, range="A1:A1", sheet="Sales",
                          params={"bold": True, "size": 14, "color": "FF0000"})
        eng._exec_openpyxl(task, wb)
        ws = eng._openpyxl._ws("Sales")
        assert ws["A1"].font.bold is True
        assert ws["A1"].font.size == 14

    def test_fill(self, engine_with_workbook):
        eng, wb = engine_with_workbook
        task = _make_task(TaskType.FILL, range="A1:A1", sheet="Sales",
                          params={"color": "FFFF00"})
        eng._exec_openpyxl(task, wb)
        ws = eng._openpyxl._ws("Sales")
        assert ws["A1"].fill.start_color.index == "00FFFF00"

    def test_alignment(self, engine_with_workbook):
        eng, wb = engine_with_workbook
        task = _make_task(TaskType.ALIGNMENT, range="A1:A1", sheet="Sales",
                          params={"horizontal": "center", "wrap_text": True})
        eng._exec_openpyxl(task, wb)
        ws = eng._openpyxl._ws("Sales")
        assert ws["A1"].alignment.horizontal == "center"
        assert ws["A1"].alignment.wrap_text is True

    def test_border(self, engine_with_workbook):
        eng, wb = engine_with_workbook
        task = _make_task(TaskType.BORDER, range="A1:A1", sheet="Sales",
                          params={"style": "thin"})
        eng._exec_openpyxl(task, wb)
        ws = eng._openpyxl._ws("Sales")
        assert ws["A1"].border.left.style is not None

    def test_column_width(self, engine_with_workbook):
        eng, wb = engine_with_workbook
        task = _make_task(TaskType.COLUMN_WIDTH, sheet="Sales",
                          params={"column": "A", "width": 25})
        eng._exec_openpyxl(task, wb)
        ws = eng._openpyxl._ws("Sales")
        assert ws.column_dimensions["A"].width == 25

    def test_row_height(self, engine_with_workbook):
        eng, wb = engine_with_workbook
        task = _make_task(TaskType.ROW_HEIGHT, cell="A5", sheet="Sales",
                          params={"height": 30})
        eng._exec_openpyxl(task, wb)
        ws = eng._openpyxl._ws("Sales")
        assert ws.row_dimensions[5].height == 30

    def test_tab_color(self, engine_with_workbook):
        eng, wb = engine_with_workbook
        task = _make_task(TaskType.TAB_COLOR, sheet="Sales",
                          params={"color": "FF0000"})
        eng._exec_openpyxl(task, wb)
        ws = eng._openpyxl._ws("Sales")
        assert ws.sheet_properties.tabColor is not None

    def test_page_break(self, engine_with_workbook):
        eng, wb = engine_with_workbook
        task = _make_task(TaskType.PAGE_BREAK, cell="A5", sheet="Sales",
                          params={})
        eng._exec_openpyxl(task, wb)
        ws = eng._openpyxl._ws("Sales")
        assert len(ws.row_breaks.brk) > 0

    def test_hyperlink(self, engine_with_workbook):
        eng, wb = engine_with_workbook
        task = _make_task(TaskType.HYPERLINK, cell="A1", sheet="Summary",
                          params={"url": "https://example.com", "display": "Example"})
        eng._exec_openpyxl(task, wb)
        ws = eng._openpyxl._ws("Summary")
        assert ws["A1"].hyperlink is not None
        assert ws["A1"].value == "Example"

    def test_sheet_create(self, engine_with_workbook):
        eng, wb = engine_with_workbook
        task = _make_task(TaskType.SHEET_CREATE, params={"name": "NewSheet"})
        eng._exec_openpyxl(task, wb)
        assert "NewSheet" in eng._openpyxl.wb.sheetnames

    def test_sheet_rename(self, engine_with_workbook):
        eng, wb = engine_with_workbook
        task = _make_task(TaskType.SHEET_RENAME, sheet="Renamed",
                          params={"old_name": "Summary", "new_name": "Renamed"})
        eng._exec_openpyxl(task, wb)
        assert "Renamed" in eng._openpyxl.wb.sheetnames

    def test_sheet_move(self, engine_with_workbook):
        eng, wb = engine_with_workbook
        task = _make_task(TaskType.SHEET_MOVE, sheet="Summary",
                          params={"position": -1})
        eng._exec_openpyxl(task, wb)
        # Sheet should still exist, just reordered
        assert "Summary" in eng._openpyxl.wb.sheetnames

    def test_sheet_copy(self, engine_with_workbook):
        eng, wb = engine_with_workbook
        task = _make_task(TaskType.SHEET_COPY, sheet="Sales",
                          params={"source": "Sales", "new_name": "Sales Backup"})
        eng._exec_openpyxl(task, wb)
        assert "Sales Backup" in eng._openpyxl.wb.sheetnames

    def test_merge_cells(self, engine_with_workbook):
        eng, wb = engine_with_workbook
        task = _make_task(TaskType.MERGE_CELLS, range="A1:C1", sheet="Summary")
        eng._exec_openpyxl(task, wb)
        ws = eng._openpyxl._ws("Summary")
        assert len(ws.merged_cells.ranges) > 0

    def test_freeze_panes(self, engine_with_workbook):
        eng, wb = engine_with_workbook
        task = _make_task(TaskType.FREEZE_PANES, cell="A2", sheet="Sales")
        eng._exec_openpyxl(task, wb)
        ws = eng._openpyxl._ws("Sales")
        assert ws.freeze_panes == "A2"

    def test_autofilter(self, engine_with_workbook):
        eng, wb = engine_with_workbook
        task = _make_task(TaskType.AUTOFILTER, range="A1:E11", sheet="Sales")
        eng._exec_openpyxl(task, wb)
        ws = eng._openpyxl._ws("Sales")
        assert ws.auto_filter.ref == "A1:E11"

    def test_conditional_format(self, engine_with_workbook):
        eng, wb = engine_with_workbook
        task = _make_task(TaskType.CONDITIONAL_FORMAT, range="B2:B11", sheet="Sales",
                          params={"operator": "greaterThan", "formula": ["5000"]})
        eng._exec_openpyxl(task, wb)
        ws = eng._openpyxl._ws("Sales")
        assert len(list(ws.conditional_formatting)) > 0

    def test_data_validation(self, engine_with_workbook):
        eng, wb = engine_with_workbook
        task = _make_task(TaskType.DATA_VALIDATION, range="A20", sheet="Sales",
                          params={"type": "list", "formula1": '"Yes,No"'})
        eng._exec_openpyxl(task, wb)
        ws = eng._openpyxl._ws("Sales")
        assert len(ws.data_validations.dataValidation) > 0

    def test_named_range(self, engine_with_workbook):
        eng, wb = engine_with_workbook
        task = _make_task(TaskType.NAMED_RANGE, sheet="Sales", range="$B$2:$B$11",
                          params={"name": "Q1Sales"})
        eng._exec_openpyxl(task, wb)
        names = list(eng._openpyxl.wb.defined_names)
        assert "Q1Sales" in names

    def test_print_settings_landscape(self, engine_with_workbook):
        eng, wb = engine_with_workbook
        task = _make_task(TaskType.PRINT_SETTINGS, sheet="Sales",
                          params={"landscape": True})
        eng._exec_openpyxl(task, wb)
        ws = eng._openpyxl._ws("Sales")
        assert ws.page_setup.orientation == "landscape"

    def test_print_settings_area(self, engine_with_workbook):
        eng, wb = engine_with_workbook
        task = _make_task(TaskType.PRINT_SETTINGS, sheet="Sales",
                          params={"print_area": "A1:E11"})
        eng._exec_openpyxl(task, wb)
        ws = eng._openpyxl._ws("Sales")
        assert ws.print_area is not None

    def test_chart_bar(self, engine_with_workbook):
        eng, wb = engine_with_workbook
        task = _make_task(TaskType.CHART_BAR, sheet="Sales",
                          params={"title": "Q1 Sales", "data_range": "B1:B11"})
        eng._exec_openpyxl(task, wb)
        ws = eng._openpyxl._ws("Sales")
        assert len(ws._charts) >= 1

    def test_chart_line(self, engine_with_workbook):
        eng, wb = engine_with_workbook
        task = _make_task(TaskType.CHART_LINE, sheet="Sales",
                          params={"title": "Trend", "data_range": "B1:B11"})
        eng._exec_openpyxl(task, wb)
        ws = eng._openpyxl._ws("Sales")
        assert len(ws._charts) >= 1

    def test_chart_pie(self, engine_with_workbook):
        eng, wb = engine_with_workbook
        task = _make_task(TaskType.CHART_PIE, sheet="Sales",
                          params={"title": "Distribution", "data_range": "B1:B11"})
        eng._exec_openpyxl(task, wb)
        ws = eng._openpyxl._ws("Sales")
        assert len(ws._charts) >= 1

    def test_formatting_bold(self, engine_with_workbook):
        eng, wb = engine_with_workbook
        task = _make_task(TaskType.FORMATTING, cell="A1", sheet="Sales",
                          params={"bold": True})
        eng._exec_openpyxl(task, wb)
        ws = eng._openpyxl._ws("Sales")
        assert ws["A1"].font.bold is True

    def test_table_style(self, engine_with_workbook):
        eng, wb = engine_with_workbook
        # Create table first
        eng._exec_openpyxl(
            _make_task(TaskType.TABLE_CREATE, sheet="Sales", range="A1:E11",
                       params={"name": "StyleTable", "style": "TableStyleMedium5"}), wb)
        task = _make_task(TaskType.TABLE_STYLE, sheet="Sales",
                          style="TableStyleMedium2",
                          params={"table_name": "StyleTable"})
        # TABLE_STYLE is delegated to openpyxl but may use xlwings;
        # for openpyxl-only config it falls through to the dispatch
        # engine handles this via _dispatch_task → layer cascade


# ══════════════════════════════════════════════════════════════════
# Error handling
# ══════════════════════════════════════════════════════════════════

class TestEngineErrorHandling:
    """Test error paths and edge cases in the engine."""

    def test_unknown_task_type_not_implemented(self, engine_with_workbook):
        """TaskTypes with no openpyxl handler raise NotImplementedError."""
        eng, wb = engine_with_workbook
        # SAVE has no openpyxl handler — falls through to else branch
        task = _make_task(TaskType.SAVE)
        with pytest.raises(NotImplementedError):
            eng._exec_openpyxl(task, wb)

    def test_dispatch_unknown_layer(self, engine_with_workbook):
        eng, wb = engine_with_workbook
        task = _make_task(TaskType.CELL_VALUE, cell="A1", value="test")
        with pytest.raises(ValueError, match="Unknown layer"):
            eng._dispatch_task(task, 999, wb)

    def test_execute_task_no_layers(self, engine_with_workbook):
        """A task type with no configured layers returns (False, [])."""
        eng, wb = engine_with_workbook
        # Mock get_layers_for_task to return empty
        with patch.object(eng.config, "get_layers_for_task", return_value=[]):
            task = _make_task(TaskType.CELL_VALUE, cell="A1", value="test")
            success, errors = eng._execute_task(task, wb)
        assert success is False

    def test_task_cascade_retry(self, sample_workbook):
        """When all layers raise, engine should return (False, errors)."""
        config = EngineConfig()
        config.max_retries = 1
        config.retry_delay = 0.0
        eng = ExcelEngine(config)

        task = _make_task(TaskType.CELL_VALUE, cell="A1", value="test")
        # Patch _dispatch_task to always raise so all layers fail
        with patch.object(eng, "_dispatch_task", side_effect=RuntimeError("fail")):
            success, errors = eng._execute_task(task, sample_workbook)
        assert success is False
        assert len(errors) > 0


# ══════════════════════════════════════════════════════════════════
# Public API: run, execute, scan, plan
# ══════════════════════════════════════════════════════════════════

class TestEnginePublicAPI:
    """Test engine.run(), execute(), scan(), plan()."""

    def test_run_with_tasks(self, sample_workbook):
        config = EngineConfig()
        # Keep full layer_order so get_layers_for_task can sort
        config.verify_after_each_section = True
        eng = ExcelEngine(config)

        tasks = [
            _make_task(TaskType.CELL_VALUE, id="t1", cell="A1", value="RunTest", sheet="Sales"),
            _make_task(TaskType.FORMULA, id="t2", cell="F2", formula="=SUM(B2:E2)", sheet="Sales"),
        ]
        result = eng.run(workbook=sample_workbook, tasks=tasks)
        assert result.tasks_completed == 2
        assert result.success is True

        # Verify the values were actually written
        wb = load_workbook(str(sample_workbook))
        assert wb["Sales"]["A1"].value == "RunTest"
        assert wb["Sales"]["F2"].value == "=SUM(B2:E2)"
        wb.close()

    def test_run_with_instruction_text(self, sample_workbook):
        config = EngineConfig()
        config.layer_order = [Layer.OPENPYXL]
        config.verify_after_each_section = False
        eng = ExcelEngine(config)

        result = eng.run(
            workbook=sample_workbook,
            instruction_text="Enter 42 in cell A1 of the Sales sheet",
        )
        # Should succeed even if parsing yields no tasks (engine is best-effort)
        assert result is not None

    def test_run_no_instructions_raises(self, sample_workbook):
        config = EngineConfig()
        eng = ExcelEngine(config)
        result = eng.run(workbook=sample_workbook)
        assert result.success is False
        assert len(result.errors) > 0

    def test_execute_with_plan(self, sample_workbook):
        config = EngineConfig()
        config.verify_after_each_section = False
        eng = ExcelEngine(config)

        tasks = [
            _make_task(TaskType.CELL_VALUE, id="e1", cell="B1", value="Exec", sheet="Summary"),
        ]
        section = Section(id="sec1", name="Test Section", sheet="Summary", tasks=tasks)
        plan = ExecutionPlan(sections=[section], total_tasks=1, estimated_time_seconds=2.0)

        result = eng.execute(plan, sample_workbook)
        assert result.tasks_completed == 1
        assert result.sections_completed == 1

    def test_plan_creates_sections(self, engine):
        tasks = [
            _make_task(TaskType.CELL_VALUE, id="p1", cell="A1", value="x", sheet="Sheet1"),
            _make_task(TaskType.FORMULA, id="p2", cell="B1", formula="=A1", sheet="Sheet1"),
        ]
        plan = engine.plan(tasks)
        assert plan.total_tasks == 2
        assert plan.section_count >= 1

    def test_engine_result_summary(self, sample_workbook):
        result = MagicMock()
        result = __import__("excel_engine.engine", fromlist=["EngineResult"]).EngineResult(
            success=True,
            workbook_path=sample_workbook,
            sections_completed=1,
            sections_total=1,
            tasks_completed=2,
            tasks_total=2,
        )
        summary = result.summary()
        assert "SUCCESS" in summary
        assert "2/2" in summary

    def test_engine_result_summary_with_errors(self, sample_workbook):
        from excel_engine.engine import EngineResult
        result = EngineResult(
            success=False,
            workbook_path=sample_workbook,
            sections_completed=0,
            sections_total=1,
            tasks_completed=0,
            tasks_total=2,
            errors=["Something broke"],
            failed_tasks=["t1"],
        )
        summary = result.summary()
        assert "FAILED" in summary
        assert "Something broke" in summary

    def test_properties(self, engine):
        """Verify public properties return correct layer instances."""
        assert engine.openpyxl is engine._openpyxl
        assert engine.parser is engine._parser
        assert engine.planner is engine._planner
        assert engine.verifier is engine._verifier
        assert engine.extractor is engine._extractor

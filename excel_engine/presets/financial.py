"""Investment banking / financial modeling best practices as an opt-in preset.

This module provides color-coding conventions, number formats, and font
standards commonly used in investment banking and financial modeling.
"""

from __future__ import annotations

import logging
from dataclasses import dataclass
from typing import Optional

from openpyxl.styles import Font, PatternFill

logger = logging.getLogger(__name__)


@dataclass
class FinancialPreset:
    """Investment banking / financial model formatting standards."""

    # Color coding (RGB hex)
    input_font_color: str = "0000FF"       # Blue — hardcoded inputs
    formula_font_color: str = "000000"     # Black — all formulas
    crosssheet_font_color: str = "008000"  # Green — cross-sheet links
    external_font_color: str = "FF0000"    # Red — external links
    assumption_fill_color: str = "FFFF00"  # Yellow — key assumptions

    # Number formats
    currency_format: str = '$#,##0;($#,##0);"-"'
    currency_mm_format: str = '$#,##0.0;($#,##0.0);"-"'
    percentage_format: str = "0.0%"
    multiple_format: str = '0.0"x"'
    year_format: str = "@"  # Text format for years

    # Font
    font_name: str = "Arial"
    font_size: int = 10
    header_font_size: int = 11
    header_bold: bool = True


# ---------------------------------------------------------------------------
# Private helpers
# ---------------------------------------------------------------------------

def _detect_cell_category(cell) -> str:
    """Classify a cell as 'formula', 'crosssheet', 'external', 'input', or 'text'.

    Args:
        cell: An openpyxl cell object.

    Returns:
        One of ``'external'``, ``'crosssheet'``, ``'formula'``, ``'input'``,
        or ``'text'``.
    """
    value = cell.value
    if value is None:
        return "text"

    if isinstance(value, str) and value.startswith("="):
        if "[" in value:
            return "external"
        if "!" in value:
            return "crosssheet"
        return "formula"

    if isinstance(value, (int, float)):
        return "input"

    return "text"


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def apply_financial_preset(
    wb,
    preset: Optional[FinancialPreset] = None,
    sheets: Optional[list[str]] = None,
) -> dict:
    """Apply financial formatting preset to a workbook.

    Args:
        wb: openpyxl ``Workbook`` object.
        preset: :class:`FinancialPreset` config (uses defaults if *None*).
        sheets: Specific sheet names to process (*None* → all sheets).

    Returns:
        Summary dict with counts of cells colored by category::

            {
                "inputs_colored": n,
                "formulas_colored": n,
                "crosssheet_colored": n,
                "external_colored": n,
                "total_cells": n,
            }
    """
    if preset is None:
        preset = FinancialPreset()

    color_map: dict[str, str] = {
        "input": preset.input_font_color,
        "formula": preset.formula_font_color,
        "crosssheet": preset.crosssheet_font_color,
        "external": preset.external_font_color,
    }

    summary = {
        "inputs_colored": 0,
        "formulas_colored": 0,
        "crosssheet_colored": 0,
        "external_colored": 0,
        "total_cells": 0,
    }

    target_sheets = _resolve_sheets(wb, sheets)

    for ws in target_sheets:
        logger.debug("Processing sheet '%s'", ws.title)
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue

                summary["total_cells"] += 1
                category = _detect_cell_category(cell)

                # Determine base font properties
                is_header = cell.row == 1
                font_size = preset.header_font_size if is_header else preset.font_size
                bold = preset.header_bold if is_header else False

                if category in color_map:
                    cell.font = Font(
                        name=preset.font_name,
                        size=font_size,
                        bold=bold,
                        color=color_map[category],
                    )
                    _increment_summary(summary, category)
                else:
                    # Text or other — still apply base font
                    cell.font = Font(
                        name=preset.font_name,
                        size=font_size,
                        bold=bold,
                    )

    logger.info(
        "Financial preset applied — %d cells processed, %d inputs, "
        "%d formulas, %d crosssheet, %d external",
        summary["total_cells"],
        summary["inputs_colored"],
        summary["formulas_colored"],
        summary["crosssheet_colored"],
        summary["external_colored"],
    )
    return summary


def apply_number_formats(
    wb,
    preset: Optional[FinancialPreset] = None,
    currency_columns: Optional[list[str]] = None,
    percentage_columns: Optional[list[str]] = None,
    multiple_columns: Optional[list[str]] = None,
    year_columns: Optional[list[str]] = None,
    sheets: Optional[list[str]] = None,
) -> None:
    """Apply financial number formats to specified columns.

    Columns are identified by their letter (e.g. ``"B"``, ``"D"``).  Only
    cells that already contain a value are reformatted.

    Args:
        wb: openpyxl ``Workbook`` object.
        preset: :class:`FinancialPreset` config (uses defaults if *None*).
        currency_columns: Column letters to receive the currency format.
        percentage_columns: Column letters to receive the percentage format.
        multiple_columns: Column letters to receive the multiple format.
        year_columns: Column letters to receive the year (text) format.
        sheets: Specific sheet names to process (*None* → all sheets).
    """
    if preset is None:
        preset = FinancialPreset()

    format_map: dict[str, str] = {}
    for col in (currency_columns or []):
        format_map[col.upper()] = preset.currency_format
    for col in (percentage_columns or []):
        format_map[col.upper()] = preset.percentage_format
    for col in (multiple_columns or []):
        format_map[col.upper()] = preset.multiple_format
    for col in (year_columns or []):
        format_map[col.upper()] = preset.year_format

    if not format_map:
        logger.debug("No columns specified for number formatting; nothing to do.")
        return

    target_sheets = _resolve_sheets(wb, sheets)

    for ws in target_sheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                col_letter = _column_letter(cell.column)
                if col_letter in format_map:
                    cell.number_format = format_map[col_letter]

    logger.info("Number formats applied to columns: %s", sorted(format_map))


# ---------------------------------------------------------------------------
# Internal utilities
# ---------------------------------------------------------------------------

def _resolve_sheets(wb, sheets: Optional[list[str]] = None) -> list:
    """Return worksheet objects for the requested sheet names."""
    if sheets is None:
        return list(wb.worksheets)
    resolved = []
    for name in sheets:
        if name in wb.sheetnames:
            resolved.append(wb[name])
        else:
            logger.warning("Sheet '%s' not found in workbook — skipping.", name)
    return resolved


def _column_letter(col_index: int) -> str:
    """Convert a 1-based column index to its letter representation."""
    from openpyxl.utils import get_column_letter

    return get_column_letter(col_index)


def _increment_summary(summary: dict, category: str) -> None:
    """Increment the appropriate summary counter for *category*."""
    key_map = {
        "input": "inputs_colored",
        "formula": "formulas_colored",
        "crosssheet": "crosssheet_colored",
        "external": "external_colored",
    }
    key = key_map.get(category)
    if key:
        summary[key] += 1

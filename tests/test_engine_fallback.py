"""
Tests for engine fallback paths — layer cascade, retry logic,
error recovery, partial completion, and the EngineResult model.
"""

from __future__ import annotations

import time
from dataclasses import dataclass, field
from pathlib import Path
from unittest.mock import MagicMock, patch, PropertyMock

import pytest

from excel_engine.config import EngineConfig, TaskType, Layer, TASK_LAYER_MAP
from excel_engine.engine import ExcelEngine, EngineResult
from excel_engine.parsers.task_extractor import Task
from excel_engine.planner.task_planner import ExecutionPlan, Section
from excel_engine.recovery import ErrorClassifier, RecoveryStrategy, TaskError
from excel_engine.recalc import RecalcResult


# ── Helpers ──


def _make_task(
    task_id="t1",
    task_type=TaskType.CELL_VALUE,
    description="test task",
    sheet="Sheet1",
    cell="A1",
    value="100",
    formula=None,
    params=None,
    range_=None,
):
    return Task(
        id=task_id,
        task_type=task_type,
        description=description,
        sheet=sheet,
        cell=cell,
        range=range_,
        value=value,
        formula=formula,
        params=params or {},
    )


def _make_plan(tasks, section_id="s1", section_name="Section 1"):
    section = Section(id=section_id, name=section_name, sheet="Sheet1", tasks=tasks)
    plan = MagicMock(spec=ExecutionPlan)
    plan.sections = [section]
    plan.section_count = 1
    plan.total_tasks = len(tasks)
    plan.summary.return_value = "Plan: 1 section"
    return plan


# ── EngineResult ──


class TestEngineResult:
    def test_success_summary(self, tmp_path):
        r = EngineResult(
            success=True,
            workbook_path=tmp_path / "test.xlsx",
            sections_completed=2,
            sections_total=2,
            tasks_completed=5,
            tasks_total=5,
            elapsed_seconds=3.5,
        )
        s = r.summary()
        assert "SUCCESS" in s
        assert "2/2" in s
        assert "5/5" in s

    def test_failure_summary_with_errors(self, tmp_path):
        r = EngineResult(
            success=False,
            workbook_path=tmp_path / "test.xlsx",
            sections_completed=1,
            sections_total=2,
            tasks_completed=3,
            tasks_total=5,
            errors=["Task t1 failed", "Task t2 failed"],
            failed_tasks=["t1", "t2"],
            elapsed_seconds=5.0,
        )
        s = r.summary()
        assert "FAILED" in s
        assert "1/2" in s
        assert "t1" in s

    def test_summary_with_task_errors(self, tmp_path):
        r = EngineResult(
            success=False,
            workbook_path=tmp_path / "test.xlsx",
            sections_completed=0,
            sections_total=1,
            tasks_completed=0,
            tasks_total=1,
            task_errors=[
                TaskError("t1", TaskType.CELL_VALUE, Layer.OPENPYXL, "transient", "timeout", 0.0),
                TaskError("t1", TaskType.CELL_VALUE, Layer.XLWINGS, "permanent", "crash", 0.0),
                TaskError("t2", TaskType.FORMULA, Layer.OPENPYXL, "layer_incompatible", "n/a", 0.0),
            ],
        )
        s = r.summary()
        assert "transient=1" in s
        assert "permanent=1" in s
        assert "layer_incompatible=1" in s

    def test_summary_with_formula_errors(self, tmp_path):
        recalc = RecalcResult(success=True, total_formulas=10, total_errors=0)
        r = EngineResult(
            success=True,
            workbook_path=tmp_path / "test.xlsx",
            sections_completed=1,
            sections_total=1,
            tasks_completed=1,
            tasks_total=1,
            formula_errors=recalc,
        )
        s = r.summary()
        assert "10 formulas" in s

    def test_summary_formula_errors_present(self, tmp_path):
        recalc = RecalcResult(
            success=True,
            total_formulas=10,
            total_errors=2,
            error_summary={"#REF!": {"count": 2}},
        )
        r = EngineResult(
            success=True,
            workbook_path=tmp_path / "test.xlsx",
            sections_completed=1,
            sections_total=1,
            tasks_completed=1,
            tasks_total=1,
            formula_errors=recalc,
        )
        s = r.summary()
        assert "2 errors" in s
        assert "#REF!" in s

    def test_summary_formula_skipped(self, tmp_path):
        recalc = RecalcResult(success=False, skipped=True, warning="LibreOffice not found")
        r = EngineResult(
            success=True,
            workbook_path=tmp_path / "test.xlsx",
            sections_completed=1,
            sections_total=1,
            tasks_completed=1,
            tasks_total=1,
            formula_errors=recalc,
        )
        s = r.summary()
        assert "skipped" in s


# ── ErrorClassifier ──


class TestErrorClassifier:
    def test_transient_timeout(self):
        exc = Exception("AppleEvent timed out after 30s")
        assert ErrorClassifier.classify(exc, Layer.APPLESCRIPT) == "transient"

    def test_transient_not_responding(self):
        exc = Exception("Excel is not responding")
        assert ErrorClassifier.classify(exc, Layer.XLWINGS) == "transient"

    def test_transient_busy(self):
        exc = Exception("Application is busy")
        assert ErrorClassifier.classify(exc, Layer.SYSTEM_EVENTS) == "transient"

    def test_transient_connection(self):
        exc = Exception("connection refused")
        assert ErrorClassifier.classify(exc, Layer.XLWINGS) == "transient"

    def test_permanent_file_not_found(self):
        exc = FileNotFoundError("No such file")
        assert ErrorClassifier.classify(exc, Layer.OPENPYXL) == "permanent"

    def test_permanent_permission(self):
        exc = PermissionError("Access denied")
        assert ErrorClassifier.classify(exc, Layer.OPENPYXL) == "permanent"

    def test_permanent_no_such_sheet(self):
        exc = Exception("No such sheet 'Missing'")
        assert ErrorClassifier.classify(exc, Layer.OPENPYXL) == "permanent"

    def test_permanent_not_supported(self):
        exc = Exception("Feature not supported")
        assert ErrorClassifier.classify(exc, Layer.OPENPYXL) == "permanent"

    def test_layer_incompatible_not_implemented(self):
        exc = NotImplementedError("not implemented for chart_histogram")
        assert ErrorClassifier.classify(exc, Layer.OPENPYXL) == "layer_incompatible"

    def test_layer_incompatible_attribute_error(self):
        exc = AttributeError("'NoneType' has no attribute 'x'")
        assert ErrorClassifier.classify(exc, Layer.XLWINGS) == "layer_incompatible"

    def test_unknown_error_is_permanent(self):
        exc = ValueError("some unknown error")
        assert ErrorClassifier.classify(exc, Layer.OPENPYXL) == "permanent"

    def test_transient_system_events_error(self):
        exc = Exception("System Events got an error: can't get element")
        assert ErrorClassifier.classify(exc, Layer.SYSTEM_EVENTS) == "transient"

    def test_transient_app_not_running(self):
        exc = Exception("Application isn't running")
        assert ErrorClassifier.classify(exc, Layer.APPLESCRIPT) == "transient"


# ── RecoveryStrategy ──


class TestRecoveryStrategy:
    def test_should_retry_transient(self):
        s = RecoveryStrategy(max_retries=3)
        assert s.should_retry("transient", 0) is True
        assert s.should_retry("transient", 1) is True
        assert s.should_retry("transient", 2) is True
        assert s.should_retry("transient", 3) is False

    def test_should_not_retry_permanent(self):
        s = RecoveryStrategy(max_retries=3)
        assert s.should_retry("permanent", 0) is False

    def test_should_not_retry_layer_incompatible(self):
        s = RecoveryStrategy(max_retries=3)
        assert s.should_retry("layer_incompatible", 0) is False

    def test_get_delay_increases(self):
        s = RecoveryStrategy(base_delay=1.0, max_delay=30.0)
        d0 = s.get_delay(0)
        d1 = s.get_delay(1)
        d2 = s.get_delay(2)
        # base * 2^attempt is 1, 2, 4 before jitter
        assert d0 >= 1.0
        assert d1 >= 2.0
        assert d2 >= 4.0

    def test_get_delay_capped(self):
        s = RecoveryStrategy(base_delay=1.0, max_delay=5.0)
        d10 = s.get_delay(10)
        # 1 * 2^10 = 1024, but capped at 5 + jitter
        assert d10 <= 5.0 * 1.1 + 0.1


# ── EngineConfig ──


class TestEngineConfig:
    def test_get_layers_for_known_task(self):
        config = EngineConfig()
        layers = config.get_layers_for_task(TaskType.CELL_VALUE)
        assert Layer.OPENPYXL in layers

    def test_get_layers_for_openpyxl_only_task(self):
        config = EngineConfig(layer_order=[Layer.OPENPYXL])
        layers = config.get_layers_for_task(TaskType.CELL_VALUE)
        assert layers == [Layer.OPENPYXL]

    def test_get_layers_empty_for_unavailable(self):
        config = EngineConfig(layer_order=[Layer.PYAUTOGUI])
        layers = config.get_layers_for_task(TaskType.CELL_VALUE)
        # CELL_VALUE maps to OPENPYXL, XLWINGS, APPLESCRIPT — none is PYAUTOGUI
        assert layers == []

    def test_get_layers_respects_order(self):
        config = EngineConfig(
            layer_order=[Layer.APPLESCRIPT, Layer.OPENPYXL, Layer.XLWINGS]
        )
        layers = config.get_layers_for_task(TaskType.CELL_VALUE)
        assert layers[0] == Layer.APPLESCRIPT

    def test_get_layers_unknown_task_type(self):
        """A task type not in the map returns empty list."""
        config = EngineConfig()
        # Create a mock task type that's not in the map
        mock_tt = MagicMock()
        mock_tt.value = "unknown_task_xyz"
        layers = config.get_layers_for_task(mock_tt)
        assert layers == []


# ── Engine._execute_task — Layer Cascade ──


class TestExecuteTask:
    @patch("time.sleep")  # skip actual delays in retry
    def test_success_on_first_layer(self, mock_sleep):
        engine = ExcelEngine()
        task = _make_task()
        wb = Path("/fake/wb.xlsx")

        with patch.object(engine, "_dispatch_task") as mock_dispatch:
            mock_dispatch.return_value = None  # success
            ok, errors = engine._execute_task(task, wb)

        assert ok is True
        assert len(errors) == 0

    @patch("time.sleep")
    def test_cascade_to_second_layer(self, mock_sleep):
        """First layer raises NotImplementedError, falls through to second."""
        config = EngineConfig(layer_order=[
            Layer.OPENPYXL, Layer.XLWINGS, Layer.APPLESCRIPT,
            Layer.VBA, Layer.SYSTEM_EVENTS, Layer.PYAUTOGUI,
        ])
        engine = ExcelEngine(config=config)
        task = _make_task(task_type=TaskType.CELL_VALUE)
        wb = Path("/fake/wb.xlsx")

        call_count = 0

        def side_effect(t, layer, w):
            nonlocal call_count
            call_count += 1
            if layer == Layer.OPENPYXL:
                raise NotImplementedError("not supported here")
            # xlwings succeeds

        with patch.object(engine, "_dispatch_task", side_effect=side_effect):
            ok, errors = engine._execute_task(task, wb)

        assert ok is True
        assert call_count == 2
        assert len(errors) == 1
        # NotImplementedError is classified as "layer_incompatible"
        assert errors[0].error_type in ("layer_incompatible", "permanent")

    @patch("time.sleep")
    def test_all_layers_fail(self, mock_sleep):
        engine = ExcelEngine()
        task = _make_task(task_type=TaskType.CELL_VALUE)
        wb = Path("/fake/wb.xlsx")

        with patch.object(engine, "_dispatch_task",
                          side_effect=NotImplementedError("nope")):
            ok, errors = engine._execute_task(task, wb)

        assert ok is False
        assert len(errors) > 0

    @patch("time.sleep")
    def test_transient_retry_then_success(self, mock_sleep):
        """Transient error triggers retry, then succeeds."""
        config = EngineConfig(
            max_retries=3, retry_delay=0.01,
            layer_order=[Layer.OPENPYXL, Layer.XLWINGS, Layer.APPLESCRIPT,
                         Layer.VBA, Layer.SYSTEM_EVENTS, Layer.PYAUTOGUI],
        )
        engine = ExcelEngine(config=config)
        task = _make_task(task_type=TaskType.CELL_VALUE)
        wb = Path("/fake/wb.xlsx")

        attempt = 0

        def side_effect(t, layer, w):
            nonlocal attempt
            attempt += 1
            if attempt <= 2:
                raise Exception("AppleEvent timed out")
            # succeed on 3rd attempt

        with patch.object(engine, "_dispatch_task", side_effect=side_effect):
            ok, errors = engine._execute_task(task, wb)

        assert ok is True
        assert attempt == 3
        # 2 transient errors recorded
        transient = [e for e in errors if e.error_type == "transient"]
        assert len(transient) == 2

    @patch("time.sleep")
    def test_transient_exhausts_retries_then_cascades(self, mock_sleep):
        """Transient errors exhaust retries then cascade to next layer."""
        config = EngineConfig(
            max_retries=2, retry_delay=0.01,
            layer_order=[Layer.OPENPYXL, Layer.XLWINGS, Layer.APPLESCRIPT,
                         Layer.VBA, Layer.SYSTEM_EVENTS, Layer.PYAUTOGUI],
        )
        engine = ExcelEngine(config=config)
        task = _make_task(task_type=TaskType.CELL_VALUE)
        wb = Path("/fake/wb.xlsx")

        layers_tried = []

        def side_effect(t, layer, w):
            layers_tried.append(layer)
            if layer == Layer.OPENPYXL:
                raise Exception("Application is busy")
            # xlwings succeeds

        with patch.object(engine, "_dispatch_task", side_effect=side_effect):
            ok, errors = engine._execute_task(task, wb)

        assert ok is True
        # OPENPYXL retried max_retries times, then cascaded to XLWINGS
        openpyxl_attempts = [l for l in layers_tried if l == Layer.OPENPYXL]
        assert len(openpyxl_attempts) == config.max_retries + 1  # initial + retries
        assert Layer.XLWINGS in layers_tried

    @patch("time.sleep")
    def test_no_layers_for_task(self, mock_sleep):
        """Task type with no layers returns failure immediately."""
        config = EngineConfig(layer_order=[Layer.PYAUTOGUI])
        engine = ExcelEngine(config=config)
        task = _make_task(task_type=TaskType.CELL_VALUE)
        wb = Path("/fake/wb.xlsx")

        ok, errors = engine._execute_task(task, wb)
        assert ok is False


# ── Engine._dispatch_task ──


class TestDispatchTask:
    def test_unknown_layer_raises(self):
        engine = ExcelEngine()
        task = _make_task()
        wb = Path("/fake/wb.xlsx")

        with pytest.raises(ValueError, match="Unknown layer"):
            engine._dispatch_task(task, MagicMock(), wb)

    def test_dispatch_to_openpyxl(self):
        engine = ExcelEngine()
        task = _make_task()
        wb = Path("/fake/wb.xlsx")

        with patch.object(engine, "_exec_openpyxl") as mock:
            engine._dispatch_task(task, Layer.OPENPYXL, wb)
            mock.assert_called_once_with(task, wb)

    def test_dispatch_to_xlwings(self):
        engine = ExcelEngine()
        task = _make_task()
        wb = Path("/fake/wb.xlsx")

        with patch.object(engine, "_exec_xlwings") as mock:
            engine._dispatch_task(task, Layer.XLWINGS, wb)
            mock.assert_called_once_with(task, wb)

    def test_dispatch_to_applescript(self):
        engine = ExcelEngine()
        task = _make_task()
        wb = Path("/fake/wb.xlsx")

        with patch.object(engine, "_exec_applescript") as mock:
            engine._dispatch_task(task, Layer.APPLESCRIPT, wb)
            mock.assert_called_once_with(task, wb)

    def test_dispatch_to_system_events(self):
        engine = ExcelEngine()
        task = _make_task()
        wb = Path("/fake/wb.xlsx")

        with patch.object(engine, "_exec_system_events") as mock:
            engine._dispatch_task(task, Layer.SYSTEM_EVENTS, wb)
            mock.assert_called_once_with(task, wb)

    def test_dispatch_to_vba(self):
        engine = ExcelEngine()
        task = _make_task()
        wb = Path("/fake/wb.xlsx")

        with patch.object(engine, "_exec_vba") as mock:
            engine._dispatch_task(task, Layer.VBA, wb)
            mock.assert_called_once_with(task, wb)

    def test_dispatch_to_pyautogui(self):
        engine = ExcelEngine()
        task = _make_task()
        wb = Path("/fake/wb.xlsx")

        with patch.object(engine, "_exec_pyautogui") as mock:
            engine._dispatch_task(task, Layer.PYAUTOGUI, wb)
            mock.assert_called_once_with(task, wb)


# ── Engine._exec_openpyxl task type dispatch ──


class TestExecOpenpyxlDispatch:
    @pytest.fixture(autouse=True)
    def setup_engine(self, tmp_path):
        self.engine = ExcelEngine()
        self.wb_path = tmp_path / "test.xlsx"
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "Name"
        ws["B1"] = "Value"
        for i in range(2, 12):
            ws[f"A{i}"] = f"Item{i}"
            ws[f"B{i}"] = i * 10
        wb.save(str(self.wb_path))
        wb.close()

    def test_formula_task(self):
        task = _make_task(task_type=TaskType.FORMULA, cell="C1", formula="=SUM(B2:B11)")
        self.engine._exec_openpyxl(task, self.wb_path)
        assert self.engine._openpyxl.get_value("C1", sheet="Sheet1") == "=SUM(B2:B11)"

    def test_cell_value_task(self):
        task = _make_task(task_type=TaskType.CELL_VALUE, cell="C1", value="Hello")
        self.engine._exec_openpyxl(task, self.wb_path)
        assert self.engine._openpyxl.get_value("C1", sheet="Sheet1") == "Hello"

    def test_table_create_task(self):
        task = _make_task(
            task_type=TaskType.TABLE_CREATE,
            params={"name": "MyTable", "ref": "A1:B11"},
        )
        self.engine._exec_openpyxl(task, self.wb_path)

    def test_number_format_task(self):
        task = _make_task(
            task_type=TaskType.NUMBER_FORMAT,
            cell="B2", range_=None,
            params={"format": "$#,##0.00"},
        )
        self.engine._exec_openpyxl(task, self.wb_path)

    def test_number_format_range_task(self):
        task = _make_task(
            task_type=TaskType.NUMBER_FORMAT,
            cell=None, range_="B2:B11",
            params={"format": "0.00%"},
        )
        self.engine._exec_openpyxl(task, self.wb_path)

    def test_font_task(self):
        task = _make_task(
            task_type=TaskType.FONT, cell="A1",
            params={"bold": True, "size": 14, "color": "FF0000"},
        )
        self.engine._exec_openpyxl(task, self.wb_path)

    def test_fill_task(self):
        task = _make_task(
            task_type=TaskType.FILL, cell="A1",
            params={"color": "FFFF00"},
        )
        self.engine._exec_openpyxl(task, self.wb_path)

    def test_alignment_task(self):
        task = _make_task(
            task_type=TaskType.ALIGNMENT, cell="A1",
            params={"horizontal": "center", "vertical": "top"},
        )
        self.engine._exec_openpyxl(task, self.wb_path)

    def test_border_task(self):
        task = _make_task(
            task_type=TaskType.BORDER, cell="A1",
            params={"style": "thick"},
        )
        self.engine._exec_openpyxl(task, self.wb_path)

    def test_column_width_task(self):
        task = _make_task(
            task_type=TaskType.COLUMN_WIDTH, cell="B1",
            params={"width": 20},
        )
        self.engine._exec_openpyxl(task, self.wb_path)

    def test_column_width_explicit_column(self):
        task = _make_task(
            task_type=TaskType.COLUMN_WIDTH, cell=None,
            params={"column": "C", "width": 25},
        )
        self.engine._exec_openpyxl(task, self.wb_path)

    def test_conditional_format_task(self):
        task = _make_task(
            task_type=TaskType.CONDITIONAL_FORMAT,
            range_="B2:B11",
            params={"operator": "greaterThan", "formula": ["50"]},
        )
        self.engine._exec_openpyxl(task, self.wb_path)

    def test_chart_bar_task(self):
        task = _make_task(
            task_type=TaskType.CHART_BAR,
            params={"title": "Bar", "data_range": "B1:B11"},
        )
        self.engine._exec_openpyxl(task, self.wb_path)

    def test_chart_line_task(self):
        task = _make_task(
            task_type=TaskType.CHART_LINE,
            params={"title": "Line"},
        )
        self.engine._exec_openpyxl(task, self.wb_path)

    def test_chart_pie_task(self):
        task = _make_task(
            task_type=TaskType.CHART_PIE,
            params={"title": "Pie"},
        )
        self.engine._exec_openpyxl(task, self.wb_path)

    def test_chart_scatter_task(self):
        task = _make_task(
            task_type=TaskType.CHART_SCATTER,
            params={"title": "Scatter", "x_range": "A2:A11", "y_range": "B2:B11"},
        )
        self.engine._exec_openpyxl(task, self.wb_path)

    def test_chart_area_task(self):
        task = _make_task(
            task_type=TaskType.CHART_AREA,
            params={"title": "Area"},
        )
        self.engine._exec_openpyxl(task, self.wb_path)

    def test_chart_combo_task(self):
        ws = self.engine._openpyxl.wb or None
        # Need to open workbook first to add data
        task = _make_task(
            task_type=TaskType.CHART_COMBO,
            params={
                "title": "Combo",
                "bar_data_range": "B1:B11",
                "line_data_range": "B1:B11",
            },
        )
        self.engine._exec_openpyxl(task, self.wb_path)

    def test_named_range_task(self):
        task = _make_task(
            task_type=TaskType.NAMED_RANGE,
            params={"name": "MyRange"},
            range_="$A$1:$B$11",
        )
        self.engine._exec_openpyxl(task, self.wb_path)

    def test_data_validation_list_task(self):
        task = _make_task(
            task_type=TaskType.DATA_VALIDATION,
            cell="C2", range_=None,
            params={"type": "list", "formula1": '"Yes,No"'},
        )
        self.engine._exec_openpyxl(task, self.wb_path)

    def test_data_validation_values_shortcut(self):
        task = _make_task(
            task_type=TaskType.DATA_VALIDATION,
            cell="C2", range_=None,
            params={"type": "list", "values": ["A", "B", "C"]},
        )
        self.engine._exec_openpyxl(task, self.wb_path)

    def test_freeze_panes_task(self):
        task = _make_task(task_type=TaskType.FREEZE_PANES, cell="A2")
        self.engine._exec_openpyxl(task, self.wb_path)

    def test_autofilter_task(self):
        task = _make_task(
            task_type=TaskType.AUTOFILTER,
            range_="A1:B11",
        )
        self.engine._exec_openpyxl(task, self.wb_path)

    def test_sheet_create_task(self):
        task = _make_task(
            task_type=TaskType.SHEET_CREATE,
            params={"name": "NewSheet"},
        )
        self.engine._exec_openpyxl(task, self.wb_path)

    def test_sheet_rename_task(self):
        task = _make_task(
            task_type=TaskType.SHEET_RENAME,
            params={"old_name": "Sheet1", "new_name": "Renamed"},
        )
        self.engine._exec_openpyxl(task, self.wb_path)

    def test_merge_cells_task(self):
        task = _make_task(task_type=TaskType.MERGE_CELLS, range_="A1:B1")
        self.engine._exec_openpyxl(task, self.wb_path)

    def test_print_settings_landscape(self):
        task = _make_task(
            task_type=TaskType.PRINT_SETTINGS,
            params={"landscape": True, "print_area": "A1:B11"},
        )
        self.engine._exec_openpyxl(task, self.wb_path)

    def test_text_function_task(self):
        task = _make_task(
            task_type=TaskType.TEXT_FUNCTION,
            cell="C1", formula="=UPPER(A1)",
        )
        self.engine._exec_openpyxl(task, self.wb_path)

    def test_lookup_function_value_fallback(self):
        """When formula is None but value is set, use set_value."""
        task = _make_task(
            task_type=TaskType.LOOKUP_FUNCTION,
            cell="C1", formula=None, value="FallbackVal",
        )
        self.engine._exec_openpyxl(task, self.wb_path)

    def test_table_total_row_task(self):
        # First create a table
        create_task = _make_task(
            task_type=TaskType.TABLE_CREATE,
            params={"name": "T1", "ref": "A1:B11"},
        )
        self.engine._exec_openpyxl(create_task, self.wb_path)

        task = _make_task(
            task_type=TaskType.TABLE_TOTAL_ROW,
            params={"name": "T1"},
        )
        self.engine._exec_openpyxl(task, self.wb_path)

    def test_row_height_task(self):
        task = _make_task(
            task_type=TaskType.ROW_HEIGHT,
            cell="A5",
            params={"size": 30},
        )
        self.engine._exec_openpyxl(task, self.wb_path)

    def test_tab_color_task(self):
        task = _make_task(
            task_type=TaskType.TAB_COLOR,
            params={"color": "00FF00"},
        )
        self.engine._exec_openpyxl(task, self.wb_path)

    def test_page_break_task(self):
        task = _make_task(
            task_type=TaskType.PAGE_BREAK,
            cell="A10",
        )
        self.engine._exec_openpyxl(task, self.wb_path)

    def test_hyperlink_task(self):
        task = _make_task(
            task_type=TaskType.HYPERLINK,
            cell="A1",
            params={"url": "https://example.com", "display": "Example"},
        )
        self.engine._exec_openpyxl(task, self.wb_path)

    def test_sheet_move_task(self):
        # Create a second sheet so move makes sense
        create_task = _make_task(
            task_type=TaskType.SHEET_CREATE,
            params={"name": "Sheet2"},
        )
        self.engine._exec_openpyxl(create_task, self.wb_path)

        task = _make_task(
            task_type=TaskType.SHEET_MOVE,
            params={"position": -1},
        )
        task.sheet = "Sheet2"
        self.engine._exec_openpyxl(task, self.wb_path)

    def test_sheet_copy_task(self):
        task = _make_task(
            task_type=TaskType.SHEET_COPY,
            params={"new_name": "Sheet1 Copy"},
        )
        task.sheet = "Sheet1"
        self.engine._exec_openpyxl(task, self.wb_path)

    def test_formatting_task_with_bold(self):
        task = _make_task(
            task_type=TaskType.FORMATTING,
            cell="A1",
            params={"bold": True, "font_size": 14},
        )
        self.engine._exec_openpyxl(task, self.wb_path)

    def test_unknown_task_type_raises(self):
        task = _make_task(task_type=TaskType.GOAL_SEEK)
        with pytest.raises(NotImplementedError, match="not implemented"):
            self.engine._exec_openpyxl(task, self.wb_path)


# ── Engine._exec_pyautogui ──


class TestExecPyautogui:
    def test_raises_not_implemented(self):
        engine = ExcelEngine()
        task = _make_task()
        with pytest.raises(NotImplementedError, match="PyAutoGUI"):
            engine._exec_pyautogui(task, Path("/fake.xlsx"))


# ── Engine.run — full pipeline ──


class TestEngineRun:
    def test_run_with_no_input_raises(self):
        engine = ExcelEngine()
        result = engine.run(workbook=Path("/fake/wb.xlsx"))
        assert result.success is False
        assert any("instructions" in e.lower() or "Provide" in e for e in result.errors)

    def test_run_with_instruction_text(self, tmp_path):
        wb = tmp_path / "test.xlsx"
        from openpyxl import Workbook
        workbook = Workbook()
        workbook.save(str(wb))
        workbook.close()

        engine = ExcelEngine(config=EngineConfig(
            verify_after_each_section=False,
            recalculate_formulas=False,
        ))

        with patch.object(engine._extractor, "extract", return_value=[]), \
             patch.object(engine._planner, "plan") as mock_plan:
            mock_plan.return_value = MagicMock(
                sections=[], section_count=0, total_tasks=0,
                summary=MagicMock(return_value="empty"),
            )
            result = engine.run(workbook=wb, instruction_text="Do nothing")

        assert result.success is True

    def test_run_with_tasks_list(self, tmp_path):
        wb = tmp_path / "test.xlsx"
        from openpyxl import Workbook
        workbook = Workbook()
        workbook.save(str(wb))
        workbook.close()

        engine = ExcelEngine(config=EngineConfig(
            verify_after_each_section=False,
            recalculate_formulas=False,
        ))

        tasks = [_make_task()]

        with patch.object(engine._planner, "plan") as mock_plan, \
             patch.object(engine, "_execute_section", return_value=True):
            section = MagicMock()
            section.tasks = tasks
            section.task_count = 1
            mock_plan.return_value = MagicMock(
                sections=[section], section_count=1, total_tasks=1,
                summary=MagicMock(return_value="1 section"),
            )
            result = engine.run(workbook=wb, tasks=tasks)


# ── Engine._save_workbook ──


class TestSaveWorkbook:
    def test_save_via_openpyxl(self, tmp_path):
        engine = ExcelEngine()
        engine._openpyxl.wb = MagicMock()
        with patch.object(engine._openpyxl, "save") as mock_save:
            engine._save_workbook(tmp_path / "test.xlsx")
            mock_save.assert_called_once()

    def test_save_via_xlwings(self, tmp_path):
        engine = ExcelEngine()
        engine._openpyxl.wb = None
        engine._xlwings._wb = MagicMock()
        engine._save_workbook(tmp_path / "test.xlsx")

    def test_save_via_applescript_fallback(self, tmp_path):
        engine = ExcelEngine()
        engine._openpyxl.wb = None
        engine._xlwings._wb = None
        with patch.object(engine._applescript, "save"):
            engine._save_workbook(tmp_path / "test.xlsx")

    def test_save_all_fail(self, tmp_path):
        engine = ExcelEngine()
        engine._openpyxl.wb = None
        engine._xlwings._wb = None
        with patch.object(engine._applescript, "save", side_effect=Exception("fail")):
            engine._save_workbook(tmp_path / "test.xlsx")  # logs warning, doesn't raise


# ── Engine._cleanup ──


class TestCleanup:
    def test_cleanup_no_errors(self):
        engine = ExcelEngine()
        with patch.object(engine._openpyxl, "close"), \
             patch.object(engine._xlwings, "disconnect"), \
             patch.object(engine._verifier, "close"):
            engine._cleanup()

    def test_cleanup_with_errors(self):
        engine = ExcelEngine()
        with patch.object(engine._openpyxl, "close", side_effect=Exception("fail")), \
             patch.object(engine._xlwings, "disconnect", side_effect=Exception("fail")), \
             patch.object(engine._verifier, "close", side_effect=Exception("fail")):
            engine._cleanup()  # should not raise


# ── Engine._execute_section ──


class TestExecuteSection:
    @patch("time.sleep")
    def test_section_with_all_tasks_passing(self, mock_sleep, tmp_path):
        engine = ExcelEngine(config=EngineConfig(verify_after_each_section=False))
        t1 = _make_task("t1")
        t2 = _make_task("t2")
        tasks = [t1, t2]
        section = Section(id="s1", name="Sec1", sheet="Sheet1", tasks=tasks)
        result = EngineResult(
            success=False,
            workbook_path=tmp_path / "test.xlsx",
            sections_completed=0,
            sections_total=1,
            tasks_completed=0,
            tasks_total=2,
        )

        with patch.object(engine, "_execute_task", return_value=(True, [])):
            ok = engine._execute_section(section, tmp_path / "test.xlsx", result)

        assert ok is True
        assert result.tasks_completed == 2

    @patch("time.sleep")
    def test_section_with_task_failure(self, mock_sleep, tmp_path):
        engine = ExcelEngine(config=EngineConfig(verify_after_each_section=False))
        t1 = _make_task("t1")
        t2 = _make_task("t2")
        tasks = [t1, t2]
        section = Section(id="s1", name="Sec1", sheet="Sheet1", tasks=tasks)
        result = EngineResult(
            success=False,
            workbook_path=tmp_path / "test.xlsx",
            sections_completed=0,
            sections_total=1,
            tasks_completed=0,
            tasks_total=2,
        )

        returns = [(True, []), (False, [TaskError("t2", TaskType.CELL_VALUE, Layer.OPENPYXL, "permanent", "err", 0.0)])]

        with patch.object(engine, "_execute_task", side_effect=returns):
            ok = engine._execute_section(section, tmp_path / "test.xlsx", result)

        assert ok is False
        assert result.tasks_completed == 1
        assert "t2" in result.failed_tasks

    @patch("time.sleep")
    def test_section_verification(self, mock_sleep, tmp_path):
        """Section verification runs when configured."""
        wb_path = tmp_path / "test.xlsx"
        from openpyxl import Workbook
        wb = Workbook()
        wb.save(str(wb_path))
        wb.close()

        config = EngineConfig(verify_after_each_section=True, recalculate_formulas=False)
        engine = ExcelEngine(config=config)
        t1 = _make_task("t1")
        tasks = [t1]
        section = Section(id="s1", name="Sec1", sheet="Sheet1", tasks=tasks)
        result = EngineResult(
            success=False,
            workbook_path=wb_path,
            sections_completed=0,
            sections_total=1,
            tasks_completed=0,
            tasks_total=1,
        )

        mock_sv = MagicMock()
        mock_sv.all_passed = True

        with patch.object(engine, "_execute_task", return_value=(True, [])), \
             patch.object(engine, "_save_workbook"), \
             patch.object(engine._verifier, "load"), \
             patch.object(engine._verifier, "verify_section", return_value=mock_sv), \
             patch.object(engine._verifier, "close"):
            ok = engine._execute_section(section, wb_path, result)

        assert ok is True
        assert len(result.verifications) == 1

    @patch("time.sleep")
    def test_section_verification_failure(self, mock_sleep, tmp_path):
        """Section verification failure is logged as warning, not fatal."""
        wb_path = tmp_path / "test.xlsx"
        from openpyxl import Workbook
        wb = Workbook()
        wb.save(str(wb_path))
        wb.close()

        config = EngineConfig(verify_after_each_section=True, recalculate_formulas=False)
        engine = ExcelEngine(config=config)
        t1 = _make_task("t1")
        tasks = [t1]
        section = Section(id="s1", name="Sec1", sheet="Sheet1", tasks=tasks)
        result = EngineResult(
            success=False,
            workbook_path=wb_path,
            sections_completed=0,
            sections_total=1,
            tasks_completed=0,
            tasks_total=1,
        )

        with patch.object(engine, "_execute_task", return_value=(True, [])), \
             patch.object(engine, "_save_workbook"), \
             patch.object(engine._verifier, "load", side_effect=Exception("load error")):
            ok = engine._execute_section(section, wb_path, result)

        # Task succeeded but verification error is handled gracefully
        assert ok is True


# ── Engine.execute — recalculation paths ──


class TestEngineExecuteRecalc:
    @staticmethod
    def _make_xlsx(path):
        """Create a minimal valid xlsx for tests that only exercise higher-level logic."""
        from openpyxl import Workbook
        wb = Workbook()
        wb.save(str(path))

    @patch("time.sleep")
    def test_recalc_skipped_when_disabled(self, mock_sleep, tmp_path):
        config = EngineConfig(
            recalculate_formulas=False,
            verify_after_each_section=False,
        )
        engine = ExcelEngine(config=config)
        plan = MagicMock(sections=[], section_count=0, total_tasks=0)
        wb = tmp_path / "test.xlsx"
        self._make_xlsx(wb)

        with patch("excel_engine.engine.recalculate") as mock_recalc:
            result = engine.execute(plan, wb)
            mock_recalc.assert_not_called()

        assert result.success is True

    @patch("time.sleep")
    def test_recalc_runs_on_success(self, mock_sleep, tmp_path):
        config = EngineConfig(
            recalculate_formulas=True,
            verify_after_each_section=False,
        )
        engine = ExcelEngine(config=config)
        plan = MagicMock(sections=[], section_count=0, total_tasks=0)
        wb = tmp_path / "test.xlsx"
        self._make_xlsx(wb)

        recalc_result = RecalcResult(success=True, total_formulas=5, total_errors=0)
        with patch("excel_engine.engine.recalculate", return_value=recalc_result), \
             patch.object(engine, "_cleanup"):
            result = engine.execute(plan, wb)

        assert result.formula_errors is not None
        assert result.formula_errors.total_formulas == 5

    @patch("time.sleep")
    def test_recalc_exception_handled(self, mock_sleep, tmp_path):
        config = EngineConfig(
            recalculate_formulas=True,
            verify_after_each_section=False,
        )
        engine = ExcelEngine(config=config)
        plan = MagicMock(sections=[], section_count=0, total_tasks=0)
        wb = tmp_path / "test.xlsx"
        self._make_xlsx(wb)

        with patch("excel_engine.engine.recalculate", side_effect=Exception("boom")), \
             patch.object(engine, "_cleanup"):
            result = engine.execute(plan, wb)

        assert result.formula_errors is not None
        assert result.formula_errors.success is False


# ── Engine convenience methods ──


class TestEngineConvenience:
    def test_scan(self, tmp_path):
        engine = ExcelEngine()
        inst = tmp_path / "inst.txt"
        inst.write_text("Enter 100 in cell A1")

        with patch.object(engine._parser, "parse", return_value="text"), \
             patch.object(engine._extractor, "extract", return_value=[_make_task()]):
            tasks = engine.scan(inst)
            assert len(tasks) == 1

    def test_plan(self):
        engine = ExcelEngine()
        tasks = [_make_task()]
        with patch.object(engine._planner, "plan") as mock_plan:
            mock_plan.return_value = MagicMock()
            plan = engine.plan(tasks)
            assert plan is not None

    def test_properties(self):
        engine = ExcelEngine()
        assert engine.openpyxl is engine._openpyxl
        assert engine.xlwings is engine._xlwings
        assert engine.applescript is engine._applescript
        assert engine.system_events is engine._system_events
        assert engine.vba is engine._vba
        assert engine.pyautogui is engine._pyautogui
        assert engine.parser is engine._parser
        assert engine.extractor is engine._extractor
        assert engine.planner is engine._planner
        assert engine.verifier is engine._verifier

"""
Layer 1: openpyxl — Offline Excel file manipulation.

Handles formulas, tables, formatting, charts, named ranges, data validation,
freeze panes, autofilter, column widths, and sheet management without Excel open.

KEY RULES:
  - Insert raw data columns BEFORE creating a table.
  - Structural reference formulas ([@COL]) become #REF! in openpyxl — re-enter LIVE.
  - Histogram charts are cx:chart type — cannot be created with openpyxl.
  - Table style names must be exact (e.g., "TableStyleMedium5").
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.chart import (
    AreaChart,
    BarChart,
    LineChart,
    PieChart,
    Reference,
    ScatterChart,
    Series,
)
from openpyxl.formatting.rule import CellIsRule, FormulaRule, ColorScaleRule
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
    NamedStyle,
    numbers,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.workbook.defined_name import DefinedName

logger = logging.getLogger(__name__)


class OpenpyxlLayer:
    """Layer 1 — offline Excel file manipulation via openpyxl."""

    def __init__(self) -> None:
        self.wb: Optional[Workbook] = None
        self.path: Optional[Path] = None

    # ── Workbook lifecycle ──

    def open(self, path: Path) -> Workbook:
        """Load an existing workbook (preserves VBA if .xlsm)."""
        self.path = Path(path)
        keep_vba = self.path.suffix.lower() == ".xlsm"
        self.wb = load_workbook(str(self.path), keep_vba=keep_vba)
        logger.info("Opened workbook: %s (%d sheets)", self.path.name, len(self.wb.sheetnames))
        return self.wb

    def create(self, path: Path) -> Workbook:
        """Create a new workbook."""
        self.path = Path(path)
        self.wb = Workbook()
        logger.info("Created new workbook: %s", self.path.name)
        return self.wb

    def save(self, path: Optional[Path] = None) -> None:
        """Save the workbook to disk."""
        save_path = path or self.path
        if not save_path or not self.wb:
            raise ValueError("No workbook or path to save to")
        self.wb.save(str(save_path))
        logger.info("Saved workbook: %s", save_path.name)

    def close(self) -> None:
        """Close the workbook (save first if needed)."""
        if self.wb:
            self.wb.close()
            self.wb = None

    def _ws(self, sheet_name: Optional[str] = None):
        """Get a worksheet by name, or the active sheet."""
        if not self.wb:
            raise RuntimeError("No workbook is open")
        if sheet_name:
            return self.wb[sheet_name]
        return self.wb.active

    @staticmethod
    def _qualify_range(ws, rng: str) -> str:
        """Ensure a range string has a sheet qualifier (e.g. 'Sheet!A1:B10')."""
        if "!" not in rng:
            title = ws.title
            if " " in title:
                title = f"'{title}'"
            return f"{title}!{rng}"
        return rng

    # ── Sheet Management ──

    def create_sheet(self, name: str, index: Optional[int] = None) -> None:
        """Create a new worksheet."""
        if not self.wb:
            raise RuntimeError("No workbook is open")
        self.wb.create_sheet(title=name, index=index)
        logger.info("Created sheet: %s", name)

    def rename_sheet(self, old_name: str, new_name: str) -> None:
        """Rename a worksheet."""
        ws = self._ws(old_name)
        ws.title = new_name
        logger.info("Renamed sheet '%s' → '%s'", old_name, new_name)

    def delete_sheet(self, name: str) -> None:
        """Delete a worksheet."""
        if not self.wb:
            raise RuntimeError("No workbook is open")
        del self.wb[name]
        logger.info("Deleted sheet: %s", name)

    # ── Cell Operations ──

    def set_value(self, cell: str, value: Any, sheet: Optional[str] = None) -> None:
        """Set a cell's value."""
        ws = self._ws(sheet)
        ws[cell] = value

    def get_value(self, cell: str, sheet: Optional[str] = None) -> Any:
        """Get a cell's value."""
        ws = self._ws(sheet)
        return ws[cell].value

    def set_formula(self, cell: str, formula: str, sheet: Optional[str] = None) -> None:
        """
        Set a cell's formula. For structural table references ([@COL]),
        this will produce #REF! — use xlwings/AppleScript layer instead.
        """
        ws = self._ws(sheet)
        if not formula.startswith("="):
            formula = f"={formula}"
        ws[cell] = formula
        logger.debug("Set formula %s = %s", cell, formula)

    def set_values_range(
        self, start_cell: str, data: list[list[Any]], sheet: Optional[str] = None
    ) -> None:
        """Write a 2D array of values starting at a cell."""
        ws = self._ws(sheet)
        for row_offset, row_data in enumerate(data):
            for col_offset, value in enumerate(row_data):
                ws.cell(
                    row=ws[start_cell].row + row_offset,
                    column=ws[start_cell].column + col_offset,
                    value=value,
                )

    # ── Tables ──

    def create_table(
        self,
        name: str,
        ref: str,
        style: str = "TableStyleMedium5",
        sheet: Optional[str] = None,
        show_totals: bool = False,
    ) -> None:
        """
        Create an Excel table. Insert raw data columns BEFORE calling this.
        ref format: "A1:G20"
        """
        ws = self._ws(sheet)
        style_info = TableStyleInfo(
            name=style,
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table = Table(displayName=name, ref=ref)
        table.tableStyleInfo = style_info
        if show_totals:
            table.totalsRowShown = True
        ws.add_table(table)
        logger.info("Created table '%s' at %s with style %s", name, ref, style)

    # ── Formatting ──

    def set_font(
        self,
        cells: str,
        sheet: Optional[str] = None,
        name: str = "Calibri",
        size: int = 11,
        bold: bool = False,
        italic: bool = False,
        color: str = "000000",
        underline: Optional[str] = None,
    ) -> None:
        """Apply font to a range. cells can be 'A1:D5' or 'A1'."""
        ws = self._ws(sheet)
        font = Font(
            name=name, size=size, bold=bold, italic=italic,
            color=color, underline=underline,
        )
        for row in ws[cells]:
            if not isinstance(row, tuple):
                row = (row,)
            for cell in row:
                cell.font = font

    def set_fill(
        self, cells: str, color: str, sheet: Optional[str] = None,
        fill_type: str = "solid",
    ) -> None:
        """Apply a background fill to a range."""
        ws = self._ws(sheet)
        fill = PatternFill(start_color=color, end_color=color, fill_type=fill_type)
        for row in ws[cells]:
            if not isinstance(row, tuple):
                row = (row,)
            for cell in row:
                cell.fill = fill

    def set_alignment(
        self,
        cells: str,
        sheet: Optional[str] = None,
        horizontal: str = "general",
        vertical: str = "bottom",
        wrap_text: bool = False,
        indent: int = 0,
    ) -> None:
        """Apply alignment to a range."""
        ws = self._ws(sheet)
        alignment = Alignment(
            horizontal=horizontal, vertical=vertical,
            wrap_text=wrap_text, indent=indent,
        )
        for row in ws[cells]:
            if not isinstance(row, tuple):
                row = (row,)
            for cell in row:
                cell.alignment = alignment

    def set_number_format(
        self, cells: str, fmt: str, sheet: Optional[str] = None
    ) -> None:
        """Apply a number format to a range. Use codes from ExcelConstants.NUMBER_FORMATS."""
        ws = self._ws(sheet)
        for row in ws[cells]:
            if not isinstance(row, tuple):
                row = (row,)
            for cell in row:
                cell.number_format = fmt

    def set_border(
        self,
        cells: str,
        sheet: Optional[str] = None,
        style: str = "thin",
        color: str = "000000",
    ) -> None:
        """Apply borders to all edges of cells in a range."""
        ws = self._ws(sheet)
        side = Side(style=style, color=color)
        border = Border(left=side, right=side, top=side, bottom=side)
        for row in ws[cells]:
            if not isinstance(row, tuple):
                row = (row,)
            for cell in row:
                cell.border = border

    def set_column_width(
        self, column: str, width: float, sheet: Optional[str] = None
    ) -> None:
        """Set column width. column is a letter like 'A' or 'AB'."""
        ws = self._ws(sheet)
        ws.column_dimensions[column].width = width

    def set_row_height(
        self, row: int, height: float, sheet: Optional[str] = None
    ) -> None:
        """Set row height."""
        ws = self._ws(sheet)
        ws.row_dimensions[row].height = height

    def auto_column_widths(self, sheet: Optional[str] = None) -> None:
        """Auto-fit column widths based on content (approximate)."""
        ws = self._ws(sheet)
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    # ── Conditional Formatting ──

    def add_conditional_format_cell_is(
        self,
        range_str: str,
        operator: str,
        formula: list[str],
        font: Optional[Font] = None,
        fill: Optional[PatternFill] = None,
        sheet: Optional[str] = None,
    ) -> None:
        """Add a CellIs conditional format rule."""
        ws = self._ws(sheet)
        rule = CellIsRule(
            operator=operator,
            formula=formula,
            font=font,
            fill=fill,
        )
        ws.conditional_formatting.add(range_str, rule)
        logger.info("Added CellIs CF to %s: %s %s", range_str, operator, formula)

    def add_conditional_format_formula(
        self,
        range_str: str,
        formula: str,
        font: Optional[Font] = None,
        fill: Optional[PatternFill] = None,
        sheet: Optional[str] = None,
    ) -> None:
        """Add a formula-based conditional format rule."""
        ws = self._ws(sheet)
        rule = FormulaRule(formula=[formula], font=font, fill=fill)
        ws.conditional_formatting.add(range_str, rule)

    def add_color_scale(
        self,
        range_str: str,
        start_color: str = "F8696B",
        mid_color: Optional[str] = "FFEB84",
        end_color: str = "63BE7B",
        sheet: Optional[str] = None,
    ) -> None:
        """Add a 2 or 3 color scale conditional format."""
        ws = self._ws(sheet)
        if mid_color:
            rule = ColorScaleRule(
                start_type="min", start_color=start_color,
                mid_type="percentile", mid_value=50, mid_color=mid_color,
                end_type="max", end_color=end_color,
            )
        else:
            rule = ColorScaleRule(
                start_type="min", start_color=start_color,
                end_type="max", end_color=end_color,
            )
        ws.conditional_formatting.add(range_str, rule)

    # ── Charts (NOT histogram — that's cx:chart, use SystemEventsLayer) ──

    def add_bar_chart(
        self,
        sheet: Optional[str] = None,
        title: str = "",
        data_range: str = "B1:B10",
        cats_range: Optional[str] = None,
        anchor: str = "E2",
        width: float = 15,
        height: float = 10,
        style: int = 10,
        bar_dir: str = "col",
    ) -> None:
        """Add a bar/column chart."""
        ws = self._ws(sheet)
        chart = BarChart()
        chart.type = bar_dir  # "col" for column, "bar" for horizontal
        chart.title = title
        chart.style = style
        chart.width = width
        chart.height = height

        data_ref = Reference(ws, range_string=self._qualify_range(ws, data_range))
        chart.add_data(data_ref, titles_from_data=True)

        if cats_range:
            cats_ref = Reference(ws, range_string=self._qualify_range(ws, cats_range))
            chart.set_categories(cats_ref)

        ws.add_chart(chart, anchor)
        logger.info("Added bar chart '%s' at %s", title, anchor)

    def add_line_chart(
        self,
        sheet: Optional[str] = None,
        title: str = "",
        data_range: str = "B1:B10",
        cats_range: Optional[str] = None,
        anchor: str = "E2",
        width: float = 15,
        height: float = 10,
        style: int = 10,
    ) -> None:
        """Add a line chart."""
        ws = self._ws(sheet)
        chart = LineChart()
        chart.title = title
        chart.style = style
        chart.width = width
        chart.height = height

        data_ref = Reference(ws, range_string=self._qualify_range(ws, data_range))
        chart.add_data(data_ref, titles_from_data=True)

        if cats_range:
            cats_ref = Reference(ws, range_string=self._qualify_range(ws, cats_range))
            chart.set_categories(cats_ref)

        ws.add_chart(chart, anchor)
        logger.info("Added line chart '%s' at %s", title, anchor)

    def add_pie_chart(
        self,
        sheet: Optional[str] = None,
        title: str = "",
        data_range: str = "B1:B10",
        cats_range: Optional[str] = None,
        anchor: str = "E2",
        width: float = 15,
        height: float = 10,
    ) -> None:
        """Add a pie chart."""
        ws = self._ws(sheet)
        chart = PieChart()
        chart.title = title
        chart.width = width
        chart.height = height

        data_ref = Reference(ws, range_string=self._qualify_range(ws, data_range))
        chart.add_data(data_ref, titles_from_data=True)

        if cats_range:
            cats_ref = Reference(ws, range_string=self._qualify_range(ws, cats_range))
            chart.set_categories(cats_ref)

        ws.add_chart(chart, anchor)
        logger.info("Added pie chart '%s' at %s", title, anchor)

    def add_scatter_chart(
        self,
        sheet: Optional[str] = None,
        title: str = "",
        x_range: str = "A1:A10",
        y_range: str = "B1:B10",
        anchor: str = "E2",
        width: float = 15,
        height: float = 10,
        style: int = 13,
        scatter_style: str = "line",
    ) -> None:
        """
        Add a scatter (XY) chart.

        scatter_style: 'line', 'lineMarker', 'marker', 'smooth', 'smoothMarker'
        """
        ws = self._ws(sheet)
        chart = ScatterChart()
        chart.title = title
        chart.style = style
        chart.width = width
        chart.height = height
        chart.x_axis.title = ""
        chart.y_axis.title = ""

        x_ref = Reference(ws, range_string=self._qualify_range(ws, x_range))
        y_ref = Reference(ws, range_string=self._qualify_range(ws, y_range))
        series = Series(y_ref, x_ref, title_from_data=True)
        series.graphicalProperties.line.noFill = scatter_style == "marker"
        chart.series.append(series)

        ws.add_chart(chart, anchor)
        logger.info("Added scatter chart '%s' (style=%s) at %s", title, scatter_style, anchor)

    def add_area_chart(
        self,
        sheet: Optional[str] = None,
        title: str = "",
        data_range: str = "B1:B10",
        cats_range: Optional[str] = None,
        anchor: str = "E2",
        width: float = 15,
        height: float = 10,
        style: int = 10,
        grouping: str = "standard",
    ) -> None:
        """
        Add an area chart.

        grouping: 'standard', 'stacked', 'percentStacked'
        """
        ws = self._ws(sheet)
        chart = AreaChart()
        chart.title = title
        chart.style = style
        chart.width = width
        chart.height = height
        chart.grouping = grouping

        data_ref = Reference(ws, range_string=self._qualify_range(ws, data_range))
        chart.add_data(data_ref, titles_from_data=True)

        if cats_range:
            cats_ref = Reference(ws, range_string=self._qualify_range(ws, cats_range))
            chart.set_categories(cats_ref)

        ws.add_chart(chart, anchor)
        logger.info("Added area chart '%s' (grouping=%s) at %s", title, grouping, anchor)

    def add_combo_chart(
        self,
        sheet: Optional[str] = None,
        title: str = "",
        bar_data_range: str = "B1:B10",
        line_data_range: str = "C1:C10",
        cats_range: Optional[str] = None,
        anchor: str = "E2",
        width: float = 15,
        height: float = 10,
        secondary_axis: bool = True,
    ) -> None:
        """
        Add a combo chart (bar + line, optionally with secondary axis).

        openpyxl supports this by appending a LineChart onto a BarChart.
        """
        ws = self._ws(sheet)

        bar_chart = BarChart()
        bar_chart.type = "col"
        bar_chart.title = title
        bar_chart.style = 10
        bar_chart.width = width
        bar_chart.height = height

        bar_ref = Reference(ws, range_string=self._qualify_range(ws, bar_data_range))
        bar_chart.add_data(bar_ref, titles_from_data=True)

        if cats_range:
            cats_ref = Reference(ws, range_string=self._qualify_range(ws, cats_range))
            bar_chart.set_categories(cats_ref)

        # Create the line series on secondary axis
        line_chart = LineChart()
        line_ref = Reference(ws, range_string=self._qualify_range(ws, line_data_range))
        line_chart.add_data(line_ref, titles_from_data=True)

        if secondary_axis:
            line_chart.y_axis.axId = 200
            line_chart.y_axis.crosses = "max"
            bar_chart.y_axis.crosses = "min"

        # Merge line chart into bar chart
        bar_chart += line_chart

        ws.add_chart(bar_chart, anchor)
        logger.info(
            "Added combo chart '%s' (secondary_axis=%s) at %s",
            title, secondary_axis, anchor,
        )

    # ── Named Ranges ──

    def create_named_range(
        self,
        name: str,
        sheet: str,
        range_str: str,
    ) -> None:
        """Create a workbook-scoped named range."""
        if not self.wb:
            raise RuntimeError("No workbook is open")
        defn = DefinedName(name, attr_text=f"'{sheet}'!{range_str}")
        self.wb.defined_names.add(defn)
        logger.info("Created named range '%s' → %s!%s", name, sheet, range_str)

    # ── Data Validation ──

    def add_data_validation(
        self,
        range_str: str,
        validation_type: str = "list",
        formula1: Optional[str] = None,
        formula2: Optional[str] = None,
        allow_blank: bool = True,
        prompt_title: str = "",
        prompt_message: str = "",
        error_title: str = "",
        error_message: str = "",
        sheet: Optional[str] = None,
    ) -> None:
        """Add data validation to a range."""
        ws = self._ws(sheet)
        dv = DataValidation(
            type=validation_type,
            formula1=formula1,
            formula2=formula2,
            allow_blank=allow_blank,
        )
        dv.prompt = prompt_message
        dv.promptTitle = prompt_title
        dv.error = error_message
        dv.errorTitle = error_title
        dv.add(range_str)
        ws.add_data_validation(dv)
        logger.info("Added %s validation to %s", validation_type, range_str)

    # ── Freeze Panes ──

    def freeze_panes(self, cell: str, sheet: Optional[str] = None) -> None:
        """Freeze panes at the given cell (e.g., 'A2' freezes row 1)."""
        ws = self._ws(sheet)
        ws.freeze_panes = cell
        logger.info("Freeze panes at %s", cell)

    # ── AutoFilter ──

    def set_autofilter(self, range_str: str, sheet: Optional[str] = None) -> None:
        """Enable autofilter on a range."""
        ws = self._ws(sheet)
        ws.auto_filter.ref = range_str
        logger.info("Set autofilter on %s", range_str)

    # ── Merge Cells ──

    def merge_cells(self, range_str: str, sheet: Optional[str] = None) -> None:
        """Merge cells in a range."""
        ws = self._ws(sheet)
        ws.merge_cells(range_str)

    def unmerge_cells(self, range_str: str, sheet: Optional[str] = None) -> None:
        """Unmerge cells in a range."""
        ws = self._ws(sheet)
        ws.unmerge_cells(range_str)

    # ── Print Settings ──

    def set_print_area(self, range_str: str, sheet: Optional[str] = None) -> None:
        """Set the print area."""
        ws = self._ws(sheet)
        ws.print_area = range_str

    def set_print_title_rows(self, rows: str, sheet: Optional[str] = None) -> None:
        """Set rows to repeat at top when printing (e.g., '1:1')."""
        ws = self._ws(sheet)
        ws.print_title_rows = rows

    def set_page_orientation(
        self, landscape: bool = True, sheet: Optional[str] = None
    ) -> None:
        """Set page orientation."""
        ws = self._ws(sheet)
        ws.page_setup.orientation = "landscape" if landscape else "portrait"

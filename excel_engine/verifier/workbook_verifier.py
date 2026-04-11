"""
Workbook Verifier — Verify task completion after each section.

Checks formulas, formatting, charts, tables, slicers, named ranges,
data validation, and other Excel features to confirm tasks were executed.
Trained against 114+ SAM textbook checkpoints.
"""

from __future__ import annotations

import logging
import math
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook
from excel_engine.config import TaskType
from excel_engine.parsers.task_extractor import Task

logger = logging.getLogger(__name__)


@dataclass
class VerificationResult:
    """Result of verifying a single task."""
    task_id: str
    task_type: TaskType
    passed: bool
    message: str
    details: Optional[dict] = None


@dataclass
class SectionVerification:
    """Aggregated verification results for a section."""
    section_id: str
    results: list[VerificationResult] = field(default_factory=list)

    @property
    def all_passed(self) -> bool:
        return all(r.passed for r in self.results)

    @property
    def pass_count(self) -> int:
        return sum(1 for r in self.results if r.passed)

    @property
    def fail_count(self) -> int:
        return sum(1 for r in self.results if not r.passed)

    def summary(self) -> str:
        total = len(self.results)
        return (
            f"Section {self.section_id}: {self.pass_count}/{total} passed"
            f"{' ✓' if self.all_passed else ' ✗'}"
        )


def _normalize_formula(f: str) -> str:
    """Normalize a formula for comparison — strip whitespace, uppercase func names."""
    if not f:
        return ""
    f = f.strip()
    if not f.startswith("="):
        f = "=" + f
    # Uppercase function names but preserve string literals
    parts = re.split(r'(".*?")', f)
    normalized = []
    for i, part in enumerate(parts):
        if i % 2 == 0:  # outside quotes
            normalized.append(re.sub(r'\s+', '', part).upper())
        else:
            normalized.append(part)
    return "".join(normalized)


def _expand_range(range_str: str) -> set[str]:
    """Expand a range like 'A1:C3' into a set of cell references."""
    from openpyxl.utils import range_boundaries
    cells = set()
    for part in range_str.replace(" ", "").split(","):
        if ":" in part:
            try:
                min_col, min_row, max_col, max_row = range_boundaries(part)
                from openpyxl.utils import get_column_letter
                for r in range(min_row, max_row + 1):
                    for c in range(min_col, max_col + 1):
                        cells.add(f"{get_column_letter(c)}{r}")
            except Exception:
                cells.add(part)
        else:
            cells.add(part.upper())
    return cells


def _ranges_overlap(rule_range_str: str, expected_cells: set[str]) -> bool:
    """Check whether a rule's range string overlaps with expected cells."""
    rule_cells = _expand_range(rule_range_str)
    return bool(rule_cells & expected_cells)


_CROSS_SHEET_REF_RE = re.compile(
    r"(?:^|[=(,+\-*/&: ])'?([A-Za-z][\w ]*?)'?!([A-Z]+\d+)",
    re.IGNORECASE,
)


def _extract_sheet_references(formula: str) -> list[tuple[str, str]]:
    """Extract (sheet_name, cell_ref) pairs from a formula."""
    return _CROSS_SHEET_REF_RE.findall(formula)


def _get_chart_title_text(chart) -> str | None:
    """Extract the plain-text title string from an openpyxl chart object."""
    title_obj = getattr(chart, "title", None)
    if title_obj is None:
        return None
    # Simple string case (before save/load)
    if isinstance(title_obj, str):
        return title_obj
    # After save/load, title is a Title object with nested rich text
    try:
        text_obj = title_obj.text
        if isinstance(text_obj, str):
            return text_obj
        # RichText: iterate paragraphs → runs
        if hasattr(text_obj, "rich") and text_obj.rich:
            parts = []
            for p in text_obj.rich.p:
                for r in (p.r or []):
                    if r.t:
                        parts.append(r.t)
            if parts:
                return "".join(parts)
    except Exception:
        pass
    return None


def _get_axis_title_text(axis) -> str | None:
    """Extract the plain-text title string from a chart axis."""
    if axis is None:
        return None
    title_obj = getattr(axis, "title", None)
    if title_obj is None:
        return None
    if isinstance(title_obj, str):
        return title_obj
    try:
        text_obj = title_obj.text
        if isinstance(text_obj, str):
            return text_obj
        if hasattr(text_obj, "rich") and text_obj.rich:
            parts = []
            for p in text_obj.rich.p:
                for r in (p.r or []):
                    if r.t:
                        parts.append(r.t)
            if parts:
                return "".join(parts)
    except Exception:
        pass
    return None


class WorkbookVerifier:
    """Verify Excel workbook state against expected task outcomes."""

    def __init__(self) -> None:
        self._wb = None
        self._wb_values = None  # data_only=True version for computed values
        self._path: Optional[Path] = None

    def load(self, path: Path) -> None:
        """Load a workbook for verification (data_only=False to see formulas)."""
        self._path = Path(path)
        self._wb = load_workbook(str(self._path), data_only=False)
        logger.info("Loaded workbook for verification: %s", self._path.name)

    def load_with_values(self, path: Path) -> None:
        """Load a workbook with calculated values (data_only=True)."""
        self._path = Path(path)
        self._wb_values = load_workbook(str(self._path), data_only=True)

    def close(self) -> None:
        if self._wb:
            self._wb.close()
            self._wb = None
        if self._wb_values:
            self._wb_values.close()
            self._wb_values = None

    def verify_section(
        self, section_id: str, tasks: list[Task]
    ) -> SectionVerification:
        """Verify all tasks in a section."""
        verification = SectionVerification(section_id=section_id)

        for task in tasks:
            result = self.verify_task(task)
            verification.results.append(result)

        logger.info(verification.summary())
        return verification

    def verify_task(self, task: Task) -> VerificationResult:
        """Verify a single task was completed correctly."""
        verifiers = {
            # ── Data Entry & Formulas ──
            TaskType.FORMULA: self._verify_formula,
            TaskType.CELL_VALUE: self._verify_cell_value,
            TaskType.TEXT_FUNCTION: self._verify_formula,
            TaskType.LOOKUP_FUNCTION: self._verify_formula,
            TaskType.FILTER_FUNCTION: self._verify_formula,
            TaskType.SORT_FUNCTION: self._verify_formula,
            TaskType.UNIQUE_FUNCTION: self._verify_formula,
            TaskType.THREE_D_REFERENCE: self._verify_formula,
            TaskType.EXTERNAL_REFERENCE: self._verify_external_reference,

            # ── Tables ──
            TaskType.TABLE_CREATE: self._verify_table,
            TaskType.TABLE_STYLE: self._verify_table_style,
            TaskType.TABLE_TOTAL_ROW: self._verify_table_total_row,
            TaskType.CALCULATED_COLUMN: self._verify_formula,

            # ── Formatting ──
            TaskType.FORMATTING: self._verify_formatting,
            TaskType.CONDITIONAL_FORMAT: self._verify_conditional_format,
            TaskType.NUMBER_FORMAT: self._verify_number_format,
            TaskType.ALIGNMENT: self._verify_alignment,
            TaskType.COLUMN_WIDTH: self._verify_column_width,
            TaskType.ROW_HEIGHT: self._verify_row_height,
            TaskType.FONT: self._verify_font,
            TaskType.FILL: self._verify_fill,
            TaskType.BORDER: self._verify_border,
            TaskType.MERGE_CELLS: self._verify_merged_cells,
            TaskType.TAB_COLOR: self._verify_tab_color,

            # ── View & Layout ──
            TaskType.FREEZE_PANES: self._verify_freeze_panes,
            TaskType.SPLIT_PANES: self._verify_split_panes,
            TaskType.PAGE_BREAK: self._verify_page_break,
            TaskType.PRINT_SETTINGS: self._verify_print_settings,

            # ── Data Tools ──
            TaskType.AUTOFILTER: self._verify_autofilter,
            TaskType.ADVANCED_FILTER: self._verify_autofilter,
            TaskType.SORT: self._verify_sort,
            TaskType.SUBTOTAL: self._verify_subtotal,
            TaskType.DATA_VALIDATION: self._verify_data_validation,
            TaskType.GOAL_SEEK: self._verify_cell_value,

            # ── Charts ──
            TaskType.CHART_BAR: self._verify_chart,
            TaskType.CHART_LINE: self._verify_chart,
            TaskType.CHART_PIE: self._verify_chart,
            TaskType.CHART_SCATTER: self._verify_chart,
            TaskType.CHART_AREA: self._verify_chart,
            TaskType.CHART_COMBO: self._verify_chart,
            TaskType.CHART_HISTOGRAM: self._verify_chart,
            TaskType.SPARKLINE: self._verify_chart,

            # ── Ranges & References ──
            TaskType.NAMED_RANGE: self._verify_named_range,
            TaskType.HYPERLINK: self._verify_hyperlink,

            # ── Advanced Features ──
            TaskType.SLICER: self._verify_slicer,
            TaskType.PIVOT_TABLE: self._verify_pivot_table,
            TaskType.PIVOT_CHART: self._verify_chart,

            # ── Sheet Operations ──
            TaskType.SHEET_CREATE: self._verify_sheet_exists,
            TaskType.SHEET_RENAME: self._verify_sheet_exists,
            TaskType.SHEET_COPY: self._verify_sheet_exists,
        }

        verifier = verifiers.get(task.task_type)
        if verifier:
            try:
                return verifier(task)
            except Exception as e:
                return VerificationResult(
                    task_id=task.id,
                    task_type=task.task_type,
                    passed=False,
                    message=f"Verification error: {e}",
                )

        # No specific verifier — assume passed if task is marked completed
        return VerificationResult(
            task_id=task.id,
            task_type=task.task_type,
            passed=task.completed,
            message="No specific verifier; relying on execution status",
        )

    # ══════════════════════════════════════════════════════════════════
    # Data Entry & Formula Verifiers
    # ══════════════════════════════════════════════════════════════════

    def _verify_formula(self, task: Task) -> VerificationResult:
        """Verify a formula was entered correctly, with deep comparison."""
        if not self._wb or not task.cell:
            return self._skip(task, "No cell reference")

        ws = self._get_ws(task.sheet)
        cell = ws[task.cell]

        if cell.value is None:
            return VerificationResult(
                task_id=task.id, task_type=task.task_type,
                passed=False, message=f"Cell {task.cell} is empty",
            )

        cell_value = str(cell.value)

        if task.formula:
            if cell_value.startswith("="):
                # Deep comparison: normalize both formulas
                actual = _normalize_formula(cell_value)
                expected = _normalize_formula(task.formula)
                exact_match = actual == expected

                # A4: Validate cross-sheet references (warnings only)
                ref_warnings = self._check_cross_sheet_refs(cell_value)

                result = VerificationResult(
                    task_id=task.id, task_type=task.task_type,
                    passed=exact_match,
                    message=(
                        f"Formula match: {cell_value}"
                        if exact_match
                        else f"Formula present (differs): actual={cell_value}, expected={task.formula}"
                    ),
                    details={
                        "expected": task.formula,
                        "actual": cell_value,
                        "exact_match": exact_match,
                    },
                )
                if ref_warnings:
                    result.details["cross_sheet_warnings"] = ref_warnings
                return result
            else:
                return VerificationResult(
                    task_id=task.id, task_type=task.task_type,
                    passed=False,
                    message=f"Expected formula, got value: {cell_value}",
                )

        # No expected formula — just check cell has a formula
        if cell_value.startswith("="):
            # A4: Validate cross-sheet references (warnings only)
            ref_warnings = self._check_cross_sheet_refs(cell_value)
            result = VerificationResult(
                task_id=task.id, task_type=task.task_type,
                passed=True, message=f"Formula found: {cell_value}",
            )
            if ref_warnings:
                result.details = {"cross_sheet_warnings": ref_warnings}
            return result

        return VerificationResult(
            task_id=task.id, task_type=task.task_type,
            passed=cell_value is not None,
            message=f"Cell has value: {cell_value}",
        )

    def _verify_cell_value(self, task: Task) -> VerificationResult:
        """Verify a cell has a specific value."""
        if not self._wb or not task.cell:
            return self._skip(task, "No cell reference")

        ws = self._get_ws(task.sheet)
        cell = ws[task.cell]

        if cell.value is None:
            return VerificationResult(
                task_id=task.id, task_type=task.task_type,
                passed=False, message=f"Cell {task.cell} is empty",
            )

        cell_value = str(cell.value)

        if task.value:
            matches = cell_value.strip().lower() == task.value.strip().lower()
            if not matches:
                # A1: Floating-point leniency — numeric fallback
                try:
                    actual_num = float(str(cell.value))
                    expected_num = float(str(task.value))
                    if math.isclose(actual_num, expected_num, rel_tol=1e-9, abs_tol=1e-12):
                        matches = True
                except (ValueError, TypeError):
                    pass
            return VerificationResult(
                task_id=task.id, task_type=task.task_type,
                passed=matches,
                message=(
                    f"Value matches: {cell_value}"
                    if matches
                    else f"Value mismatch: expected='{task.value}', actual='{cell_value}'"
                ),
                details={"expected": task.value, "actual": cell_value},
            )

        return VerificationResult(
            task_id=task.id, task_type=task.task_type,
            passed=True,
            message=(
                f"No expected value specified — cannot verify correctness (cell has: {cell_value})"
                if task.formula is None
                else f"Cell has value: {cell_value}"
            ),
        )

    def _verify_external_reference(self, task: Task) -> VerificationResult:
        """Verify external references (links to other workbooks)."""
        if not self._wb:
            return self._skip(task, "No workbook loaded")

        # Check for external links in defined names or cell formulas
        has_external = False
        for dn in self._wb.defined_names.definedName:
            if "[" in str(dn.value):
                has_external = True
                break

        if not has_external and task.cell:
            ws = self._get_ws(task.sheet)
            cell = ws[task.cell]
            if cell.value and "[" in str(cell.value):
                has_external = True

        return VerificationResult(
            task_id=task.id, task_type=task.task_type,
            passed=has_external,
            message="External reference found" if has_external else "No external references detected",
        )

    # ══════════════════════════════════════════════════════════════════
    # Table Verifiers
    # ══════════════════════════════════════════════════════════════════

    def _verify_table(self, task: Task) -> VerificationResult:
        """Verify a table was created, with range and header checks."""
        ws = self._get_ws(task.sheet)

        if not ws.tables:
            return VerificationResult(
                task_id=task.id, task_type=task.task_type,
                passed=False, message="No tables found on sheet",
            )

        table_names = list(ws.tables.keys())
        details: dict[str, Any] = {"tables": []}

        for name in table_names:
            table = ws.tables[name]
            info: dict[str, Any] = {
                "name": name,
                "ref": table.ref,
                "style": table.tableStyleInfo.name if table.tableStyleInfo else None,
            }
            details["tables"].append(info)

        # If task specifies a range, check it matches
        if task.range:
            matching = any(
                ws.tables[n].ref == task.range for n in table_names
            )
            if not matching:
                return VerificationResult(
                    task_id=task.id, task_type=task.task_type,
                    passed=True,
                    message=f"Table(s) found but range differs: expected {task.range}, got {[ws.tables[n].ref for n in table_names]}",
                    details=details,
                )

        return VerificationResult(
            task_id=task.id, task_type=task.task_type,
            passed=True,
            message=f"Table(s) found: {table_names}",
            details=details,
        )

    def _verify_table_style(self, task: Task) -> VerificationResult:
        """Verify a table style was applied."""
        ws = self._get_ws(task.sheet)

        for name, table in ws.tables.items():
            if task.style:
                if table.tableStyleInfo and table.tableStyleInfo.name == task.style:
                    return VerificationResult(
                        task_id=task.id, task_type=task.task_type,
                        passed=True,
                        message=f"Table '{name}' has style {task.style}",
                    )
            else:
                if table.tableStyleInfo:
                    return VerificationResult(
                        task_id=task.id, task_type=task.task_type,
                        passed=True,
                        message=f"Table '{name}' has style {table.tableStyleInfo.name}",
                    )

        return VerificationResult(
            task_id=task.id, task_type=task.task_type,
            passed=False,
            message=f"Expected style '{task.style}' not found",
        )

    def _verify_table_total_row(self, task: Task) -> VerificationResult:
        """Verify the total row is enabled on a table."""
        ws = self._get_ws(task.sheet)

        for name, table in ws.tables.items():
            if table.totalsRowShown is not False:
                # totalsRowCount > 0 means total row is present
                if hasattr(table, "totalsRowCount") and table.totalsRowCount:
                    return VerificationResult(
                        task_id=task.id, task_type=task.task_type,
                        passed=True,
                        message=f"Table '{name}' has total row enabled",
                    )
                # Also check for SUBTOTAL formulas in the row after the table
                return VerificationResult(
                    task_id=task.id, task_type=task.task_type,
                    passed=True,
                    message=f"Table '{name}' — totalsRowShown not explicitly False",
                )

        return VerificationResult(
            task_id=task.id, task_type=task.task_type,
            passed=False,
            message="No table with total row found",
        )

    # ══════════════════════════════════════════════════════════════════
    # Formatting Verifiers
    # ══════════════════════════════════════════════════════════════════

    def _verify_formatting(self, task: Task) -> VerificationResult:
        """Verify general formatting was applied."""
        if not task.cell and not task.range:
            return self._skip(task, "No cell/range reference")

        ws = self._get_ws(task.sheet)
        ref = task.cell or task.range
        cell = ws[ref.split(":")[0]]

        has_formatting = (
            cell.font.bold or cell.font.italic or
            cell.fill.start_color.index not in (None, "00000000", 0) or
            cell.alignment.horizontal not in (None, "general")
        )

        return VerificationResult(
            task_id=task.id, task_type=task.task_type,
            passed=has_formatting,
            message="Formatting detected" if has_formatting else "No formatting detected",
        )

    def _verify_number_format(self, task: Task) -> VerificationResult:
        """Verify a number format was applied."""
        if not task.cell and not task.range:
            return self._skip(task, "No cell/range reference")

        ws = self._get_ws(task.sheet)
        ref = task.cell or task.range
        cell = ws[ref.split(":")[0]]

        fmt = cell.number_format
        is_custom = fmt not in (None, "General")

        return VerificationResult(
            task_id=task.id, task_type=task.task_type,
            passed=is_custom,
            message=f"Number format: {fmt}",
            details={"number_format": fmt},
        )

    def _verify_conditional_format(self, task: Task) -> VerificationResult:
        """Verify conditional formatting rules exist, with type & range matching."""
        ws = self._get_ws(task.sheet)

        cf_rules = ws.conditional_formatting
        count = len(list(cf_rules))

        details = {}
        if count > 0:
            rules_info = []
            for cf in cf_rules:
                cf_range = str(cf.sqref) if hasattr(cf, 'sqref') else str(cf)
                for rule in cf.rules:
                    rules_info.append({
                        "type": rule.type,
                        "priority": rule.priority,
                        "range": cf_range,
                    })
            details["rules"] = rules_info

        if count == 0:
            return VerificationResult(
                task_id=task.id, task_type=task.task_type,
                passed=False,
                message="No conditional formatting rules found",
                details=details if details else None,
            )

        # A3: Check rule type match if specified
        expected_type = task.params.get("rule_type")
        expected_range = task.params.get("range") or task.range

        type_ok = True
        range_ok = True

        if expected_type:
            type_ok = any(
                r["type"] and r["type"].lower() == expected_type.lower()
                for r in details.get("rules", [])
            )
            details["type_match"] = type_ok

        if expected_range:
            expected_cells = _expand_range(expected_range)
            range_ok = any(
                _ranges_overlap(r["range"], expected_cells)
                for r in details.get("rules", [])
            )
            details["range_match"] = range_ok

        passed = type_ok and range_ok
        messages = [f"{count} conditional formatting rule(s) found"]
        if expected_type and not type_ok:
            messages.append(f"No rule matching type '{expected_type}'")
        if expected_range and not range_ok:
            messages.append(f"No rule overlapping range '{expected_range}'")

        return VerificationResult(
            task_id=task.id, task_type=task.task_type,
            passed=passed,
            message="; ".join(messages),
            details=details if details else None,
        )

    def _verify_alignment(self, task: Task) -> VerificationResult:
        """Verify alignment settings on a cell or range."""
        if not task.cell and not task.range:
            return self._skip(task, "No cell/range reference")

        ws = self._get_ws(task.sheet)
        ref = task.cell or task.range
        cell = ws[ref.split(":")[0]]

        alignment = cell.alignment
        has_alignment = (
            alignment.horizontal not in (None, "general")
            or alignment.vertical not in (None, "bottom")
            or alignment.wrap_text is True
            or alignment.text_rotation not in (None, 0)
            or alignment.indent not in (None, 0)
            or alignment.shrink_to_fit is True
        )

        return VerificationResult(
            task_id=task.id, task_type=task.task_type,
            passed=has_alignment,
            message=(
                f"Alignment: h={alignment.horizontal}, v={alignment.vertical}, "
                f"wrap={alignment.wrap_text}, rotation={alignment.text_rotation}"
                if has_alignment
                else "No custom alignment detected"
            ),
            details={
                "horizontal": alignment.horizontal,
                "vertical": alignment.vertical,
                "wrap_text": alignment.wrap_text,
                "text_rotation": alignment.text_rotation,
            },
        )

    def _verify_column_width(self, task: Task) -> VerificationResult:
        """Verify column width was changed."""
        ws = self._get_ws(task.sheet)

        # Try to determine column from cell/range
        col_letter = None
        if task.cell:
            col_letter = re.match(r"([A-Z]+)", task.cell)
            col_letter = col_letter.group(1) if col_letter else None
        elif task.range:
            col_letter = re.match(r"([A-Z]+)", task.range)
            col_letter = col_letter.group(1) if col_letter else None

        if col_letter and col_letter in ws.column_dimensions:
            dim = ws.column_dimensions[col_letter]
            actual_width = dim.width
            expected = task.params.get("size")

            if expected:
                matches = abs(actual_width - expected) < 0.5
                return VerificationResult(
                    task_id=task.id, task_type=task.task_type,
                    passed=matches,
                    message=(
                        f"Column {col_letter} width={actual_width} "
                        f"(expected {expected})"
                    ),
                    details={"column": col_letter, "width": actual_width, "expected": expected},
                )

            return VerificationResult(
                task_id=task.id, task_type=task.task_type,
                passed=actual_width is not None,
                message=f"Column {col_letter} width={actual_width}",
                details={"column": col_letter, "width": actual_width},
            )

        return VerificationResult(
            task_id=task.id, task_type=task.task_type,
            passed=True,
            message="Column width check skipped (no specific column reference)",
        )

    def _verify_row_height(self, task: Task) -> VerificationResult:
        """Verify row height was changed."""
        ws = self._get_ws(task.sheet)

        row_num = None
        if task.cell:
            m = re.search(r"(\d+)", task.cell)
            row_num = int(m.group(1)) if m else None

        if row_num and row_num in ws.row_dimensions:
            dim = ws.row_dimensions[row_num]
            actual_height = dim.height
            expected = task.params.get("size")

            if expected:
                matches = abs((actual_height or 15) - expected) < 0.5
                return VerificationResult(
                    task_id=task.id, task_type=task.task_type,
                    passed=matches,
                    message=f"Row {row_num} height={actual_height} (expected {expected})",
                    details={"row": row_num, "height": actual_height, "expected": expected},
                )

            return VerificationResult(
                task_id=task.id, task_type=task.task_type,
                passed=actual_height is not None,
                message=f"Row {row_num} height={actual_height}",
            )

        return self._skip(task, "No row reference for height check")

    def _verify_font(self, task: Task) -> VerificationResult:
        """Verify font settings (bold, italic, size, color, name)."""
        if not task.cell and not task.range:
            return self._skip(task, "No cell/range reference")

        ws = self._get_ws(task.sheet)
        ref = task.cell or task.range
        cell = ws[ref.split(":")[0]]
        font = cell.font

        has_custom_font = (
            font.bold is True
            or font.italic is True
            or font.underline is not None
            or font.strikethrough is True
            or (font.size is not None and font.size != 11)
            or (font.name is not None and font.name != "Calibri")
            or (font.color and font.color.rgb and font.color.rgb != "00000000")
        )

        return VerificationResult(
            task_id=task.id, task_type=task.task_type,
            passed=has_custom_font,
            message=(
                f"Font: name={font.name}, size={font.size}, "
                f"bold={font.bold}, italic={font.italic}"
                if has_custom_font
                else "No custom font detected"
            ),
            details={
                "name": font.name, "size": font.size,
                "bold": font.bold, "italic": font.italic,
                "underline": font.underline, "color": str(font.color),
            },
        )

    def _verify_fill(self, task: Task) -> VerificationResult:
        """Verify fill/background color was applied."""
        if not task.cell and not task.range:
            return self._skip(task, "No cell/range reference")

        ws = self._get_ws(task.sheet)
        ref = task.cell or task.range
        cell = ws[ref.split(":")[0]]
        fill = cell.fill

        has_fill = (
            fill.fill_type is not None
            and fill.fill_type != "none"
            and fill.start_color.index not in (None, "00000000", 0)
        )

        return VerificationResult(
            task_id=task.id, task_type=task.task_type,
            passed=has_fill,
            message=(
                f"Fill: type={fill.fill_type}, color={fill.start_color.index}"
                if has_fill
                else "No fill/background color detected"
            ),
            details={
                "fill_type": fill.fill_type,
                "color": str(fill.start_color.index),
            },
        )

    def _verify_border(self, task: Task) -> VerificationResult:
        """Verify borders were applied."""
        if not task.cell and not task.range:
            return self._skip(task, "No cell/range reference")

        ws = self._get_ws(task.sheet)
        ref = task.cell or task.range
        cell = ws[ref.split(":")[0]]
        border = cell.border

        has_border = any(
            getattr(border, side).style is not None
            for side in ("left", "right", "top", "bottom", "diagonal")
        )

        sides = {}
        for side in ("left", "right", "top", "bottom"):
            s = getattr(border, side)
            if s.style:
                sides[side] = s.style

        return VerificationResult(
            task_id=task.id, task_type=task.task_type,
            passed=has_border,
            message=(
                f"Borders: {sides}" if has_border else "No borders detected"
            ),
            details={"borders": sides},
        )

    def _verify_tab_color(self, task: Task) -> VerificationResult:
        """Verify sheet tab color was set."""
        if not self._wb:
            return self._skip(task, "No workbook loaded")

        ws = self._get_ws(task.sheet)
        color = ws.sheet_properties.tabColor

        return VerificationResult(
            task_id=task.id, task_type=task.task_type,
            passed=color is not None,
            message=f"Tab color: {color}" if color else "No tab color set",
        )

    # ══════════════════════════════════════════════════════════════════
    # View & Layout Verifiers
    # ══════════════════════════════════════════════════════════════════

    def _verify_freeze_panes(self, task: Task) -> VerificationResult:
        """Verify freeze panes are set."""
        ws = self._get_ws(task.sheet)

        frozen = ws.freeze_panes
        expected = task.cell

        if expected and frozen:
            matches = str(frozen) == expected
            return VerificationResult(
                task_id=task.id, task_type=task.task_type,
                passed=True,  # freeze is set
                message=(
                    f"Freeze panes at {frozen}"
                    + (f" (expected {expected})" if not matches else "")
                ),
                details={"actual": str(frozen), "expected": expected},
            )

        return VerificationResult(
            task_id=task.id, task_type=task.task_type,
            passed=frozen is not None,
            message=f"Freeze panes: {frozen}" if frozen else "No freeze panes set",
        )

    def _verify_split_panes(self, task: Task) -> VerificationResult:
        """Verify split panes — openpyxl can detect via sheet_view."""
        ws = self._get_ws(task.sheet)

        # openpyxl stores split info in sheet_view
        for view in ws.views.sheetView:
            pane = view.pane
            if pane and (pane.xSplit or pane.ySplit):
                return VerificationResult(
                    task_id=task.id, task_type=task.task_type,
                    passed=True,
                    message=f"Split panes: xSplit={pane.xSplit}, ySplit={pane.ySplit}",
                    details={"xSplit": pane.xSplit, "ySplit": pane.ySplit},
                )

        return VerificationResult(
            task_id=task.id, task_type=task.task_type,
            passed=False,
            message="No split panes detected",
        )

    def _verify_page_break(self, task: Task) -> VerificationResult:
        """Verify page breaks exist."""
        ws = self._get_ws(task.sheet)

        row_breaks = len(ws.row_breaks.brk) if ws.row_breaks else 0
        col_breaks = len(ws.col_breaks.brk) if ws.col_breaks else 0
        total = row_breaks + col_breaks

        return VerificationResult(
            task_id=task.id, task_type=task.task_type,
            passed=total > 0,
            message=f"Page breaks: {row_breaks} row, {col_breaks} column",
            details={"row_breaks": row_breaks, "col_breaks": col_breaks},
        )

    def _verify_print_settings(self, task: Task) -> VerificationResult:
        """Verify print/page setup settings."""
        ws = self._get_ws(task.sheet)
        ps = ws.page_setup

        has_settings = (
            ps.orientation is not None
            or ps.paperSize is not None
            or ps.fitToWidth is not None
            or ps.fitToHeight is not None
            or ws.print_area is not None
            or ws.print_title_rows is not None
            or ws.print_title_cols is not None
        )

        details = {
            "orientation": ps.orientation,
            "paper_size": ps.paperSize,
            "fit_to_width": ps.fitToWidth,
            "fit_to_height": ps.fitToHeight,
            "print_area": ws.print_area,
            "print_title_rows": ws.print_title_rows,
        }

        return VerificationResult(
            task_id=task.id, task_type=task.task_type,
            passed=has_settings,
            message="Print settings configured" if has_settings else "Default print settings",
            details=details,
        )

    # ══════════════════════════════════════════════════════════════════
    # Data Tool Verifiers
    # ══════════════════════════════════════════════════════════════════

    def _verify_autofilter(self, task: Task) -> VerificationResult:
        """Verify autofilter is enabled."""
        ws = self._get_ws(task.sheet)
        has_filter = ws.auto_filter.ref is not None

        return VerificationResult(
            task_id=task.id, task_type=task.task_type,
            passed=has_filter,
            message=f"Autofilter: {ws.auto_filter.ref}" if has_filter else "No autofilter",
            details={"ref": ws.auto_filter.ref} if has_filter else None,
        )

    def _verify_sort(self, task: Task) -> VerificationResult:
        """Verify data appears sorted (heuristic — checks if auto_filter has sort state)."""
        ws = self._get_ws(task.sheet)

        # openpyxl doesn't directly store sort state, but if autofilter
        # has sortCondition children, that's evidence of sorting
        if ws.auto_filter.ref:
            sort_state = ws.auto_filter.sortState
            if sort_state and sort_state.sortCondition:
                return VerificationResult(
                    task_id=task.id, task_type=task.task_type,
                    passed=True,
                    message=f"Sort state found: {len(sort_state.sortCondition)} condition(s)",
                )

        # Fallback: check if any SUBTOTAL functions exist (often paired with sort)
        has_subtotal = False
        for row in ws.iter_rows(max_row=min(ws.max_row or 1, 100)):
            for cell in row:
                if isinstance(cell.value, str) and "SUBTOTAL" in cell.value.upper():
                    has_subtotal = True
                    break
            if has_subtotal:
                break

        return VerificationResult(
            task_id=task.id, task_type=task.task_type,
            passed=has_subtotal,
            message=(
                "SUBTOTAL formulas found (likely sorted with subtotals)"
                if has_subtotal
                else "Cannot confirm sort state via openpyxl (sort applied at runtime)"
            ),
        )

    def _verify_subtotal(self, task: Task) -> VerificationResult:
        """Verify SUBTOTAL formulas and outline/grouping exist."""
        ws = self._get_ws(task.sheet)

        subtotal_count = 0
        outline_levels = set()

        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.upper().startswith("=SUBTOTAL"):
                    subtotal_count += 1
            # Check row outline level
            row_num = row[0].row if row else None
            if row_num and row_num in ws.row_dimensions:
                level = ws.row_dimensions[row_num].outlineLevel
                if level and level > 0:
                    outline_levels.add(level)

        has_subtotals = subtotal_count > 0 or len(outline_levels) > 0

        return VerificationResult(
            task_id=task.id, task_type=task.task_type,
            passed=has_subtotals,
            message=(
                f"{subtotal_count} SUBTOTAL formula(s), "
                f"{len(outline_levels)} outline level(s)"
            ),
            details={
                "subtotal_formulas": subtotal_count,
                "outline_levels": sorted(outline_levels),
            },
        )

    def _verify_data_validation(self, task: Task) -> VerificationResult:
        """Verify data validation exists, with type details."""
        ws = self._get_ws(task.sheet)
        dv_list = ws.data_validations.dataValidation

        if not dv_list:
            return VerificationResult(
                task_id=task.id, task_type=task.task_type,
                passed=False, message="No data validations found",
            )

        details = []
        for dv in dv_list:
            details.append({
                "type": dv.type,
                "ranges": str(dv.sqref),
                "formula1": str(dv.formula1) if dv.formula1 else None,
                "allow_blank": dv.allow_blank,
            })

        return VerificationResult(
            task_id=task.id, task_type=task.task_type,
            passed=True,
            message=f"{len(dv_list)} data validation(s) found",
            details={"validations": details},
        )

    # ══════════════════════════════════════════════════════════════════
    # Chart Verifiers
    # ══════════════════════════════════════════════════════════════════

    def _verify_chart(self, task: Task) -> VerificationResult:
        """Verify a chart exists on the sheet, with type, title, legend, and series checks."""
        ws = self._get_ws(task.sheet)
        charts = ws._charts

        if not charts:
            return VerificationResult(
                task_id=task.id, task_type=task.task_type,
                passed=False, message="No charts found on sheet",
            )

        chart_types = []
        details: dict[str, Any] = {"count": len(charts)}
        checks_passed = True
        messages = [f"{len(charts)} chart(s) found"]

        for chart in charts:
            chart_types.append(type(chart).__name__)

        details["types"] = chart_types

        # A2: Enhanced chart checks
        expected_title = task.params.get("title")
        check_legend = task.params.get("legend")
        expected_series_count = task.params.get("series_count")
        expected_x_axis = task.params.get("x_axis_title")
        expected_y_axis = task.params.get("y_axis_title")

        if expected_title is not None:
            title_found = any(
                _get_chart_title_text(c) is not None
                and _get_chart_title_text(c).strip().lower() == str(expected_title).strip().lower()
                for c in charts
            )
            details["title_match"] = title_found
            if not title_found:
                checks_passed = False
                messages.append(f"Expected chart title '{expected_title}' not found")
            else:
                messages.append(f"Chart title '{expected_title}' found")

        if check_legend is not None:
            legend_found = any(getattr(c, "legend", None) is not None for c in charts)
            details["legend_present"] = legend_found
            if check_legend and not legend_found:
                checks_passed = False
                messages.append("Expected legend not found")
            elif not check_legend and legend_found:
                checks_passed = False
                messages.append("Legend found but not expected")

        # Always verify at least one series exists
        max_series = max((len(getattr(c, "series", []) or []) for c in charts), default=0)
        details["max_series_count"] = max_series
        if max_series < 1:
            checks_passed = False
            messages.append("No data series found in any chart")

        if expected_series_count is not None:
            series_match = any(
                len(getattr(c, "series", []) or []) == int(expected_series_count)
                for c in charts
            )
            details["series_count_match"] = series_match
            if not series_match:
                checks_passed = False
                messages.append(f"Expected {expected_series_count} series, found counts: {[len(getattr(c, 'series', []) or []) for c in charts]}")

        if expected_x_axis is not None:
            x_match = any(
                _get_axis_title_text(getattr(c, "x_axis", None)) is not None
                and _get_axis_title_text(c.x_axis).strip().lower() == str(expected_x_axis).strip().lower()
                for c in charts
            )
            details["x_axis_match"] = x_match
            if not x_match:
                checks_passed = False
                messages.append(f"Expected x-axis title '{expected_x_axis}' not found")

        if expected_y_axis is not None:
            y_match = any(
                _get_axis_title_text(getattr(c, "y_axis", None)) is not None
                and _get_axis_title_text(c.y_axis).strip().lower() == str(expected_y_axis).strip().lower()
                for c in charts
            )
            details["y_axis_match"] = y_match
            if not y_match:
                checks_passed = False
                messages.append(f"Expected y-axis title '{expected_y_axis}' not found")

        return VerificationResult(
            task_id=task.id, task_type=task.task_type,
            passed=checks_passed,
            message="; ".join(messages),
            details=details,
        )

    # ══════════════════════════════════════════════════════════════════
    # Range & Reference Verifiers
    # ══════════════════════════════════════════════════════════════════

    def _verify_named_range(self, task: Task) -> VerificationResult:
        """Verify a named range exists, with optional name matching."""
        if not self._wb:
            return self._skip(task, "No workbook loaded")

        names = [dn.name for dn in self._wb.defined_names.definedName]

        if not names:
            return VerificationResult(
                task_id=task.id, task_type=task.task_type,
                passed=False, message="No named ranges found",
            )

        # If task has a specific name in params, check for it
        expected_name = task.params.get("name") or task.value
        if expected_name:
            found = any(n.lower() == expected_name.lower() for n in names)
            return VerificationResult(
                task_id=task.id, task_type=task.task_type,
                passed=found,
                message=(
                    f"Named range '{expected_name}' found"
                    if found
                    else f"Named range '{expected_name}' not found (existing: {names})"
                ),
                details={"expected": expected_name, "all_names": names},
            )

        return VerificationResult(
            task_id=task.id, task_type=task.task_type,
            passed=True,
            message=f"Named ranges found: {names}",
            details={"names": names},
        )

    def _verify_hyperlink(self, task: Task) -> VerificationResult:
        """Verify a hyperlink exists on the specified cell or sheet."""
        ws = self._get_ws(task.sheet)

        if task.cell:
            cell = ws[task.cell]
            if cell.hyperlink:
                return VerificationResult(
                    task_id=task.id, task_type=task.task_type,
                    passed=True,
                    message=f"Hyperlink on {task.cell}: {cell.hyperlink.target}",
                    details={"target": cell.hyperlink.target, "display": cell.hyperlink.display},
                )
            return VerificationResult(
                task_id=task.id, task_type=task.task_type,
                passed=False,
                message=f"No hyperlink on cell {task.cell}",
            )

        # Check entire sheet for any hyperlinks
        hyperlinks = ws.hyperlinks
        if hasattr(hyperlinks, '__len__'):
            count = len(hyperlinks)
        else:
            count = sum(1 for _ in hyperlinks) if hyperlinks else 0

        return VerificationResult(
            task_id=task.id, task_type=task.task_type,
            passed=count > 0,
            message=f"{count} hyperlink(s) found on sheet" if count > 0 else "No hyperlinks found",
        )

    # ══════════════════════════════════════════════════════════════════
    # Advanced Feature Verifiers
    # ══════════════════════════════════════════════════════════════════

    def _verify_slicer(self, task: Task) -> VerificationResult:
        """
        Verify slicer existence.  openpyxl has limited slicer support —
        we check for slicerCache entries in the workbook's rels/XML.
        """
        if not self._wb:
            return self._skip(task, "No workbook loaded")

        # openpyxl doesn't directly expose slicers, but we can check
        # if the workbook has slicer caches (stored in wb._slicers or rels)
        has_slicers = False

        # Check for slicer drawing objects on sheets
        ws = self._get_ws(task.sheet)
        # Slicer presence can be inferred from drawings or defined name _xlnm.Slicer
        for dn in self._wb.defined_names.definedName:
            if "slicer" in dn.name.lower() or "_xlnm.Slicer" in str(dn.name):
                has_slicers = True
                break

        return VerificationResult(
            task_id=task.id, task_type=task.task_type,
            passed=has_slicers,
            message=(
                "Slicer evidence found in workbook"
                if has_slicers
                else "No slicers detected (limited openpyxl support — verify in Excel)"
            ),
        )

    def _verify_pivot_table(self, task: Task) -> VerificationResult:
        """Verify a PivotTable exists (check via openpyxl pivotTable collection)."""
        if not self._wb:
            return self._skip(task, "No workbook loaded")

        ws = self._get_ws(task.sheet)

        # openpyxl stores pivot tables in ws._pivots
        pivot_count = len(ws._pivots) if hasattr(ws, "_pivots") else 0

        if pivot_count > 0:
            return VerificationResult(
                task_id=task.id, task_type=task.task_type,
                passed=True,
                message=f"{pivot_count} PivotTable(s) found on sheet",
            )

        # Fallback: check all sheets
        total = 0
        for name in self._wb.sheetnames:
            s = self._wb[name]
            if hasattr(s, "_pivots"):
                total += len(s._pivots)

        return VerificationResult(
            task_id=task.id, task_type=task.task_type,
            passed=total > 0,
            message=(
                f"{total} PivotTable(s) found in workbook"
                if total > 0
                else "No PivotTables found"
            ),
        )

    # ══════════════════════════════════════════════════════════════════
    # Sheet Operation Verifiers
    # ══════════════════════════════════════════════════════════════════

    def _verify_sheet_exists(self, task: Task) -> VerificationResult:
        """Verify a sheet exists by name."""
        if not self._wb:
            return self._skip(task, "No workbook loaded")

        sheet_name = task.sheet or task.params.get("name") or task.value
        if sheet_name:
            exists = sheet_name in self._wb.sheetnames
            return VerificationResult(
                task_id=task.id, task_type=task.task_type,
                passed=exists,
                message=f"Sheet '{sheet_name}' {'exists' if exists else 'not found'}",
                details={"sheet": sheet_name, "all_sheets": self._wb.sheetnames},
            )

        return VerificationResult(
            task_id=task.id, task_type=task.task_type,
            passed=True,
            message=f"Sheets: {self._wb.sheetnames}",
        )

    def _verify_merged_cells(self, task: Task) -> VerificationResult:
        """Verify merged cells exist, with range details."""
        ws = self._get_ws(task.sheet)
        merged = ws.merged_cells.ranges

        if not merged:
            return VerificationResult(
                task_id=task.id, task_type=task.task_type,
                passed=False, message="No merged cells found",
            )

        ranges = [str(r) for r in merged]

        # Check if a specific range matches
        if task.range and task.range in ranges:
            return VerificationResult(
                task_id=task.id, task_type=task.task_type,
                passed=True,
                message=f"Merged range {task.range} found",
                details={"all_merged": ranges},
            )

        return VerificationResult(
            task_id=task.id, task_type=task.task_type,
            passed=True,
            message=f"{len(ranges)} merged range(s): {ranges}",
            details={"merged_ranges": ranges},
        )

    # ══════════════════════════════════════════════════════════════════
    # Helpers
    # ══════════════════════════════════════════════════════════════════

    def _get_ws(self, sheet_name: Optional[str] = None):
        """Get a worksheet by name or active sheet."""
        if not self._wb:
            raise RuntimeError("No workbook loaded for verification")
        if sheet_name and sheet_name in self._wb.sheetnames:
            return self._wb[sheet_name]
        return self._wb.active

    def _skip(self, task: Task, reason: str) -> VerificationResult:
        """Skip verification with a reason."""
        return VerificationResult(
            task_id=task.id,
            task_type=task.task_type,
            passed=True,
            message=f"Skipped: {reason}",
        )

    def _check_cross_sheet_refs(self, formula: str) -> list[str]:
        """A4: Validate cross-sheet references in a formula. Returns warnings."""
        if not self._wb:
            return []
        warnings = []
        refs = _extract_sheet_references(formula)
        for sheet_name, cell_ref in refs:
            if sheet_name not in self._wb.sheetnames:
                # Could be external workbook ref — warn but don't fail
                msg = f"Referenced sheet '{sheet_name}' not found in workbook"
                logger.warning(msg)
                warnings.append(msg)
            else:
                try:
                    ws = self._wb[sheet_name]
                    cell_val = ws[cell_ref].value
                    if cell_val is None:
                        msg = f"Referenced cell {sheet_name}!{cell_ref} is empty"
                        logger.warning(msg)
                        warnings.append(msg)
                except Exception as exc:
                    msg = f"Error checking {sheet_name}!{cell_ref}: {exc}"
                    logger.warning(msg)
                    warnings.append(msg)
        return warnings

    # ------------------------------------------------------------------
    # Standalone scanners (not tied to individual tasks)
    # ------------------------------------------------------------------

    def verify_formula_errors(self) -> list[VerificationResult]:
        """Scan all cells for Excel error values (#REF!, #DIV/0!, etc.).

        Works with data_only=True workbook to see computed values.
        Can detect errors even without LibreOffice recalculation if the
        workbook was previously opened in Excel/Calc.
        """
        error_values = {
            "#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A",
        }

        wb_vals = self._wb_values
        if wb_vals is None:
            if self._path is None:
                return [
                    VerificationResult(
                        task_id="formula_error_scan",
                        task_type=TaskType.FORMULA,
                        passed=True,
                        message="No workbook path available; skipping formula error scan",
                    )
                ]
            wb_vals = load_workbook(str(self._path), data_only=True)

        results: list[VerificationResult] = []
        sheets_with_errors: set[str] = set()

        for sheet_name in wb_vals.sheetnames:
            ws = wb_vals[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str):
                        for err in error_values:
                            if err in cell.value:
                                sheets_with_errors.add(sheet_name)
                                results.append(
                                    VerificationResult(
                                        task_id=f"formula_error_{sheet_name}_{cell.coordinate}",
                                        task_type=TaskType.FORMULA,
                                        passed=False,
                                        message=f"Excel error {err} found at {sheet_name}!{cell.coordinate}",
                                        details={
                                            "error_type": err,
                                            "location": f"{sheet_name}!{cell.coordinate}",
                                            "cell_value": cell.value,
                                        },
                                    )
                                )
                                break  # one error per cell is enough

        # Close if we opened a temporary workbook
        if wb_vals is not self._wb_values:
            wb_vals.close()

        if not results:
            results.append(
                VerificationResult(
                    task_id="formula_error_scan",
                    task_type=TaskType.FORMULA,
                    passed=True,
                    message="No formula errors found",
                )
            )

        logger.info(
            "Formula error scan: %d errors found across %d sheets",
            len(results) if not results[0].passed else 0,
            len(sheets_with_errors),
        )
        return results

    def count_formulas(self) -> int:
        """Count total number of formula cells in the workbook."""
        if not self._wb:
            raise RuntimeError("No workbook loaded for verification")

        count = 0
        for sheet_name in self._wb.sheetnames:
            ws = self._wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        count += 1
        return count

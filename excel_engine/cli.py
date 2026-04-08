"""
CLI — Standalone command-line interface for the Excel Engine.

Usage:
    excel-engine run <workbook> <instructions>
    excel-engine parse <instructions>
    excel-engine verify <workbook>
    excel-engine info
    excel-engine check-env
    excel-engine interactive
"""

from __future__ import annotations

import argparse
import json
import logging
import shutil
import subprocess
import sys
from dataclasses import asdict
from pathlib import Path
from typing import Optional

import excel_engine
from excel_engine.config import EngineConfig, Layer, TASK_LAYER_MAP


def _setup_logging(verbose: bool) -> None:
    level = logging.DEBUG if verbose else logging.INFO
    fmt = "%(levelname)s: %(message)s" if not verbose else (
        "%(asctime)s %(name)s %(levelname)s: %(message)s"
    )
    logging.basicConfig(level=level, format=fmt, stream=sys.stderr)


def _load_config(config_path: Optional[str]) -> EngineConfig:
    """Load EngineConfig from a JSON file or return defaults."""
    if not config_path:
        return EngineConfig()

    path = Path(config_path)
    if not path.exists():
        print(f"Error: config file not found: {path}", file=sys.stderr)
        sys.exit(1)

    data = json.loads(path.read_text())
    return EngineConfig(**{
        k: v for k, v in data.items()
        if k in EngineConfig.__dataclass_fields__
    })


def _write_output(data: dict, output_path: Optional[str]) -> None:
    """Write result data as JSON to a file or stdout."""
    text = json.dumps(data, indent=2, default=str)
    if output_path:
        Path(output_path).write_text(text + "\n")
        print(f"Results written to {output_path}")
    else:
        print(text)


# ── Commands ──


def cmd_run(args: argparse.Namespace) -> int:
    """Full assignment completion: parse → plan → execute → verify."""
    from excel_engine.engine import ExcelEngine, EngineResult

    workbook = Path(args.workbook).resolve()
    instructions = Path(args.instructions).resolve()

    if not workbook.exists():
        print(f"Error: workbook not found: {workbook}", file=sys.stderr)
        return 1
    if not instructions.exists():
        print(f"Error: instructions not found: {instructions}", file=sys.stderr)
        return 1

    config = _load_config(args.config)
    engine = ExcelEngine(config=config)

    print(f"── Excel Engine v{excel_engine.__version__} ──")
    print(f"Workbook:     {workbook.name}")
    print(f"Instructions: {instructions.name}")
    print()

    if args.dry_run:
        # Parse and plan only
        text = engine.parser.parse(instructions)
        tasks = engine.extractor.extract(text)
        plan = engine.planner.plan(tasks)
        print("DRY RUN — execution plan:")
        print(plan.summary())
        print(f"\nEstimated time: {plan.estimated_time_seconds:.0f}s")

        if args.output:
            _write_output({
                "mode": "dry_run",
                "sections": plan.section_count,
                "tasks": plan.total_tasks,
                "estimated_seconds": plan.estimated_time_seconds,
                "plan_summary": plan.summary(),
            }, args.output)
        return 0

    result = engine.run(workbook=workbook, instructions=instructions)
    print(result.summary())

    if args.output:
        _write_output({
            "success": result.success,
            "workbook": str(result.workbook_path),
            "sections_completed": result.sections_completed,
            "sections_total": result.sections_total,
            "tasks_completed": result.tasks_completed,
            "tasks_total": result.tasks_total,
            "elapsed_seconds": result.elapsed_seconds,
            "errors": result.errors,
        }, args.output)

    return 0 if result.success else 1


def cmd_parse(args: argparse.Namespace) -> int:
    """Parse instructions into structured JSON tasks."""
    from excel_engine.parsers.instruction_parser import InstructionParser
    from excel_engine.parsers.task_extractor import TaskExtractor

    instructions = Path(args.instructions).resolve()
    if not instructions.exists():
        print(f"Error: instructions not found: {instructions}", file=sys.stderr)
        return 1

    parser = InstructionParser()
    extractor = TaskExtractor()

    text = parser.parse(instructions)
    tasks = extractor.extract(text)

    task_dicts = []
    for t in tasks:
        d = {
            "id": t.id,
            "task_type": t.task_type.value,
            "description": t.description,
            "sheet": t.sheet,
            "cell": t.cell,
            "range": t.range,
            "value": t.value,
            "formula": t.formula,
            "style": t.style,
            "params": t.params,
            "depends_on": t.depends_on,
        }
        task_dicts.append({k: v for k, v in d.items() if v})

    print(f"Parsed {len(tasks)} tasks from {instructions.name}")

    output_data = {
        "source": str(instructions),
        "task_count": len(tasks),
        "tasks": task_dicts,
    }

    if args.output:
        _write_output(output_data, args.output)
    else:
        print(json.dumps(output_data, indent=2, default=str))

    return 0


def cmd_verify(args: argparse.Namespace) -> int:
    """Verify workbook completion status."""
    from excel_engine.verifier.workbook_verifier import WorkbookVerifier

    workbook = Path(args.workbook).resolve()
    if not workbook.exists():
        print(f"Error: workbook not found: {workbook}", file=sys.stderr)
        return 1

    verifier = WorkbookVerifier()
    print(f"Verifying: {workbook.name}")

    # If instructions provided, do a full task-based verification
    if hasattr(args, "instructions") and args.instructions:
        from excel_engine.parsers.instruction_parser import InstructionParser
        from excel_engine.parsers.task_extractor import TaskExtractor
        from excel_engine.planner.task_planner import TaskPlanner

        inst_path = Path(args.instructions).resolve()
        parser = InstructionParser()
        extractor = TaskExtractor()

        text = parser.parse(inst_path)
        tasks = extractor.extract(text)
        planner = TaskPlanner(config=_load_config(args.config))
        plan = planner.plan(tasks)

        verifier.load(workbook)
        results = []
        for section in plan.sections:
            sv = verifier.verify_section(section.id, section.tasks)
            results.append(sv)
            print(sv.summary())
        verifier.close()

        overall = all(sv.all_passed for sv in results)
        print(f"\nOverall: {'✓ ALL PASSED' if overall else '✗ SOME FAILED'}")

        if args.output:
            _write_output({
                "workbook": str(workbook),
                "overall_passed": overall,
                "sections": [
                    {
                        "id": sv.section_id,
                        "passed": sv.all_passed,
                        "pass_count": sv.pass_count,
                        "fail_count": sv.fail_count,
                    }
                    for sv in results
                ],
            }, args.output)

        return 0 if overall else 1

    # Basic workbook inspection without instructions
    from openpyxl import load_workbook as openpyxl_load
    wb = openpyxl_load(workbook, data_only=True)
    sheets = wb.sheetnames

    print(f"  Sheets: {len(sheets)} — {', '.join(sheets)}")
    for name in sheets:
        ws = wb[name]
        print(f"  [{name}] {ws.max_row}r × {ws.max_column}c")
    wb.close()
    return 0


def cmd_info(args: argparse.Namespace) -> int:
    """Show engine version, layers, and configuration."""
    config = _load_config(getattr(args, "config", None))

    print(f"Excel Engine v{excel_engine.__version__}")
    print(f"Author: {excel_engine.__author__}")
    print()

    print("Automation Layers:")
    for layer in Layer:
        preferred_tasks = [
            tt.value for tt, layers in TASK_LAYER_MAP.items()
            if layers and layers[0] == layer
        ]
        count = len(preferred_tasks)
        print(f"  {layer.value}. {layer.name:<15} ({count} primary task types)")

    print()
    print("Task Types:")
    from excel_engine.config import TaskType
    for tt in TaskType:
        layers = TASK_LAYER_MAP.get(tt, [])
        chain = " → ".join(l.name for l in layers)
        print(f"  {tt.value:<25} [{chain}]")

    print()
    print("Configuration:")
    print(f"  Scan timeout:       {config.scan_timeout}s")
    print(f"  Section timeout:    {config.section_timeout}s")
    print(f"  Max retries:        {config.max_retries}")
    print(f"  Verify per section: {config.verify_after_each_section}")
    print(f"  SAM fingerprint:    {config.sam_fingerprint_protected}")
    print(f"  Retina display:     {config.retina_display}")

    return 0


def cmd_check_env(args: argparse.Namespace) -> int:
    """Check macOS environment for Excel Engine requirements."""
    issues = []
    print("Excel Engine — Environment Check")
    print("=" * 40)

    # Python version
    v = sys.version_info
    py_ok = v >= (3, 10)
    status = "✓" if py_ok else "✗"
    print(f"  {status} Python {v.major}.{v.minor}.{v.micro}", end="")
    if not py_ok:
        print("  (need ≥ 3.10)")
        issues.append("Python 3.10+ required")
    else:
        print()

    # Platform
    import platform
    is_mac = platform.system() == "Darwin"
    status = "✓" if is_mac else "✗"
    print(f"  {status} Platform: {platform.system()} {platform.machine()}")
    if not is_mac:
        issues.append("macOS required")

    # openpyxl (required)
    try:
        import openpyxl
        print(f"  ✓ openpyxl {openpyxl.__version__}")
    except ImportError:
        print("  ✗ openpyxl — NOT INSTALLED")
        issues.append("openpyxl not installed (pip install openpyxl)")

    # xlwings (optional)
    try:
        import xlwings
        print(f"  ✓ xlwings {xlwings.__version__}")
    except ImportError:
        print("  ○ xlwings — not installed (optional: pip install xlwings)")

    # pyautogui (optional)
    try:
        import pyautogui
        print(f"  ✓ pyautogui {pyautogui.__version__}")
    except ImportError:
        print("  ○ pyautogui — not installed (optional: pip install pyautogui)")

    # python-docx (optional)
    try:
        import docx
        print(f"  ✓ python-docx")
    except ImportError:
        print("  ○ python-docx — not installed (optional: pip install python-docx)")

    # pdfplumber (optional)
    try:
        import pdfplumber
        print(f"  ✓ pdfplumber {pdfplumber.__version__}")
    except ImportError:
        print("  ○ pdfplumber — not installed (optional: pip install pdfplumber)")

    # Microsoft Excel
    print()
    print("Applications:")
    excel_path = Path("/Applications/Microsoft Excel.app")
    if excel_path.exists():
        print("  ✓ Microsoft Excel found")
    else:
        print("  ✗ Microsoft Excel — NOT FOUND")
        issues.append("Microsoft Excel not installed")

    # osascript
    if shutil.which("osascript"):
        print("  ✓ osascript available")
    else:
        print("  ✗ osascript — NOT FOUND")
        issues.append("osascript not available")

    # Accessibility permissions
    print()
    print("Permissions:")
    if is_mac:
        try:
            result = subprocess.run(
                ["osascript", "-e",
                 'tell application "System Events" to return name of first process'],
                capture_output=True, text=True, timeout=5,
            )
            if result.returncode == 0:
                print("  ✓ Accessibility (System Events) — granted")
            else:
                print("  ✗ Accessibility (System Events) — may need permission")
                issues.append("Accessibility permission may be needed")
        except Exception:
            print("  ? Accessibility — could not check")
    else:
        print("  ○ Skipped (not macOS)")

    # Summary
    print()
    if issues:
        print(f"Issues found ({len(issues)}):")
        for issue in issues:
            print(f"  • {issue}")
        return 1
    else:
        print("✓ Environment looks good!")
        return 0


# ── CLI Parser ──


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="excel-engine",
        description=f"Excel Engine v{excel_engine.__version__} — Autonomous Excel automation for macOS",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "Examples:\n"
            "  excel-engine run assignment.xlsx instructions.docx\n"
            "  excel-engine parse instructions.rtfd --output tasks.json\n"
            "  excel-engine verify completed.xlsx\n"
            "  excel-engine info\n"
            "  excel-engine check-env\n"
            "  excel-engine interactive\n"
        ),
    )

    # Global options
    parser.add_argument(
        "-v", "--verbose", action="store_true",
        help="Enable verbose/debug output",
    )
    parser.add_argument(
        "--version", action="version",
        version=f"%(prog)s {excel_engine.__version__}",
    )

    subparsers = parser.add_subparsers(dest="command", help="Available commands")

    # ── run ──
    run_p = subparsers.add_parser("run", help="Complete an Excel assignment")
    run_p.add_argument("workbook", help="Path to the Excel workbook (.xlsx)")
    run_p.add_argument("instructions", help="Path to instruction file (.docx, .rtfd, .pdf, .txt)")
    run_p.add_argument("--dry-run", action="store_true", help="Plan only, don't execute")
    run_p.add_argument(
        "--phase", choices=["1", "2", "both"], default="both",
        help="Run specific phase: 1=openpyxl, 2=xlwings, both=full pipeline",
    )
    run_p.add_argument("--config", help="Path to JSON config file")
    run_p.add_argument("--output", "-o", help="Write results to JSON file")

    # ── parse ──
    parse_p = subparsers.add_parser("parse", help="Parse instructions to structured JSON")
    parse_p.add_argument("instructions", help="Path to instruction file")
    parse_p.add_argument("--output", "-o", help="Write parsed tasks to JSON file")
    parse_p.add_argument("--config", help="Path to JSON config file")

    # ── verify ──
    verify_p = subparsers.add_parser("verify", help="Verify workbook completion")
    verify_p.add_argument("workbook", help="Path to the Excel workbook (.xlsx)")
    verify_p.add_argument("--instructions", "-i", help="Optional instruction file for task-based verification")
    verify_p.add_argument("--output", "-o", help="Write verification results to JSON file")
    verify_p.add_argument("--config", help="Path to JSON config file")

    # ── info ──
    info_p = subparsers.add_parser("info", help="Show engine version, layers, and config")
    info_p.add_argument("--config", help="Path to JSON config file")

    # ── check-env ──
    subparsers.add_parser("check-env", help="Check macOS environment readiness")

    # ── interactive ──
    subparsers.add_parser("interactive", help="Interactive guided mode")

    return parser


def main(argv: Optional[list[str]] = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)

    if not args.command:
        parser.print_help()
        return 0

    _setup_logging(args.verbose)

    dispatch = {
        "run": cmd_run,
        "parse": cmd_parse,
        "verify": cmd_verify,
        "info": cmd_info,
        "check-env": cmd_check_env,
        "interactive": _cmd_interactive,
    }

    handler = dispatch.get(args.command)
    if handler is None:
        parser.print_help()
        return 1

    try:
        return handler(args)
    except KeyboardInterrupt:
        print("\nAborted.", file=sys.stderr)
        return 130
    except Exception as exc:
        logging.getLogger(__name__).debug("Unhandled exception", exc_info=True)
        print(f"Error: {exc}", file=sys.stderr)
        return 1


def _cmd_interactive(args: argparse.Namespace) -> int:
    """Launch the interactive mode."""
    from excel_engine.interactive import interactive_session
    return interactive_session()


if __name__ == "__main__":
    sys.exit(main())

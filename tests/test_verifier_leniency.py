"""
Tests documenting verifier leniency — known edge cases.

The WorkbookVerifier has a known issue: when a CELL_VALUE task has
value=None, the cell is verified as "passed" as long as it has *any*
value, without checking against an expected value.
"""

import pytest
from openpyxl import Workbook

from excel_engine.config import TaskType
from excel_engine.parsers.task_extractor import Task
from excel_engine.verifier.workbook_verifier import WorkbookVerifier


def _task(task_type, **kw):
    kw.setdefault("id", f"len-{task_type.value}")
    kw.setdefault("description", f"Leniency {task_type.value}")
    return Task(task_type=task_type, **kw)


@pytest.fixture
def verifier_with_data(tmp_path):
    """Create a workbook with known values and return a loaded verifier."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "expected"
    ws["B1"] = 42
    ws["C1"] = None  # empty cell

    path = tmp_path / "leniency.xlsx"
    wb.save(path)
    wb.close()

    v = WorkbookVerifier()
    v.load(path)
    yield v
    v.close()


class TestVerifierLeniency:
    def test_value_none_passes_with_any_content(self, verifier_with_data):
        """Known issue: task.value=None → verifier says 'passed' for any non-empty cell."""
        task = _task(TaskType.CELL_VALUE, cell="A1", value=None, sheet="Sheet1")
        result = verifier_with_data.verify_task(task)
        # This documents the leniency: it passes even though no expected value was given
        assert result.passed is True
        assert "cannot verify" in result.message.lower() or "has value" in result.message.lower() or "skip" in result.message.lower()

    def test_value_match_passes(self, verifier_with_data):
        """Correct expected value → verification passes."""
        task = _task(TaskType.CELL_VALUE, cell="A1", value="expected", sheet="Sheet1")
        result = verifier_with_data.verify_task(task)
        assert result.passed is True
        assert "match" in result.message.lower() or "expected" in result.message.lower()

    def test_value_mismatch_fails(self, verifier_with_data):
        """Wrong expected value → verification fails."""
        task = _task(TaskType.CELL_VALUE, cell="A1", value="different", sheet="Sheet1")
        result = verifier_with_data.verify_task(task)
        assert result.passed is False
        assert "mismatch" in result.message.lower()

    def test_empty_cell_fails_even_with_none_value(self, verifier_with_data):
        """An actually empty cell fails verification even with value=None."""
        task = _task(TaskType.CELL_VALUE, cell="C1", value=None, sheet="Sheet1")
        result = verifier_with_data.verify_task(task)
        assert result.passed is False
        assert "empty" in result.message.lower()

    def test_numeric_value_match(self, verifier_with_data):
        """Numeric cell matches when string representation equals."""
        task = _task(TaskType.CELL_VALUE, cell="B1", value="42", sheet="Sheet1")
        result = verifier_with_data.verify_task(task)
        assert result.passed is True

    def test_numeric_value_mismatch(self, verifier_with_data):
        """Numeric cell fails when expected value differs."""
        task = _task(TaskType.CELL_VALUE, cell="B1", value="999", sheet="Sheet1")
        result = verifier_with_data.verify_task(task)
        assert result.passed is False

    def test_no_workbook_loaded_skips(self):
        """Verifier with no workbook loaded skips verification."""
        v = WorkbookVerifier()
        task = _task(TaskType.CELL_VALUE, cell="A1", value="expected")
        result = v.verify_task(task)
        # The _skip method returns passed=True with a "Skipped" message
        assert result.passed is True
        assert "skip" in result.message.lower()

    def test_no_cell_reference_skips(self, verifier_with_data):
        """Task with no cell reference skips verification."""
        task = _task(TaskType.CELL_VALUE, cell=None, value="expected", sheet="Sheet1")
        result = verifier_with_data.verify_task(task)
        assert result.passed is True
        assert "skip" in result.message.lower()

    def test_formula_none_value_lenient(self, verifier_with_data):
        """Formula verifier is also lenient when cell has any value."""
        # Put a formula-like value in cell A1 (already has "expected", not a formula)
        task = _task(TaskType.FORMULA, cell="A1", formula=None, sheet="Sheet1")
        result = verifier_with_data.verify_task(task)
        # Cell has a value but it's not a formula — verifier still says passed
        # because cell_value is not None
        assert result.passed is True
        assert "has value" in result.message.lower()
